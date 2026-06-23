# -*- coding: utf-8 -*-
"""
Export Excel configurable du listing de stock pour l'inventaire.
Supporte le regroupement par : rayon, forme, groupe, fournisseur.
Supporte les filtres de stock : tous, nuls (=0), non nuls (>0).
"""
import io
from decimal import Decimal

from django.http import HttpResponse

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from api.models import Produit, StockLot, PharmacySettings


# ---------------------------------------------------------------------------
# Helpers style
# ---------------------------------------------------------------------------

def _make_border(style='thin'):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill(hex_color='1F4E79'):
    return PatternFill(fill_type='solid', fgColor=hex_color)


def _group_fill(hex_color='BDD7EE'):
    return PatternFill(fill_type='solid', fgColor=hex_color)


def _subtotal_fill(hex_color='D9E1F2'):
    return PatternFill(fill_type='solid', fgColor=hex_color)


# ---------------------------------------------------------------------------
# Fonction principale
# ---------------------------------------------------------------------------

def generate_listing_excel(
    group_by: str = 'rayon',
    stock_filter: str = 'tous',
    filter_id: int = None,
    inventaire_id: int = None,
):
    """
    Génère un fichier Excel du listing de stock courant (Produit.stock).

    Paramètres
    ----------
    group_by    : 'rayon' | 'forme' | 'groupe' | 'fournisseur'
    stock_filter: 'tous' | 'zero' | 'non_zero'
    filter_id   : id de l'entité de regroupement pour filtrer (optionnel)
    inventaire_id: si fourni, liste les lignes d'un inventaire précis (optionnel)
    """
    if not HAS_OPENPYXL:
        return HttpResponse("openpyxl non installé", status=500)

    # ------------------------------------------------------------------
    # 1. Récupération des données
    # ------------------------------------------------------------------
    if inventaire_id:
        rows = _get_rows_from_inventaire(inventaire_id, group_by, stock_filter, filter_id)
        listing_type = 'inventaire'
    else:
        rows = _get_rows_from_stock(group_by, stock_filter, filter_id)
        listing_type = 'stock'

    # ------------------------------------------------------------------
    # 2. En-tête pharmacie
    # ------------------------------------------------------------------
    try:
        pharmacy = PharmacySettings.objects.get(pk=1)
        pharma_name = pharmacy.pharmacy_name or 'PHARMACIE'
        pharma_address = f"{pharmacy.address or ''} {pharmacy.city or ''}".strip()
        pharma_phone = pharmacy.phone or ''
    except Exception:
        pharma_name = 'PHARMACIE'
        pharma_address = ''
        pharma_phone = ''

    from django.utils import timezone as tz
    now_str = tz.localtime(tz.now()).strftime("%d/%m/%Y à %H:%M")

    group_labels = {
        'rayon': 'Rayon',
        'forme': 'Forme galénique',
        'groupe': 'Groupe thérapeutique',
        'fournisseur': 'Fournisseur',
    }
    group_label = group_labels.get(group_by, group_by.capitalize())

    stock_filter_labels = {
        'tous': 'Tous les stocks',
        'zero': 'Stocks nuls (Qté = 0)',
        'non_zero': 'Stocks non nuls (Qté > 0)',
    }
    stock_label = stock_filter_labels.get(stock_filter, stock_filter)

    # ------------------------------------------------------------------
    # 3. Construction du classeur
    # ------------------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Listing Stock (Lots)" if listing_type == 'stock' else "Listing Inventaire"

    thin_border = _make_border('thin')
    medium_border = _make_border('medium')

    # Fonts
    font_title = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    font_pharma = Font(name='Calibri', bold=True, size=11)
    font_info = Font(name='Calibri', size=9, italic=True)
    font_group = Font(name='Calibri', bold=True, size=10, color='1F4E79')
    font_header = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
    font_data = Font(name='Calibri', size=9)
    font_subtotal = Font(name='Calibri', bold=True, size=9)

    # Colonnes : définition selon listing_type
    if listing_type == 'inventaire':
        columns = [
            ('CIP', 14),
            ('Désignation', 38),
            ('N° Lot', 14),
            ('Exp. Lot', 12),
            ('Stock Théo.', 12),
            ('Qté Comptée', 12),
            ('Écart', 10),
            ('PMP', 12),
            ('Val. Écart', 14),
        ]
    else:
        columns = [
            ('CIP', 14),
            ('Désignation', 38),
            ('Forme', 16),
            ('Rayon', 14),
            ('N° Lot', 14),
            ('Exp. Lot', 12),
            ('Stock Lot', 10),
            ('Stock Rés.', 10),
            ('PMP', 12),
            ('Val. Stock', 14),
            ('Prix Vente', 12),
        ]

    nb_cols = len(columns)

    # Appliquer largeurs
    for i, (_, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ------------------------------------------------------------------
    # 4. Bloc en-tête pharmacie
    # ------------------------------------------------------------------
    row = 1

    # Ligne titre principale
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nb_cols)
    cell = ws.cell(row=row, column=1, value="LISTING D'INVENTAIRE DE STOCK")
    cell.font = font_title
    cell.fill = _header_fill('1F4E79')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 22
    row += 1

    # Nom pharmacie
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nb_cols)
    cell = ws.cell(row=row, column=1, value=pharma_name)
    cell.font = font_pharma
    cell.alignment = Alignment(horizontal='center')
    row += 1

    if pharma_address:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nb_cols)
        ws.cell(row=row, column=1, value=pharma_address).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=1).font = font_info
        row += 1

    if pharma_phone:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nb_cols)
        ws.cell(row=row, column=1, value=f"Tél : {pharma_phone}").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=1).font = font_info
        row += 1

    # Infos du listing
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nb_cols)
    source_label = 'Stock courant (par lots)' if listing_type == 'stock' else 'Inventaire'
    info_text = f"Édité le : {now_str}  |  Source : {source_label}  |  Regroupement : {group_label}  |  Filtre : {stock_label}"
    cell = ws.cell(row=row, column=1, value=info_text)
    cell.font = font_info
    cell.fill = PatternFill(fill_type='solid', fgColor='F2F2F2')
    cell.alignment = Alignment(horizontal='center')
    row += 1

    # Ligne vide séparateur
    row += 1

    # ------------------------------------------------------------------
    # 5. En-tête du tableau
    # ------------------------------------------------------------------
    header_row = row
    for col_idx, (col_name, _) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = font_header
        cell.fill = _header_fill('2E75B6')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 18
    row += 1

    # ------------------------------------------------------------------
    # 6. Données groupées
    # ------------------------------------------------------------------
    data_start_row = row
    grand_total_stock = 0
    grand_total_valeur = 0
    grand_total_lines = 0

    for group_name, group_rows in rows.items():
        if not group_rows:
            continue

        # En-tête du groupe
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nb_cols)
        cell = ws.cell(row=row, column=1, value=f"  ▶  {group_name.upper()}")
        cell.font = font_group
        cell.fill = _group_fill('BDD7EE')
        cell.alignment = Alignment(vertical='center')
        for col_idx in range(1, nb_cols + 1):
            ws.cell(row=row, column=col_idx).fill = _group_fill('BDD7EE')
            ws.cell(row=row, column=col_idx).border = thin_border
        ws.row_dimensions[row].height = 16
        row += 1

        group_total_stock = 0
        group_total_valeur = 0

        for r in group_rows:
            if listing_type == 'inventaire':
                vals = [
                    r.get('cip', ''),
                    r.get('name', ''),
                    r.get('lot_numero', ''),
                    r.get('lot_expiration', ''),
                    r.get('stock_theorique', 0),
                    r.get('quantite_physique', 0),
                    r.get('ecart', 0),
                    r.get('pmp', 0),
                    r.get('valeur_ecart', 0),
                ]
                group_total_stock += r.get('quantite_physique', 0)
                group_total_valeur += r.get('valeur_ecart', 0)
            else:
                vals = [
                    r.get('cip', ''),
                    r.get('name', ''),
                    r.get('forme', ''),
                    r.get('rayon', ''),
                    r.get('lot_numero', ''),
                    r.get('lot_expiration', ''),
                    r.get('stock', 0),
                    r.get('stock_reserve', 0),
                    r.get('pmp', 0),
                    r.get('valeur_stock', 0),
                    r.get('prix_vente', 0),
                ]
                group_total_stock += r.get('stock', 0)
                group_total_valeur += r.get('valeur_stock', 0)

            # Colonnes monétaires selon le mode
            money_cols = (9, 10, 11) if listing_type == 'stock' else (8, 9)
            for col_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = font_data
                cell.border = thin_border
                if isinstance(val, (int, float, Decimal)):
                    cell.alignment = Alignment(horizontal='right')
                    if col_idx in money_cols:
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = Alignment(horizontal='left')
            row += 1

        grand_total_stock += group_total_stock
        grand_total_valeur += group_total_valeur
        grand_total_lines += len(group_rows)

        # Sous-total groupe
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        subtotal_label = f"  Sous-total {group_name}  ({len(group_rows)} article(s))"
        cell = ws.cell(row=row, column=1, value=subtotal_label)
        cell.font = font_subtotal
        cell.fill = _subtotal_fill()
        cell.alignment = Alignment(horizontal='right')

        if listing_type == 'inventaire':
            stock_col = 6
            val_col = 9
        else:
            stock_col = 7
            val_col = 10

        for col_idx in range(1, nb_cols + 1):
            c = ws.cell(row=row, column=col_idx)
            c.fill = _subtotal_fill()
            c.border = thin_border
            c.font = font_subtotal

        ws.cell(row=row, column=stock_col, value=group_total_stock).number_format = '#,##0'
        ws.cell(row=row, column=stock_col).font = font_subtotal
        ws.cell(row=row, column=stock_col).fill = _subtotal_fill()
        ws.cell(row=row, column=stock_col).alignment = Alignment(horizontal='right')

        ws.cell(row=row, column=val_col, value=group_total_valeur).number_format = '#,##0'
        ws.cell(row=row, column=val_col).font = font_subtotal
        ws.cell(row=row, column=val_col).fill = _subtotal_fill()
        ws.cell(row=row, column=val_col).alignment = Alignment(horizontal='right')

        row += 1
        # Ligne vide entre groupes
        row += 1

    # ------------------------------------------------------------------
    # 7. TOTAL GÉNÉRAL
    # ------------------------------------------------------------------
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value=f"  TOTAL GÉNÉRAL  ({grand_total_lines} article(s))")
    cell.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    cell.fill = _header_fill('1F4E79')
    cell.alignment = Alignment(horizontal='right', vertical='center')

    for col_idx in range(1, nb_cols + 1):
        c = ws.cell(row=row, column=col_idx)
        c.fill = _header_fill('1F4E79')
        c.border = medium_border
        c.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')

    if listing_type == 'inventaire':
        stock_col, val_col = 6, 9
    else:
        stock_col, val_col = 7, 10

    c_stock = ws.cell(row=row, column=stock_col, value=grand_total_stock)
    c_stock.number_format = '#,##0'
    c_stock.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    c_stock.fill = _header_fill('1F4E79')
    c_stock.alignment = Alignment(horizontal='right')

    c_val = ws.cell(row=row, column=val_col, value=grand_total_valeur)
    c_val.number_format = '#,##0'
    c_val.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    c_val.fill = _header_fill('1F4E79')
    c_val.alignment = Alignment(horizontal='right')

    ws.row_dimensions[row].height = 20

    # ------------------------------------------------------------------
    # 8. Figer la ligne d'en-tête tableau
    # ------------------------------------------------------------------
    ws.freeze_panes = ws.cell(row=data_start_row, column=1)

    # ------------------------------------------------------------------
    # 9. Réponse HTTP
    # ------------------------------------------------------------------
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from django.utils import timezone as tz
    filename = f"listing_inventaire_{group_by}_{tz.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# Source de données : Stock courant (Produit)
# ---------------------------------------------------------------------------

def _get_rows_from_stock(group_by: str, stock_filter: str, filter_id=None):
    """
    Construit le dict {group_name: [rows]} à partir des lots de stock (StockLot).
    Une ligne par lot, avec N° lot, date expiration, quantité restante et PMP.
    """
    qs = StockLot.objects.filter(
        produit__isnull=False,
        produit__is_active=True,
    ).select_related(
        'produit', 'produit__rayon', 'produit__forme',
        'produit__groupe', 'produit__fournisseur',
    )

    # Filtre stock (sur quantity_remaining du lot)
    if stock_filter == 'zero':
        qs = qs.filter(quantity_remaining__lt=0)
    elif stock_filter == 'non_zero':
        qs = qs.filter(quantity_remaining__gt=0)

    # Filtre entité
    if filter_id:
        if group_by == 'rayon':
            qs = qs.filter(produit__rayon_id=filter_id)
        elif group_by == 'forme':
            qs = qs.filter(produit__forme_id=filter_id)
        elif group_by == 'groupe':
            qs = qs.filter(produit__groupe_id=filter_id)
        elif group_by == 'fournisseur':
            qs = qs.filter(produit__fournisseur_id=filter_id)

    # Tri : par groupe puis produit puis date de réception (FIFO)
    sort_map = {
        'rayon': ('produit__rayon__name', 'produit__name', 'date_reception'),
        'forme': ('produit__forme__nom', 'produit__name', 'date_reception'),
        'groupe': ('produit__groupe__nom', 'produit__name', 'date_reception'),
        'fournisseur': ('produit__fournisseur__name', 'produit__name', 'date_reception'),
    }
    qs = qs.order_by(*sort_map.get(group_by, ('produit__name', 'date_reception')))

    grouped = {}
    seen_produit_ids = set()

    for lot in qs:
        p = lot.produit
        seen_produit_ids.add(p.id)
        group_name = _get_group_name(p, group_by)
        if group_name not in grouped:
            grouped[group_name] = []

        stock_val = int(lot.quantity_remaining or 0)
        stock_reserve = int(lot.quantity_reserved or 0)
        pmp = float(lot.price_cost or p.pmp or p.cost_price or 0)
        valeur_stock = stock_val * pmp

        lot_expiration = ''
        if lot.date_expiration:
            lot_expiration = lot.date_expiration.strftime('%d/%m/%Y')

        grouped[group_name].append({
            'cip': p.cip1 or '',
            'name': p.name,
            'forme': p.forme.nom if p.forme else '',
            'rayon': p.rayon.name if p.rayon else '',
            'fournisseur': lot.fournisseur_nom or (p.fournisseur.name if p.fournisseur else ''),
            'lot_numero': lot.lot or '',
            'lot_expiration': lot_expiration,
            'stock': stock_val,
            'stock_reserve': stock_reserve,
            'pmp': round(pmp, 2),
            'valeur_stock': round(valeur_stock, 0),
            'prix_vente': float(lot.selling_price or p.selling_price or 0),
        })

    # En mode "tous" : inclure aussi les produits actifs sans aucun lot (stock nul implicite)
    if stock_filter == 'tous':
        prod_qs = Produit.objects.filter(
            is_active=True,
        ).exclude(
            id__in=seen_produit_ids,
        ).select_related('rayon', 'forme', 'groupe', 'fournisseur')

        if filter_id:
            if group_by == 'rayon':
                prod_qs = prod_qs.filter(rayon_id=filter_id)
            elif group_by == 'forme':
                prod_qs = prod_qs.filter(forme_id=filter_id)
            elif group_by == 'groupe':
                prod_qs = prod_qs.filter(groupe_id=filter_id)
            elif group_by == 'fournisseur':
                prod_qs = prod_qs.filter(fournisseur_id=filter_id)

        for p in prod_qs.order_by('name'):
            group_name = _get_group_name(p, group_by)
            if group_name not in grouped:
                grouped[group_name] = []
            pmp = float(p.pmp or p.cost_price or 0)
            grouped[group_name].append({
                'cip': p.cip1 or '',
                'name': p.name,
                'forme': p.forme.nom if p.forme else '',
                'rayon': p.rayon.name if p.rayon else '',
                'fournisseur': p.fournisseur.name if p.fournisseur else '',
                'lot_numero': '',
                'lot_expiration': '',
                'stock': 0,
                'stock_reserve': 0,
                'pmp': round(pmp, 2),
                'valeur_stock': 0.0,
                'prix_vente': float(p.selling_price or 0),
            })

    # Tri final : toutes les lignes de chaque groupe par désignation alphabétique
    for group_name in grouped:
        grouped[group_name].sort(key=lambda r: r['name'].lower())

    return grouped


# ---------------------------------------------------------------------------
# Source de données : Lignes d'un inventaire
# ---------------------------------------------------------------------------

def _get_rows_from_inventaire(inventaire_id: int, group_by: str, stock_filter: str, filter_id=None):
    """
    Construit le dict {group_name: [rows]} à partir des lignes d'un inventaire.
    """
    from api.models import LigneInventaire

    qs = LigneInventaire.objects.filter(inventaire_id=inventaire_id).select_related(
        'produit', 'produit__rayon', 'produit__forme', 'produit__groupe',
        'produit__fournisseur', 'stock_lot'
    )

    # Filtre stock (sur quantite_physique)
    if stock_filter == 'zero':
        qs = qs.filter(quantite_physique__lt=0)
    elif stock_filter == 'non_zero':
        qs = qs.filter(quantite_physique__gt=0)

    # Filtre entité
    if filter_id:
        if group_by == 'rayon':
            qs = qs.filter(produit__rayon_id=filter_id)
        elif group_by == 'forme':
            qs = qs.filter(produit__forme_id=filter_id)
        elif group_by == 'groupe':
            qs = qs.filter(produit__groupe_id=filter_id)
        elif group_by == 'fournisseur':
            qs = qs.filter(produit__fournisseur_id=filter_id)

    # Tri
    sort_map = {
        'rayon': ('produit__rayon__name', 'produit__name'),
        'forme': ('produit__forme__nom', 'produit__name'),
        'groupe': ('produit__groupe__nom', 'produit__name'),
        'fournisseur': ('produit__fournisseur__name', 'produit__name'),
    }
    qs = qs.order_by(*sort_map.get(group_by, ('produit__name',)))

    grouped = {}
    for ligne in qs:
        p = ligne.produit
        if not p:
            continue

        group_name = _get_group_name(p, group_by)
        if group_name not in grouped:
            grouped[group_name] = []

        pmp = float(ligne.pmp_snapshot or p.pmp or p.cost_price or 0)
        valeur_ecart = float(ligne.ecart) * pmp

        lot_numero = ''
        lot_expiration = ''
        if ligne.stock_lot:
            lot_numero = ligne.stock_lot.lot or ''
            if ligne.stock_lot.date_expiration:
                lot_expiration = ligne.stock_lot.date_expiration.strftime('%d/%m/%Y')

        grouped[group_name].append({
            'cip': p.cip1 or '',
            'name': p.name,
            'lot_numero': lot_numero,
            'lot_expiration': lot_expiration,
            'stock_theorique': float(ligne.stock_theorique),
            'quantite_physique': float(ligne.quantite_physique),
            'ecart': float(ligne.ecart),
            'pmp': round(pmp, 2),
            'valeur_ecart': round(valeur_ecart, 0),
        })

    return grouped


# ---------------------------------------------------------------------------
# Helper : nom du groupe
# ---------------------------------------------------------------------------

def _get_group_name(produit, group_by: str) -> str:
    if group_by == 'rayon':
        return produit.rayon.name if produit.rayon else 'SANS RAYON'
    elif group_by == 'forme':
        return produit.forme.nom if produit.forme else 'SANS FORME'
    elif group_by == 'groupe':
        return produit.groupe.nom if produit.groupe else 'SANS GROUPE'
    elif group_by == 'fournisseur':
        return produit.fournisseur.name if produit.fournisseur else 'SANS FOURNISSEUR'
    return 'AUTRES'
