from django.db import transaction
from django.db.models import F
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from ...models import (
    AvoirClient,
    DepotClient,
    Facture,
    MouvementCaisse,
    MouvementStock,
    Produit,
    StockLot,
)
from ...pagination import StandardResultsSetPagination
from ...serializers import AvoirClientSerializer
from ...services.lot_allocation_service import LotAllocationService
from ...sudo_utils import validate_sudo_mode


class AvoirClientViewSet(viewsets.ModelViewSet):
    queryset = AvoirClient.objects.select_related(
        'facture_origine', 'client', 'created_by'
    ).prefetch_related('lignes__produit', 'lignes__stock_lot')
    serializer_class = AvoirClientSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['statut', 'client', 'facture_origine', 'type_motif']

    def get_queryset(self):
        queryset = super().get_queryset()
        date_debut = self.request.query_params.get('date_debut')
        date_fin = self.request.query_params.get('date_fin')
        if date_debut:
            queryset = queryset.filter(date__gte=date_debut)
        if date_fin:
            queryset = queryset.filter(date__lte=date_fin)
        return queryset
        _validation_user, error_response = validate_sudo_mode(
            request, permission_attr='can_create_client_credit'
        )
        if error_response:
            return error_response
        return super().create(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    @transaction.atomic
    def valider(self, request, pk=None):
        validation_user, error_response = validate_sudo_mode(
            request, permission_attr='can_create_client_credit'
        )
        if error_response:
            return error_response

        refund_method = request.data.get('refund_method')
        if refund_method not in ('cash', 'credit'):
            return Response(
                {'detail': "Le mode de remboursement doit être 'cash' ou 'credit'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        avoir = AvoirClient.objects.select_for_update().get(pk=self.get_object().pk)
        if avoir.statut != AvoirClient.Statut.BROUILLON:
            return Response(
                {'detail': "Cet avoir n'est plus à l'état brouillon."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        lignes = list(avoir.lignes.select_related('produit', 'stock_lot'))
        if not lignes:
            return Response({'detail': "L'avoir ne contient aucune ligne."}, status=status.HTTP_400_BAD_REQUEST)
        if refund_method == 'credit' and not avoir.client_id:
            return Response(
                {'detail': 'Un client est requis pour un remboursement en crédit.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        product_ids_with_lots = set()
        for ligne in lignes:
            produit = Produit.objects.select_for_update().get(pk=ligne.produit_id)
            if produit.use_lot_management:
                if not ligne.stock_lot_id:
                    return Response(
                        {'detail': f'Le produit {produit.name} nécessite un lot de réintégration.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                try:
                    stock_lot = StockLot.objects.select_for_update().get(pk=ligne.stock_lot_id)
                except StockLot.DoesNotExist:
                    return Response(
                        {'detail': f'Le lot de la ligne {ligne.id} est introuvable.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                if stock_lot.produit_id != produit.pk:
                    return Response(
                        {'detail': f'Le lot de la ligne {ligne.id} ne correspond pas au produit.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                if stock_lot.date_expiration and stock_lot.date_expiration < timezone.now().date():
                    return Response(
                        {'detail': (
                            f'Le lot {stock_lot.lot} est périmé '
                            f'(expire le {stock_lot.date_expiration.strftime("%d/%m/%Y")}). '
                            'Réintégration impossible — mettre le produit au rebut.'
                        )},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                LotAllocationService.restore_to_lot(stock_lot, ligne.quantity)
                produit.stock = (produit.stock or 0) + ligne.quantity
                product_ids_with_lots.add(produit.pk)
            else:
                Produit.objects.filter(pk=produit.pk).update(stock=F('stock') + ligne.quantity)
                produit.refresh_from_db()

            MouvementStock.objects.create(
                produit=produit,
                facture=avoir.facture_origine,
                avoir_client=avoir,
                type_mouvement=MouvementStock.TypeMouvement.RETOUR,
                quantite=ligne.quantity,
                stock_apres=produit.stock,
                user=validation_user,
                description=f'Réintégration stock - Avoir client {avoir.numero}',
            )

        if product_ids_with_lots:
            LotAllocationService.sync_stock_from_lots(product_ids_with_lots)

        if refund_method == 'cash':
            MouvementCaisse.objects.create(
                type='SORTIE',
                montant=avoir.montant_total,
                motif=f'Remboursement avoir client {avoir.numero}',
                description=avoir.notes,
                user=validation_user,
                poste_caisse=avoir.facture_origine.poste_caisse if avoir.facture_origine else None,
            )
        else:
            DepotClient.objects.create(
                client=avoir.client,
                type=DepotClient.Type.DEPOT,
                montant=avoir.montant_total,
                facture=avoir.facture_origine,
                created_by=validation_user,
                notes=f'Crédit issu de l’avoir client {avoir.numero}',
            )

        avoir.statut = AvoirClient.Statut.VALIDEE
        avoir.save(update_fields=['statut'])
        return Response(self.get_serializer(avoir).data)

    @action(detail=False, methods=['get'])
    def from_invoice(self, request):
        facture_id = request.query_params.get('facture_id')
        if not facture_id:
            return Response({'detail': 'Le paramètre facture_id est requis.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            facture = Facture.objects.select_related('client').prefetch_related(
                'produits__produit', 'produits__allocations__stock_lot'
            ).get(pk=facture_id)
        except (Facture.DoesNotExist, ValueError):
            return Response({'detail': 'Facture introuvable.'}, status=status.HTTP_404_NOT_FOUND)

        lignes_data = []
        for ligne in facture.produits.all():
            produit = ligne.produit
            allocations = list(ligne.allocations.all())
            if allocations:
                for alloc in allocations:
                    lot = alloc.stock_lot
                    lignes_data.append({
                        'produit': ligne.produit_id,
                        'produit_nom': produit.name if produit else ligne.produit_nom,
                        'quantity': alloc.quantity,
                        'prix_unitaire': ligne.selling_price,
                        'remise': ligne.discount,
                        'tva': ligne.tva,
                        'lot': lot.lot if lot else '',
                        'stock_lot': lot.id if lot else None,
                        'date_expiration': lot.date_expiration.isoformat() if (lot and lot.date_expiration) else None,
                        'use_lot_management': produit.use_lot_management if produit else False,
                    })
            else:
                lignes_data.append({
                    'produit': ligne.produit_id,
                    'produit_nom': produit.name if produit else ligne.produit_nom,
                    'quantity': ligne.quantity,
                    'prix_unitaire': ligne.selling_price,
                    'remise': ligne.discount,
                    'tva': ligne.tva,
                    'lot': ligne.lot or '',
                    'stock_lot': ligne.stock_lot_id,
                    'date_expiration': None,
                    'use_lot_management': produit.use_lot_management if produit else False,
                })

        return Response({
            'facture_origine': facture.id,
            'facture_numero': facture.numero_facture,
            'client': facture.client_id,
            'client_name': facture.client.name if facture.client else facture.client_name_override,
            'montant_total': facture.total_ttc,
            'lignes': lignes_data,
        })

    @action(detail=False, methods=['get'])
    def exporter_excel(self, request):
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        queryset = self.get_queryset()
        avoirs = list(queryset.select_related('client', 'facture_origine'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Avoirs clients'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
        alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = [
            'Numéro', 'Date', 'Client', 'Facture origine',
            'Montant TTC', 'Statut', 'Motif', 'Nb lignes', 'Créé par'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        for row_idx, avoir in enumerate(avoirs, 2):
            ws.cell(row=row_idx, column=1, value=avoir.numero).border = border
            ws.cell(row=row_idx, column=2, value=avoir.date.strftime('%d/%m/%Y')).border = border
            ws.cell(row=row_idx, column=3, value=str(avoir.client) if avoir.client else avoir.facture_origine.client_name if avoir.facture_origine else '—').border = border
            ws.cell(row=row_idx, column=4, value=avoir.facture_origine.numero_facture if avoir.facture_origine else '—').border = border
            ws.cell(row=row_idx, column=5, value=float(avoir.montant_total)).border = border
            ws.cell(row=row_idx, column=6, value=avoir.statut).border = border
            ws.cell(row=row_idx, column=7, value=avoir.type_motif).border = border
            ws.cell(row=row_idx, column=8, value=avoir.lignes.count()).border = border
            ws.cell(row=row_idx, column=9, value=str(avoir.created_by) if avoir.created_by else '—').border = border

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].auto_size = True

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="avoirs_clients.xlsx"'
        wb.save(response)
        return response
