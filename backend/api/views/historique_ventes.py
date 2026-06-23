from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Sum, Count, Q, Avg, F as models_f, ExpressionWrapper, DecimalField, F, OuterRef, Subquery, Case, When
from django.db.models.functions import TruncDate, Coalesce
from django.utils import timezone
from datetime import datetime
from decimal import Decimal
from ..models import Facture, Caisse, FactureProduit, FactureProduitAllocation
import logging

logger = logging.getLogger(__name__)
business_logger = logging.getLogger('api.business')

class HistoriqueVentesViewSet(viewsets.ViewSet):
    """API endpoint for daily sales history."""
    permission_classes = [IsAuthenticated]
    
    def list(self, request):
        # Get query parameters
        from ..centralized_configs import PaginationHelper, PaginationDefaults
        date_debut = request.query_params.get('date_debut')
        date_fin = request.query_params.get('date_fin')
        page = PaginationHelper.get_page_number(request)
        page_size = PaginationHelper.get_page_size(request, PaginationDefaults.DEFAULT_REPORT_PAGE_SIZE)

        # Base queryset: only validated or paid invoices (exclude cancelled, brouillon, and proforma)
        # NOTE: L'historique des ventes INCLUT les is_divers (toutes les ventes)
        factures = Facture.objects.filter(
            status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE]
        )
        
        # Apply date filters
        if date_debut:
            factures = factures.filter(date__date__gte=date_debut)
        if date_fin:
            factures = factures.filter(date__date__lte=date_fin)

        # Group by date and aggregate
        daily_stats_query = factures.annotate(
            jour=TruncDate('date')
        ).values('jour').annotate(
            nb_ventes=Count('id'),
            ca_ht=Sum('total_ht'),
            ca_ttc=Sum('total_ttc'),  # INCLUT is_divers
            tva=Sum('total_tva'),
            total_remise=Sum('remise'),
        ).order_by('-jour')
        
        # Global totals for the period
        global_totals_agg = factures.aggregate(
            total_ttc=Sum('total_ttc'),
            total_ht=Sum('total_ht'),
            total_tva=Sum('total_tva'),
            total_ventes=Count('id'),
            total_remise=Sum('remise'),
        )
        
        # Payment totals for the entire filtered period
        # NOTE: Déjà filtré via facture__in=factures qui sont status__in=[VALIDEE, PAYEE]
        global_paiements = Caisse.objects.filter(
            facture__in=factures,
            statut='completee'
        ).exclude(mode_paiement='recouvrement').values('mode_paiement').annotate(
            total=Sum('montant')
        )
        
        pay_modes = ['especes', 'carte', 'cheque', 'virement', 'om', 'momo', 'coupon', 'en_compte']
        global_paiements_dict = {m: 0.0 for m in pay_modes}
        for p in global_paiements:
            mode = p['mode_paiement'].lower()
            if mode in global_paiements_dict:
                global_paiements_dict[mode] = float(p['total'] or 0)

        # Total count for pagination
        total_count = daily_stats_query.count()

        # Slice for current page
        start = (page - 1) * page_size
        end = start + page_size
        daily_stats = daily_stats_query[start:end]

        # ── Pré-calculs par jour en UNE requête chacun (évite N+1) ──
        from collections import defaultdict
        jours = [day['jour'] for day in daily_stats]
        factures_jour = factures.filter(date__date__in=jours) if jours else Facture.objects.none()

        paiements_par_jour = defaultdict(lambda: {m: 0.0 for m in pay_modes})
        if jours:
            for p in Caisse.objects.filter(
                facture__in=factures_jour,
                statut='completee'
            ).exclude(mode_paiement='recouvrement').annotate(
                jour=TruncDate('facture__date')
            ).values('jour', 'mode_paiement').annotate(
                total=Sum('montant')
            ):
                mode = (p['mode_paiement'] or '').lower()
                if mode in pay_modes:
                    paiements_par_jour[p['jour']][mode] = float(p['total'] or 0)

        remises_par_jour = defaultdict(float)
        if jours:
            for r in FactureProduit.objects.filter(
                facture__in=factures_jour
            ).annotate(
                jour=TruncDate('facture__date')
            ).values('jour').annotate(
                total=Sum(ExpressionWrapper(models_f('discount') * models_f('quantity'), output_field=DecimalField()))
            ):
                remises_par_jour[r['jour']] = float(r['total'] or 0)

        marges_par_jour = defaultdict(float)
        if jours:
            for m in FactureProduit.objects.filter(
                facture__in=factures_jour
            ).annotate(
                cout_unitaire=Case(
                    When(stock_lot__isnull=False, then=F('stock_lot__price_cost')),
                    When(produit__pmp__isnull=False, then=F('produit__pmp')),
                    default=F('produit__cost_price'),
                    output_field=DecimalField()
                ),
                marge_ligne=(F('selling_price') - F('cout_unitaire')) * F('quantity'),
                jour=TruncDate('facture__date')
            ).values('jour').annotate(
                marge=Sum('marge_ligne')
            ):
                marges_par_jour[m['jour']] = float(m['marge'] or 0)

        # Build results
        results = []
        for day in daily_stats:
            jour = day['jour']
            nb_ventes = day['nb_ventes'] or 0
            ca_ttc = float(day['ca_ttc'] or 0)
            panier_moyen = ca_ttc / nb_ventes if nb_ventes > 0 else 0
            remise = remises_par_jour.get(jour, 0.0) + float(day['total_remise'] or 0)
            modes = paiements_par_jour.get(jour, {m: 0.0 for m in pay_modes})

            results.append({
                'date': jour.strftime('%Y-%m-%d'),
                'nb_ventes': nb_ventes,
                'panier_moyen': round(panier_moyen, 2),
                'ca_ht': float(day['ca_ht'] or 0),
                'tva': float(day['tva'] or 0),
                'ca_ttc': ca_ttc,
                'marge': round(marges_par_jour.get(jour, 0.0), 2),
                'remise': round(remise, 2),
                'total_paiements': sum(modes.values()),
                **modes
            })
        
        return Response({
            'count': total_count,
            'results': results,
            'totals': {
                'ca_ttc': float(global_totals_agg['total_ttc'] or 0),
                'ca_ht': float(global_totals_agg['total_ht'] or 0),
                'tva': float(global_totals_agg['total_tva'] or 0),
                'nb_ventes': global_totals_agg['total_ventes'] or 0,
                'total_paiements': sum(global_paiements_dict.values()),
                **global_paiements_dict
            }
        })
    
    @action(detail=False, methods=['get'])
    def ventes_par_tranche(self, request):
        """
        Returns products sold during a time range with aggregated data.
        Columns: nom, qte_vendu, prix_vente, stock_restant
        """
        date_debut = request.query_params.get('date_debut')
        date_fin = request.query_params.get('date_fin')
        
        if not date_debut or not date_fin:
            return Response({'error': 'date_debut and date_fin are required'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Parse datetime and make timezone-aware
        try:
            debut = datetime.fromisoformat(date_debut.replace('Z', '+00:00'))
            fin = datetime.fromisoformat(date_fin.replace('Z', '+00:00'))
            
            # If datetime is naive (no timezone info), assume local timezone
            if debut.tzinfo is None:
                debut = timezone.make_aware(debut, timezone.get_current_timezone())
            if fin.tzinfo is None:
                fin = timezone.make_aware(fin, timezone.get_current_timezone())
                
        except ValueError:
            return Response({'error': 'Invalid datetime format'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Get FactureProduit for validated/paid invoices in the time range
        factures_in_range = Facture.objects.filter(
            status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
            date__gte=debut,
            date__lte=fin
        )

        facture_produits = FactureProduit.objects.filter(
            facture__in=factures_in_range
        ).select_related('produit', 'produit__rayon').values(
            'produit__id', 'produit__name', 'produit__stock', 'lot',
            'produit__cip1', 'produit__rayon__name'
        ).annotate(
            qte_vendu=Sum('quantity'),
            prix_vente=Avg('selling_price')
        ).order_by('-qte_vendu')
        
        logger.debug(f"[VENTES PAR TRANCHE] Trouvé {facture_produits.count()} produits distincts")
        
        # Format results with montant calculation
        results = []
        total_montant = 0
        
        for fp in facture_produits:
            montant = fp['qte_vendu'] * float(fp['prix_vente'])
            total_montant += montant
            results.append({
                'cip': fp['produit__cip1'] or '-',
                'nom': fp['produit__name'],
                'rayon': fp['produit__rayon__name'] or '-',
                'lot': fp['lot'] or '-',
                'qte_vendu': fp['qte_vendu'],
                'prix_vente': round(float(fp['prix_vente']), 0),
                'montant': round(montant, 0),
                'stock_restant': fp['produit__stock']
            })
        
        # Add total row at the end
        results.append({
            'cip': '-',
            'nom': 'TOTAL',
            'rayon': '-',
            'lot': '-',
            'qte_vendu': sum(r['qte_vendu'] for r in results),
            'prix_vente': '-',
            'montant': round(total_montant, 0),
            'stock_restant': '-'
        })
        
        return Response(results)

    @action(detail=False, methods=['get'])
    def exporter_excel(self, request):
        """
        Génère un export Excel de l'historique des ventes.
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        
        # Récupérer les données filtrées (sans pagination)
        date_debut = request.query_params.get('date_debut')
        date_fin = request.query_params.get('date_fin')
        
        # NOTE: L'historique des ventes INCLUT les is_divers (toutes les ventes)
        factures = Facture.objects.filter(
            status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE]
        )
        if date_debut:
            factures = factures.filter(date__date__gte=date_debut)
        if date_fin:
            factures = factures.filter(date__date__lte=date_fin)

        daily_stats = factures.annotate(
            jour=TruncDate('date')
        ).values('jour').annotate(
            nb_ventes=Count('id'),
            ca_ht=Sum('total_ht'),
            ca_ttc=Sum('total_ttc'),  # INCLUT is_divers
            tva=Sum('total_tva')
        ).order_by('-jour')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historique Ventes"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        headers = [
            "Date", "Ventes", "CA HT", "TVA", "CA TTC", 
            "Espèces", "Carte", "Chèque", "Virement", "OM", "MOMO", "Coupon", "En Compte", "Panier Moyen"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border
            
        row_idx = 2
        for day in daily_stats:
            jour = day['jour']
            nb_ventes = day['nb_ventes'] or 0
            ca_ttc = float(day['ca_ttc'] or 0)
            panier_moyen = ca_ttc / nb_ventes if nb_ventes > 0 else 0
            
            # NOTE: Filtrer aussi par statut facture pour cohérence avec ca_ttc
            paiements = Caisse.objects.filter(
                facture__date__date=jour,
                facture__status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
                statut='completee'
            ).exclude(mode_paiement='recouvrement').values('mode_paiement').annotate(
                total=Sum('montant')
            )
            
            modes = {m: 0 for m in ['especes', 'carte', 'cheque', 'virement', 'om', 'momo', 'coupon', 'en_compte']}
            for p in paiements:
                if p['mode_paiement'] in modes:
                    modes[p['mode_paiement']] = float(p['total'] or 0)
                    
            data_row = [
                jour.strftime('%d/%m/%Y'),
                nb_ventes,
                round(float(day['ca_ht'] or 0), 0),
                round(float(day['tva'] or 0), 0),
                round(ca_ttc, 0),
                round(modes['especes'], 0), round(modes['carte'], 0), round(modes['cheque'], 0),
                round(modes['virement'], 0), round(modes['om'], 0), round(modes['momo'], 0), round(modes['coupon'], 0),
                round(modes['en_compte'], 0),
                round(panier_moyen, 0)
            ]
            
            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
            row_idx += 1
            
        # Ajuster largeur colonnes
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Historique_Ventes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        return response


