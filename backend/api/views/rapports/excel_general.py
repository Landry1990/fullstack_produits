"""
Rapport général mensuel Excel — Pharmacie ZENITH
Génère un classeur .xlsx multi-feuilles pour le mois écoulé.

Feuilles :
  1. Synthèse
  2. CA & Marges (jour par jour)
  3. Remises & Avoirs
  4. Top Produits
  5. Dettes Fournisseurs
  6. Créances Clients
  7. Stock & Inventaire
  8. Achats Fournisseurs (avec précompte)
  9. État des Caisses
 10. Dépenses
 11. Synthèse Fiscale (accompte + précompte selon régime/mode)
 12. UGs (Unités Gratuites : reçues, vendues, restantes)
"""
from __future__ import annotations

import io
import os
from collections import defaultdict
from datetime import datetime, date, time, timedelta
from decimal import Decimal, ROUND_HALF_UP

import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# Palette & helpers de style
# ─────────────────────────────────────────────────────────────────────────────

PALETTE = {
    "header_bg":   "1F4E79",   # bleu marine
    "header_fg":   "FFFFFF",
    "sub_bg":      "2E75B6",   # bleu moyen (sous-titre)
    "sub_fg":      "FFFFFF",
    "alt_row":     "D6E4F0",   # bleu très clair (ligne paire)
    "total_bg":    "BDD7EE",   # bleu clair (totaux)
    "total_fg":    "1F4E79",
    "pos":         "00B050",   # vert (valeur positive)
    "neg":         "FF0000",   # rouge (valeur négative)
    "warning":     "FF9900",   # orange
    "white":       "FFFFFF",
    "light_gray":  "F2F2F2",
}

THIN = Side(style="thin", color="BFBFBF")
MED  = Side(style="medium", color="1F4E79")


def _border(sides="all", thick=False):
    s = MED if thick else THIN
    if sides == "all":
        return Border(left=s, right=s, top=s, bottom=s)
    elif sides == "bottom":
        return Border(bottom=s)
    return Border()


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=11, color="000000", italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _fmt(v, currency=True) -> str:
    if v is None:
        return "—"
    try:
        n = float(v)
        if currency:
            return f"{n:,.0f} F".replace(",", " ")
        return f"{n:,.2f}".replace(",", " ")
    except (TypeError, ValueError):
        return str(v)


def _pct(v) -> str:
    try:
        return f"{float(v):.1f} %"
    except Exception:
        return "—"


# ─────────────────────────────────────────────────────────────────────────────
# En-tête pharmacie (commun à toutes les feuilles)
# ─────────────────────────────────────────────────────────────────────────────

def _write_sheet_header(ws, pharmacy, titre: str, logo_path: str | None = None):
    """Écrit un bloc en-tête stylisé avec nom pharmacie, adresse et titre."""
    ws.sheet_view.showGridLines = False

    # Logo (colonne A, lignes 1-4) si disponible
    logo_col_offset = 0
    if logo_path and os.path.exists(logo_path):
        try:
            img = XLImage(logo_path)
            img.width, img.height = 100, 60
            ws.add_image(img, "A1")
            logo_col_offset = 2
        except Exception:
            pass

    col = logo_col_offset + 1  # colonne de départ pour le texte d'en-tête

    # Nom pharmacie
    cell = ws.cell(row=1, column=col, value=pharmacy.get("name", "PHARMACIE"))
    cell.font = _font(bold=True, size=14, color=PALETTE["header_bg"])
    cell.alignment = _align()

    # Adresse / contact
    ws.cell(row=2, column=col,
            value=f"{pharmacy.get('address', '')} — {pharmacy.get('city', '')}".strip(" —"))
    ws.cell(row=3, column=col,
            value=f"Tél : {pharmacy.get('phone', '')}  |  {pharmacy.get('email', '')}".strip(" |"))

    # Titre du rapport
    cell = ws.cell(row=4, column=col, value=titre)
    cell.font = _font(bold=True, size=13, color=PALETTE["sub_bg"])

    # Date d'édition
    now = timezone.localtime(timezone.now())
    ws.cell(row=5, column=col,
            value=f"Édité le {now.strftime('%d/%m/%Y à %H:%M')}")
    ws.cell(row=5, column=col).font = _font(italic=True, size=9, color="808080")

    # Séparateur
    ws.append([])  # ligne 6 vide
    ws.append([])  # ligne 7 vide → les données commencent à la ligne 8
    return 8  # première ligne disponible


def _write_col_headers(ws, row: int, headers: list[str]):
    """Écrit une ligne d'en-tête de colonnes stylisée."""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = _font(bold=True, color=PALETTE["header_fg"])
        cell.fill = _fill(PALETTE["header_bg"])
        cell.alignment = _align(h="center", wrap=True)
        cell.border = _border()


def _auto_width(ws, min_w=10, max_w=30):
    merged_ranges = set()
    for mr in ws.merged_cells.ranges:
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                merged_ranges.add((row, col))
    for col in ws.columns:
        length = 0
        for cell in col:
            if (cell.row, cell.column) in merged_ranges:
                continue
            length = max(length, len(str(cell.value or "")))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w, max(min_w, length + 2))


def _row_style(ws, row: int, n_cols: int, is_alt: bool = False, is_total: bool = False):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        if is_total:
            cell.font = _font(bold=True, color=PALETTE["total_fg"])
            cell.fill = _fill(PALETTE["total_bg"])
        elif is_alt:
            cell.fill = _fill(PALETTE["alt_row"])
        cell.border = _border()


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 1 — Synthèse
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_synthese(wb, data: dict, pharmacy: dict, mois_label: str, logo_path):
    ws = wb.create_sheet("Synthèse")
    start = _write_sheet_header(ws, pharmacy, f"SYNTHÈSE DU MOIS — {mois_label}", logo_path)

    ca     = data["ca"]
    marge  = data["marge"]
    divers = data["divers_ca"]

    rows = [
        ("CA PHARMACIE TTC",      ca["ca_ttc"] - divers["ca_ttc"],                True),
        ("CA DIVERS TTC",         divers["ca_ttc"],                                True),
        ("CA TOTAL TTC",          ca["ca_ttc"],                                    True),
        ("", "", False),
        ("Marge brute Pharmacie", marge["marge_pharma"],                           False),
        ("Taux marge Pharmacie",  None,                                            False),
        ("Marge brute Divers",    marge["marge_divers"],                           False),
        ("Marge brute TOTALE",    marge["marge_brute"],                            False),
        ("Taux marge TOTAL",      None,                                            False),
        ("", "", False),
        ("Remises accordées",     ca.get("total_remises", 0),                      False),
        ("  dont remises globales",  ca.get("total_remises_detail", {}).get("global", 0),   False),
        ("  dont remises lignes",    ca.get("total_remises_detail", {}).get("lignes", 0),   False),
        ("  dont fidélité",          ca.get("total_remises_detail", {}).get("fidelite", 0), False),
        ("", "", False),
        ("Nombre de ventes",      ca.get("nb_ventes", 0),                          False),
        ("Panier moyen TTC",      (ca["ca_ttc"] / ca["nb_ventes"]) if ca.get("nb_ventes") else 0, False),
        ("", "", False),
        ("Créances clients",      data["creances"]["total"],                       False),
        ("Dettes fournisseurs",   data.get("dettes_total", 0),                     False),
    ]

    # Calcul taux marge pharmacie
    ca_pharma = float(ca["ca_ttc"]) - float(divers["ca_ttc"])
    marge_pharma = float(marge.get("marge_pharma", 0))
    taux_pharma = (marge_pharma / ca_pharma * 100) if ca_pharma else 0.0
    taux_total = float(marge.get("marge_pct", 0))

    _write_col_headers(ws, start, ["Indicateur", "Montant / Valeur"])
    r = start + 1
    for i, (label, val, _is_ca) in enumerate(rows):
        if label == "":
            r += 1
            continue
        is_total = label.startswith("CA TOTAL") or label.startswith("Marge brute TOTALE")
        ws.cell(row=r, column=1, value=label).alignment = _align()
        if "Taux marge Pharmacie" in label:
            ws.cell(row=r, column=2, value=_pct(taux_pharma)).alignment = _align(h="right")
        elif "Taux marge TOTAL" in label:
            ws.cell(row=r, column=2, value=_pct(taux_total)).alignment = _align(h="right")
        elif "Nombre" in label:
            ws.cell(row=r, column=2, value=int(val)).alignment = _align(h="right")
        else:
            ws.cell(row=r, column=2, value=_fmt(val)).alignment = _align(h="right")
        _row_style(ws, r, 2, is_alt=(i % 2 == 0), is_total=is_total)
        r += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 22


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 2 — CA & Marges (jour par jour)
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_ca_marges(wb, daily_rows: list, pharmacy: dict, mois_label: str, logo_path):
    ws = wb.create_sheet("CA & Marges")
    start = _write_sheet_header(ws, pharmacy, f"CA & MARGES JOURNALIERS — {mois_label}", logo_path)

    headers = [
        "Date", "Nb Ventes",
        "CA Pharmacie TTC", "CA Divers TTC", "CA Total TTC",
        "Marge Pharmacie", "Marge Divers", "Marge Totale", "Taux Marge %",
        "Remises"
    ]
    _write_col_headers(ws, start, headers)

    totals = defaultdict(Decimal)
    r = start + 1
    for i, row in enumerate(daily_rows):
        vals = [
            row["date"],
            row["nb_ventes"],
            _fmt(row["ca_pharma"]),
            _fmt(row["ca_divers"]),
            _fmt(row["ca_total"]),
            _fmt(row["marge_pharma"]),
            _fmt(row["marge_divers"]),
            _fmt(row["marge_totale"]),
            _pct(row["taux_marge"]),
            _fmt(row["remises"]),
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v).alignment = _align(h="right" if c > 1 else "left")
        _row_style(ws, r, len(headers), is_alt=(i % 2 == 0))
        for key in ["ca_pharma", "ca_divers", "ca_total", "marge_pharma", "marge_divers", "marge_totale", "remises"]:
            totals[key] += Decimal(str(row.get(key, 0) or 0))
        totals["nb_ventes"] += int(row.get("nb_ventes", 0))
        r += 1

    # Ligne total
    ca_total_f = float(totals["ca_total"])
    marge_tot_f = float(totals["marge_totale"])
    total_vals = [
        "TOTAL", int(totals["nb_ventes"]),
        _fmt(totals["ca_pharma"]), _fmt(totals["ca_divers"]), _fmt(totals["ca_total"]),
        _fmt(totals["marge_pharma"]), _fmt(totals["marge_divers"]), _fmt(totals["marge_totale"]),
        _pct(marge_tot_f / ca_total_f * 100 if ca_total_f else 0),
        _fmt(totals["remises"]),
    ]
    for c, v in enumerate(total_vals, 1):
        ws.cell(row=r, column=c, value=v)
    _row_style(ws, r, len(headers), is_total=True)
    _auto_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 3 — Remises & Avoirs
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_remises_avoirs(wb, remises_rows: list, avoirs_rows: list, pharmacy: dict, mois_label: str, logo_path):
    ws = wb.create_sheet("Remises & Avoirs")
    start = _write_sheet_header(ws, pharmacy, f"REMISES & AVOIRS — {mois_label}", logo_path)

    # ── Remises ──
    ws.cell(row=start, column=1, value="REMISES ACCORDÉES").font = _font(bold=True, size=12, color=PALETTE["sub_bg"])
    start += 1
    h_remises = ["Date", "N° Facture", "Client", "Remise Globale (F)", "Remises Lignes (F)", "Fidélité (F)", "Total Remise (F)"]
    _write_col_headers(ws, start, h_remises)
    r = start + 1
    total_rem = Decimal("0")
    for i, row in enumerate(remises_rows):
        vals = [
            row.get("date", ""),
            row.get("numero_facture", ""),
            row.get("client", "—"),
            _fmt(row.get("remise_globale", 0)),
            _fmt(row.get("remise_lignes", 0)),
            _fmt(row.get("remise_fidelite", 0)),
            _fmt(row.get("total_remise", 0)),
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v).alignment = _align(h="right" if c > 2 else "left")
        _row_style(ws, r, len(h_remises), is_alt=(i % 2 == 0))
        total_rem += Decimal(str(row.get("total_remise", 0) or 0))
        r += 1
    # Total remises
    for c, v in enumerate(["TOTAL", "", "", "", "", "", _fmt(total_rem)], 1):
        ws.cell(row=r, column=c, value=v)
    _row_style(ws, r, len(h_remises), is_total=True)
    r += 2

    # ── Avoirs ──
    ws.cell(row=r, column=1, value="AVOIRS FOURNISSEURS").font = _font(bold=True, size=12, color=PALETTE["sub_bg"])
    r += 1
    h_avoirs = ["Date", "N° Avoir", "Fournisseur", "Montant HT (F)", "Statut"]
    _write_col_headers(ws, r, h_avoirs)
    r += 1
    total_av = Decimal("0")
    for i, row in enumerate(avoirs_rows):
        vals = [
            row.get("date", ""),
            row.get("numero", ""),
            row.get("fournisseur", "—"),
            _fmt(row.get("montant", 0)),
            row.get("statut", ""),
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v).alignment = _align(h="right" if c > 2 else "left")
        _row_style(ws, r, len(h_avoirs), is_alt=(i % 2 == 0))
        total_av += Decimal(str(row.get("montant", 0) or 0))
        r += 1
    for c, v in enumerate(["TOTAL", "", "", _fmt(total_av), ""], 1):
        ws.cell(row=r, column=c, value=v)
    _row_style(ws, r, len(h_avoirs), is_total=True)
    _auto_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 4 — Top Produits
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_top_produits(wb, top_rows: list, pharmacy: dict, mois_label: str, logo_path):
    ws = wb.create_sheet("Top Produits")
    start = _write_sheet_header(ws, pharmacy, f"TOP PRODUITS — {mois_label}", logo_path)

    headers = ["Rang", "Produit", "CIP", "Rayon", "Type", "Qté vendue", "CA TTC (F)", "Coût achat (F)", "Marge (F)", "Taux Marge %"]
    _write_col_headers(ws, start, headers)
    r = start + 1
    for i, row in enumerate(top_rows):
        vals = [
            i + 1,
            row.get("produit_nom", ""),
            row.get("cip", ""),
            row.get("rayon", ""),
            "Divers" if row.get("is_divers") else "Pharmacie",
            row.get("quantite", 0),
            _fmt(row.get("ca_ttc", 0)),
            _fmt(row.get("cout_achat", 0)),
            _fmt(row.get("marge", 0)),
            _pct(row.get("taux_marge", 0)),
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v).alignment = _align(h="right" if c > 4 else "left")
        _row_style(ws, r, len(headers), is_alt=(i % 2 == 0))
        r += 1
    _auto_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 5 — Dettes Fournisseurs
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_dettes_fournisseurs(wb, dettes_rows: list, pharmacy: dict, mois_label: str, logo_path):
    ws = wb.create_sheet("Dettes Fournisseurs")
    start = _write_sheet_header(ws, pharmacy, f"DETTES FOURNISSEURS — {mois_label}", logo_path)

    headers = ["Fournisseur", "Total Commandé (F)", "Total Payé (F)", "Solde Dû (F)", "Échu (F)", "À venir (F)", "Prochaine échéance"]
    _write_col_headers(ws, start, headers)
    r = start + 1
    totals = defaultdict(Decimal)
    for i, row in enumerate(dettes_rows):
        vals = [
            row.get("fournisseur", ""),
            _fmt(row.get("total_commande", 0)),
            _fmt(row.get("total_paye", 0)),
            _fmt(row.get("solde_du", 0)),
            _fmt(row.get("echu", 0)),
            _fmt(row.get("a_venir", 0)),
            row.get("prochaine_echeance", "—"),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = _align(h="right" if c > 1 else "left")
            if c == 4 and float(row.get("solde_du", 0) or 0) > 0:
                cell.font = _font(bold=True, color=PALETTE["neg"])
        _row_style(ws, r, len(headers), is_alt=(i % 2 == 0))
        for k, col in [("total_commande", "total_commande"), ("total_paye", "total_paye"),
                       ("solde_du", "solde_du"), ("echu", "echu"), ("a_venir", "a_venir")]:
            totals[k] += Decimal(str(row.get(k, 0) or 0))
        r += 1
    for c, v in enumerate(["TOTAL", _fmt(totals["total_commande"]), _fmt(totals["total_paye"]),
                            _fmt(totals["solde_du"]), _fmt(totals["echu"]), _fmt(totals["a_venir"]), ""], 1):
        ws.cell(row=r, column=c, value=v)
    _row_style(ws, r, len(headers), is_total=True)
    _auto_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 6 — Créances Clients
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_creances_clients(wb, creances_rows: list, pharmacy: dict, mois_label: str, logo_path):
    ws = wb.create_sheet("Créances Clients")
    start = _write_sheet_header(ws, pharmacy, f"CRÉANCES CLIENTS — {mois_label}", logo_path)

    headers = ["Client", "Type", "Nb Factures", "Montant Total (F)", "Payé (F)", "Restant Dû (F)", "Jours retard (max)"]
    _write_col_headers(ws, start, headers)
    r = start + 1
    totals = defaultdict(Decimal)
    for i, row in enumerate(creances_rows):
        retard = row.get("jours_retard", 0) or 0
        vals = [
            row.get("client_nom", ""),
            row.get("client_type", ""),
            row.get("nb_factures", 0),
            _fmt(row.get("ca_total", 0)),
            _fmt(row.get("montant_paye", 0)),
            _fmt(row.get("reste_a_payer", 0)),
            int(retard),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = _align(h="right" if c > 2 else "left")
            if c == 6 and float(row.get("reste_a_payer", 0) or 0) > 0:
                cell.font = _font(color=PALETTE["neg"])
        _row_style(ws, r, len(headers), is_alt=(i % 2 == 0))
        totals["ca_total"] += Decimal(str(row.get("ca_total", 0) or 0))
        totals["montant_paye"] += Decimal(str(row.get("montant_paye", 0) or 0))
        totals["reste_a_payer"] += Decimal(str(row.get("reste_a_payer", 0) or 0))
        r += 1
    for c, v in enumerate(["TOTAL", "", "", _fmt(totals["ca_total"]), _fmt(totals["montant_paye"]), _fmt(totals["reste_a_payer"]), ""], 1):
        ws.cell(row=r, column=c, value=v)
    _row_style(ws, r, len(headers), is_total=True)
    _auto_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 7 — Stock & Inventaire
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_stock(wb, stock_rows: list, pharmacy: dict, mois_label: str, logo_path):
    ws = wb.create_sheet("Stock & Inventaire")
    start = _write_sheet_header(ws, pharmacy, f"STOCK & INVENTAIRE — {mois_label}", logo_path)

    headers = ["Rayon", "Nb Références", "Qté totale", "Valeur Stock PMP (F)", "Stock minimum", "Alertes rupture", "Alertes péremption"]
    _write_col_headers(ws, start, headers)
    r = start + 1
    totals = defaultdict(Decimal)
    for i, row in enumerate(stock_rows):
        vals = [
            row.get("rayon", ""),
            row.get("nb_refs", 0),
            row.get("qte_totale", 0),
            _fmt(row.get("valeur_pmp", 0)),
            row.get("stock_minimum", 0),
            row.get("alertes_rupture", 0),
            row.get("alertes_peremption", 0),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = _align(h="right" if c > 1 else "left")
            if c == 6 and int(row.get("alertes_rupture", 0) or 0) > 0:
                cell.font = _font(color=PALETTE["neg"])
        _row_style(ws, r, len(headers), is_alt=(i % 2 == 0))
        totals["nb_refs"] += Decimal(str(row.get("nb_refs", 0) or 0))
        totals["qte_totale"] += Decimal(str(row.get("qte_totale", 0) or 0))
        totals["valeur_pmp"] += Decimal(str(row.get("valeur_pmp", 0) or 0))
        totals["alertes_rupture"] += Decimal(str(row.get("alertes_rupture", 0) or 0))
        r += 1
    for c, v in enumerate(["TOTAL", int(totals["nb_refs"]), int(totals["qte_totale"]),
                            _fmt(totals["valeur_pmp"]), "", int(totals["alertes_rupture"]), ""], 1):
        ws.cell(row=r, column=c, value=v)
    _row_style(ws, r, len(headers), is_total=True)
    _auto_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 8 — Achats Fournisseurs
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_fiscal(wb, data: dict, pharmacy: dict, mois_label: str, logo_path):
    """Feuille Synthèse Fiscale : accompte, précompte, et ventilation selon régime/mode."""
    from api.models import PharmacySettings

    ps = PharmacySettings.objects.first()
    if not ps:
        return

    ws = wb.create_sheet("Synthèse Fiscale")
    start = _write_sheet_header(ws, pharmacy, f"SYNTHÈSE FISCALE — {mois_label}", logo_path)

    regime = ps.regime_fiscal
    mode = ps.mode_imposition
    regime_label = ps.get_regime_fiscal_display()
    mode_label = ps.get_mode_imposition_display()

    r = start + 1
    # Section : Configuration
    ws.cell(row=r, column=1, value="Régime fiscal").font = _font(bold=True)
    ws.cell(row=r, column=2, value=regime_label).alignment = _align(h="right")
    r += 1
    ws.cell(row=r, column=1, value="Mode d'imposition").font = _font(bold=True)
    ws.cell(row=r, column=2, value=mode_label).alignment = _align(h="right")
    r += 2

    # Section : Chiffre d'affaires
    ca_data = data.get("ca", {})
    ws.cell(row=r, column=1, value="CHIFFRE D'AFFAIRES").font = _font(bold=True, size=11)
    r += 1
    for label, key in [("CA HT", "ca_ht"), ("CA TTC", "ca_ttc"), ("TVA collectée", "tva_collectee")]:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=_fmt(ca_data.get(key, 0))).alignment = _align(h="right")
        r += 1
    r += 1

    # Section : Achats & Précompte
    achats_data = data.get("achats_fiscaux", {})
    ws.cell(row=r, column=1, value="ACHATS & PRÉCOMPTE").font = _font(bold=True, size=11)
    r += 1
    ws.cell(row=r, column=1, value="Total achats HT")
    ws.cell(row=r, column=2, value=_fmt(achats_data.get("total_achats_ht", 0))).alignment = _align(h="right")
    r += 1
    if mode == 'DROIT_COMMUN':
        taux_prec = float(ps.taux_precompte_reel) if regime == 'REEL' else float(ps.taux_precompte_simplifie)
        ws.cell(row=r, column=1, value=f"Taux précompte ({regime_label})")
        ws.cell(row=r, column=2, value=f"{taux_prec}%").alignment = _align(h="right")
        r += 1
        ws.cell(row=r, column=1, value="Total précompte").font = _font(bold=True)
        ws.cell(row=r, column=2, value=_fmt(achats_data.get("total_precompte", 0))).alignment = _align(h="right")
        r += 1
    else:
        ws.cell(row=r, column=1, value="Précompte (non applicable en Marge Administrée)")
        ws.cell(row=r, column=2, value="—").alignment = _align(h="right")
        r += 1
    r += 1

    # Section : Accompte
    accompte_data = data.get("accompte_fiscal", {})
    ws.cell(row=r, column=1, value="ACCOMPTE (sur CA)").font = _font(bold=True, size=11)
    r += 1
    ws.cell(row=r, column=1, value=accompte_data.get("base_label", "Base"))
    ws.cell(row=r, column=2, value=_fmt(accompte_data.get("base_imposition", 0))).alignment = _align(h="right")
    r += 1
    ws.cell(row=r, column=1, value="Taux")
    ws.cell(row=r, column=2, value=f"{accompte_data.get('taux', 0)}%").alignment = _align(h="right")
    r += 1
    ws.cell(row=r, column=1, value="Accompte base")
    ws.cell(row=r, column=2, value=_fmt(accompte_data.get("accompte_base", 0))).alignment = _align(h="right")
    r += 1
    ws.cell(row=r, column=1, value=f"CAC ({accompte_data.get('taux_cac', 0)}%)")
    ws.cell(row=r, column=2, value=_fmt(accompte_data.get("accompte_cac", 0))).alignment = _align(h="right")
    r += 1
    ws.cell(row=r, column=1, value="Accompte total").font = _font(bold=True)
    ws.cell(row=r, column=2, value=_fmt(accompte_data.get("accompte_total", 0))).alignment = _align(h="right")
    r += 1

    _auto_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Feuille UGs — Unités Gratuites (achetées, vendues, restantes)
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_ugs(wb, data: dict, pharmacy: dict, mois_label: str, logo_path):
    """Feuille UGs : unités gratuites reçues en achat, vendues, et restantes en stock."""
    from api.models import StockLot, FactureProduit, Produit

    ws = wb.create_sheet("UGs (Unités Gratuites)")
    start = _write_sheet_header(ws, pharmacy, f"UNITÉS GRATUITES (UGs) — {mois_label}", logo_path)

    headers = ["Produit", "Fournisseur", "UG reçues", "UG vendues", "UG restantes", "Lot #"]
    _write_col_headers(ws, start, headers)
    r = start + 1
    totals = defaultdict(int)

    # Récupérer tous les lots avec UGs
    lots = StockLot.objects.filter(
        quantity_free__gt=0
    ).select_related('produit', 'fournisseur').order_by('-quantity_free')

    for lot in lots:
        prod_name = lot.produit.name if lot.produit else (lot.produit_nom or "—")
        fournisseur = lot.fournisseur.name if lot.fournisseur else (lot.fournisseur_nom or "—")
        ug_recues = lot.quantity_free
        ug_vendues = lot.quantity_free - lot.quantity_free_remaining
        ug_restantes = lot.quantity_free_remaining

        vals = [prod_name, fournisseur, ug_recues, ug_vendues, ug_restantes, f"#{lot.id}"]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v).alignment = _align(h="right" if c > 2 else "left")
        _row_style(ws, r, len(headers), is_alt=(r % 2 == 0))
        totals["ug_recues"] += ug_recues
        totals["ug_vendues"] += ug_vendues
        totals["ug_restantes"] += ug_restantes
        r += 1

    # Ligne totale
    for c, v in enumerate(["TOTAL", "", totals["ug_recues"], totals["ug_vendues"],
                            totals["ug_restantes"], ""], 1):
        ws.cell(row=r, column=c, value=v)
    _row_style(ws, r, len(headers), is_total=True)
    r += 2

    # Section résumé
    ws.cell(row=r, column=1, value="RÉSUMÉ").font = _font(bold=True, size=11)
    r += 1
    ws.cell(row=r, column=1, value="Total UGs reçues (achats)")
    ws.cell(row=r, column=2, value=totals["ug_recues"]).alignment = _align(h="right")
    r += 1
    ws.cell(row=r, column=1, value="Total UGs vendues (avec factures)")
    ws.cell(row=r, column=2, value=totals["ug_vendues"]).alignment = _align(h="right")
    r += 1
    ws.cell(row=r, column=1, value="Total UGs restantes (en stock)")
    ws.cell(row=r, column=2, value=totals["ug_restantes"]).alignment = _align(h="right")
    r += 1
    ws.cell(row=r, column=1, value="Taux de rotation UG")
    taux = (totals["ug_vendues"] / totals["ug_recues"] * 100) if totals["ug_recues"] else 0
    ws.cell(row=r, column=2, value=f"{taux:.1f}%").alignment = _align(h="right")

    _auto_width(ws)


def _sheet_achats(wb, achats_rows: list, pharmacy: dict, mois_label: str, logo_path):
    ws = wb.create_sheet("Achats Fournisseurs")
    start = _write_sheet_header(ws, pharmacy, f"ACHATS FOURNISSEURS — {mois_label}", logo_path)

    headers = ["Fournisseur", "Nb Commandes", "Montant HT (F)", "Précompte (F)", "Nb Avoirs", "Montant Avoirs (F)", "Net (F)"]
    _write_col_headers(ws, start, headers)
    r = start + 1
    totals = defaultdict(Decimal)
    for i, row in enumerate(achats_rows):
        net = Decimal(str(row.get("montant_total", 0) or 0))
        precompte = Decimal(str(row.get("precompte", 0) or 0))
        vals = [
            row.get("fournisseur_nom", ""),
            row.get("nb_commandes", 0),
            _fmt(net + Decimal(str(row.get("montant_avoirs", 0) or 0))),
            _fmt(precompte),
            row.get("nb_avoirs", 0),
            _fmt(row.get("montant_avoirs", 0)),
            _fmt(net),
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v).alignment = _align(h="right" if c > 1 else "left")
        _row_style(ws, r, len(headers), is_alt=(i % 2 == 0))
        totals["montant_total"] += net
        totals["montant_avoirs"] += Decimal(str(row.get("montant_avoirs", 0) or 0))
        totals["precompte"] += precompte
        r += 1
    for c, v in enumerate(["TOTAL", "", _fmt(totals["montant_total"] + totals["montant_avoirs"]),
                            _fmt(totals["precompte"]),
                            "", _fmt(totals["montant_avoirs"]), _fmt(totals["montant_total"])], 1):
        ws.cell(row=r, column=c, value=v)
    _row_style(ws, r, len(headers), is_total=True)
    _auto_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 9 — État des Caisses
# ─────────────────────────────────────────────────────────────────────────────

def _write_clotures_table(ws, rows: list, start_row: int) -> int:
    """Écrit un tableau de clôtures caisse à partir de start_row. Retourne la prochaine ligne disponible."""
    headers = [
        "Date", "Caisse / Vendeur", "Heure ouverture", "Heure fermeture",
        "Ventes (F)", "Entrées (F)", "Sorties (F)",
        "Montant théorique (F)", "Montant réel (F)", "Écart (F)", "Observation"
    ]
    _write_col_headers(ws, start_row, headers)
    r = start_row + 1
    totals = defaultdict(Decimal)
    for i, row in enumerate(rows):
        ecart = float(row.get("ecart_caisse", 0) or 0)
        vals = [
            row.get("date", ""),
            row.get("caisse", ""),
            row.get("heure_ouverture", ""),
            row.get("heure_fermeture", ""),
            _fmt(row.get("total_ventes", 0)),
            _fmt(row.get("total_entrees", 0)),
            _fmt(row.get("total_sorties", 0)),
            _fmt(row.get("montant_theorique", 0)),
            _fmt(row.get("montant_reel", 0)),
            _fmt(ecart),
            row.get("observation", ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = _align(h="right" if c in (5, 6, 7, 8, 9, 10) else "left")
            if c == 10:
                if ecart < -100:
                    cell.font = _font(color=PALETTE["neg"], bold=True)
                elif ecart > 100:
                    cell.font = _font(color=PALETTE["pos"], bold=True)
        _row_style(ws, r, len(headers), is_alt=(i % 2 == 0))
        for k in ["total_ventes", "total_entrees", "total_sorties", "montant_theorique", "montant_reel", "ecart_caisse"]:
            totals[k] += Decimal(str(row.get(k, 0) or 0))
        r += 1
    totals_vals = ["TOTAL", "", "", "",
                   _fmt(totals["total_ventes"]), _fmt(totals["total_entrees"]), _fmt(totals["total_sorties"]),
                   _fmt(totals["montant_theorique"]), _fmt(totals["montant_reel"]), _fmt(totals["ecart_caisse"]), ""]
    for c, v in enumerate(totals_vals, 1):
        ws.cell(row=r, column=c, value=v)
    _row_style(ws, r, len(headers), is_total=True)
    return r + 1


def _sheet_etat_caisses(wb, clotures_rows: list, pharmacy: dict, mois_label: str, logo_path):
    ws = wb.create_sheet("État des Caisses")
    start = _write_sheet_header(ws, pharmacy, f"ÉTAT DES CAISSES — {mois_label}", logo_path)

    # ── Tableau global ────────────────────────────────────────────────────────
    cell = ws.cell(row=start, column=1, value="RÉCAPITULATIF GÉNÉRAL")
    cell.font = _font(bold=True, size=12, color=PALETTE["sub_bg"])
    r = _write_clotures_table(ws, clotures_rows, start + 1)

    # ── Sous-tableaux par caissier ────────────────────────────────────────────
    # Regrouper par caissier (en conservant l'ordre d'apparition)
    caissiers: dict = {}
    for row in clotures_rows:
        nom = row.get("caissier") or "Inconnu"
        if nom not in caissiers:
            caissiers[nom] = []
        caissiers[nom].append(row)

    if len(caissiers) > 1:
        r += 1  # ligne vide de séparation
        for nom, rows in caissiers.items():
            cell = ws.cell(row=r, column=1, value=f"CAISSIER : {nom.upper()}")
            cell.font = _font(bold=True, size=11, color=PALETTE["header_bg"])
            cell.fill = _fill(PALETTE["alt_row"])
            r += 1
            r = _write_clotures_table(ws, rows, r)
            r += 1  # séparation entre caissiers

    _auto_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Feuille 10 — Dépenses
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_depenses(wb, depenses_rows: list, pharmacy: dict, mois_label: str, logo_path):
    ws = wb.create_sheet("Dépenses")
    start = _write_sheet_header(ws, pharmacy, f"DÉPENSES — {mois_label}", logo_path)

    headers = ["Date", "Heure", "Caisse", "Type", "Motif", "Description", "Montant (F)", "Saisi par"]
    _write_col_headers(ws, start, headers)
    r = start + 1
    total_entrees = Decimal("0")
    total_sorties = Decimal("0")
    for i, row in enumerate(depenses_rows):
        typ = row.get("type", "SORTIE")
        montant = Decimal(str(row.get("montant", 0) or 0))
        dt = row.get("date")
        date_str = dt.strftime("%d/%m/%Y") if isinstance(dt, (datetime, date)) else str(dt or "")
        heure_str = dt.strftime("%H:%M") if isinstance(dt, datetime) else ""
        vals = [
            date_str,
            heure_str,
            row.get("caisse", "—"),
            typ,
            row.get("motif", ""),
            row.get("description", ""),
            _fmt(montant),
            row.get("user", ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = _align(h="right" if c == 7 else "left")
            if c == 4:
                cell.font = _font(color=PALETTE["neg"] if typ == "SORTIE" else PALETTE["pos"])
        _row_style(ws, r, len(headers), is_alt=(i % 2 == 0))
        if typ == "SORTIE":
            total_sorties += montant
        else:
            total_entrees += montant
        r += 1

    # Sous-totaux
    r += 1
    for label, val, color in [
        ("Total SORTIES (dépenses)", total_sorties, PALETTE["neg"]),
        ("Total ENTRÉES", total_entrees, PALETTE["pos"]),
        ("SOLDE", total_entrees - total_sorties, PALETTE["header_bg"]),
    ]:
        ws.cell(row=r, column=5, value=label).font = _font(bold=True, color=color)
        ws.cell(row=r, column=7, value=_fmt(val)).font = _font(bold=True, color=color)
        r += 1
    _auto_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Collecte des données
# ─────────────────────────────────────────────────────────────────────────────

def _compute_accompte_fiscal_inline(ca_ht, factures_qs):
    """Calcule l'accompte fiscal selon le régime et mode configurés dans PharmacySettings."""
    from api.models import PharmacySettings, FactureProduit

    ps = PharmacySettings.objects.first()
    if not ps:
        return {}

    regime = ps.regime_fiscal
    mode = ps.mode_imposition
    taux_cac = ps.taux_cac

    if mode == 'MARGE_ADMINISTREE':
        cout_ventes = Decimal('0')
        for fp in FactureProduit.objects.filter(facture__in=factures_qs).select_related('produit'):
            cout_ventes += Decimal(str(fp.produit.cost_price or 0)) * Decimal(str(fp.quantity))
        marge_brute = ca_ht - cout_ventes
        taux_imposition = ps.taux_marge_brute
        base_imposition = marge_brute
        base_label = 'Marge brute'
    else:
        if regime == 'REEL':
            taux_imposition = ps.taux_accompte_reel
        else:
            taux_imposition = ps.taux_accompte_simplifie
        base_imposition = ca_ht
        base_label = "Chiffre d'affaires HT"

    accompte_base = base_imposition * taux_imposition / Decimal('100')
    accompte_cac = accompte_base * taux_cac / Decimal('100')
    accompte_total = accompte_base + accompte_cac

    return {
        "base_label": base_label,
        "base_imposition": int(base_imposition.quantize(Decimal('1'), rounding=ROUND_HALF_UP)),
        "taux": float(taux_imposition),
        "accompte_base": int(accompte_base.quantize(Decimal('1'), rounding=ROUND_HALF_UP)),
        "taux_cac": float(taux_cac),
        "accompte_cac": int(accompte_cac.quantize(Decimal('1'), rounding=ROUND_HALF_UP)),
        "accompte_total": int(accompte_total.quantize(Decimal('1'), rounding=ROUND_HALF_UP)),
    }


def _collect_data(date_debut, date_fin):
    """Collecte toutes les données nécessaires aux 10 feuilles."""
    from django.db.models import (
        Count, DecimalField, ExpressionWrapper, F, OuterRef,
        Q, Subquery, Sum, Value, Max
    )
    from django.db.models.functions import Coalesce, TruncDate
    from api.views.rapports.tz_utils import local_trunc_date
    from api.models import (
        Facture, FactureProduit, FactureProduitAllocation,
        ClotureCaisse, MouvementCaisse, Produit, Rayon,
        Fournisseur, CommandeProduit, PaiementFournisseur,
        Commande, Avoir, Client, Caisse, StockLot,
    )
    from api.services.margin_service import MarginService

    # ── Factures de la période ──────────────────────────────────────────────
    factures = Facture.objects.filter(
        status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
        date__gte=date_debut, date__lt=date_fin,
    )

    # ── CA global ──────────────────────────────────────────────────────────
    # Sous-requête pour montant divers par facture
    divers_sub = FactureProduitAllocation.objects.filter(
        facture_produit__facture=OuterRef("pk"),
        stock_lot__is_divers=True,
    ).values("facture_produit__facture").annotate(
        total=Coalesce(Sum(F("selling_price") * F("quantity"), output_field=DecimalField()), Decimal("0"))
    ).values("total")

    factures_ann = factures.annotate(
        divers_amount=Coalesce(Subquery(divers_sub, output_field=DecimalField()), Decimal("0")),
    )
    agg = factures_ann.aggregate(
        ca_total_ttc=Coalesce(Sum("total_ttc"), Decimal("0")),
        ca_total_ht=Coalesce(Sum("total_ht"), Decimal("0")),
        ca_divers_ttc=Coalesce(Sum("divers_amount"), Decimal("0")),
        remises=Coalesce(Sum("remise"), Decimal("0")),
        remises_fid=Coalesce(Sum("montant_fidelite"), Decimal("0")),
        nb_ventes=Count("id"),
    )
    remises_lignes = FactureProduit.objects.filter(
        facture__in=factures
    ).aggregate(
        total=Coalesce(Sum(F("discount") * F("quantity"), output_field=DecimalField()), Decimal("0"))
    )["total"]

    ca_total   = agg["ca_total_ttc"]
    ca_divers  = agg["ca_divers_ttc"]
    ca_pharma  = ca_total - ca_divers
    total_remises = agg["remises"] + agg["remises_fid"] + remises_lignes

    # ── Marges ──────────────────────────────────────────────────────────────
    margin_stats = MarginService.calculate_period_margin_with_discounts(
        date_debut=date_debut, date_fin=date_fin,
        factures_qs=factures, exclude_is_divers=False
    )
    margin_pharma_stats = MarginService.calculate_period_margin_with_discounts(
        date_debut=date_debut, date_fin=date_fin,
        factures_qs=factures, exclude_is_divers=True
    )
    marge_totale  = margin_stats["marge_brute"]
    marge_pharma  = margin_pharma_stats["marge_brute"]
    marge_divers  = marge_totale - marge_pharma

    # ── CA & Marges jour par jour ────────────────────────────────────────────
    # Utiliser TruncDate (cohérent avec le calcul des allocations ci-dessous)
    daily_data: dict = {}
    for f in factures_ann.annotate(day=local_trunc_date("date")).values(
        "day", "total_ttc", "divers_amount", "remise", "montant_fidelite"
    ):
        day = f["day"]
        if day not in daily_data:
            daily_data[day] = {"ca_total": Decimal("0"), "ca_divers": Decimal("0"),
                               "remises": Decimal("0"), "nb_ventes": 0}
        daily_data[day]["ca_total"]  += f["total_ttc"] or Decimal("0")
        daily_data[day]["ca_divers"] += f["divers_amount"] or Decimal("0")
        daily_data[day]["remises"]   += (f["remise"] or Decimal("0")) + (f["montant_fidelite"] or Decimal("0"))
        daily_data[day]["nb_ventes"] += 1

    # Marges par jour via allocations
    alloc_day = FactureProduitAllocation.objects.filter(
        facture_produit__facture__in=factures
    ).annotate(
        day=local_trunc_date("facture_produit__facture__date")
    ).values("day", "stock_lot__is_divers").annotate(
        rev=Coalesce(Sum(F("selling_price") * F("quantity"), output_field=DecimalField()), Decimal("0")),
        cost=Coalesce(Sum(F("cost_price") * F("quantity"), output_field=DecimalField()), Decimal("0")),
    )
    for row in alloc_day:
        day = row["day"]
        if day not in daily_data:
            continue  # allocation orpheline (décalage timezone) — ignorer
        marge_ligne = (row["rev"] or Decimal("0")) - (row["cost"] or Decimal("0"))
        if row["stock_lot__is_divers"]:
            daily_data[day]["marge_divers"] = daily_data[day].get("marge_divers", Decimal("0")) + marge_ligne
        else:
            daily_data[day]["marge_pharma"] = daily_data[day].get("marge_pharma", Decimal("0")) + marge_ligne

    daily_rows = []
    for day in sorted(daily_data):
        d = daily_data[day]
        ca_t = float(d["ca_total"])
        mp   = float(d.get("marge_pharma", 0))
        md   = float(d.get("marge_divers", 0))
        mt   = mp + md
        daily_rows.append({
            "date":         day.strftime("%d/%m/%Y"),
            "nb_ventes":    d["nb_ventes"],
            "ca_pharma":    float(d["ca_total"]) - float(d["ca_divers"]),
            "ca_divers":    float(d["ca_divers"]),
            "ca_total":     ca_t,
            "marge_pharma": mp,
            "marge_divers": md,
            "marge_totale": mt,
            "taux_marge":   (mt / ca_t * 100) if ca_t else 0,
            "remises":      float(d["remises"]),
        })

    # ── Remises (détail par facture) ─────────────────────────────────────────
    remises_rows = []
    for f in factures.select_related("client").prefetch_related("produits"):
        rem_lignes = sum(
            (fp.discount or Decimal("0")) * fp.quantity
            for fp in f.produits.all()
        )
        total_rem_f = (f.remise or Decimal("0")) + (f.montant_fidelite or Decimal("0")) + rem_lignes
        if total_rem_f > 0:
            remises_rows.append({
                "date":           f.date.strftime("%d/%m/%Y"),
                "numero_facture": f.numero_facture or f"F-{f.pk}",
                "client":         f.client.name if f.client else (f.client_name_override or "Comptant"),
                "remise_globale": float(f.remise or 0),
                "remise_lignes":  float(rem_lignes),
                "remise_fidelite":float(f.montant_fidelite or 0),
                "total_remise":   float(total_rem_f),
            })

    # ── Avoirs fournisseurs ──────────────────────────────────────────────────
    avoirs_rows = []
    for a in Avoir.objects.filter(
        date__gte=date_debut.date(), date__lt=date_fin.date(), status="VALIDEE"
    ).select_related("fournisseur"):
        avoirs_rows.append({
            "date":        a.date.strftime("%d/%m/%Y") if a.date else "",
            "numero":      a.numero or f"AV-{a.pk}",
            "fournisseur": a.fournisseur.name if a.fournisseur else "—",
            "montant":     float(a.total_ht),
            "statut":      a.status,
        })

    # ── Top Produits ─────────────────────────────────────────────────────────
    alloc_prod = FactureProduitAllocation.objects.filter(
        facture_produit__facture__in=factures
    ).values(
        "facture_produit__produit__id",
        "facture_produit__produit__name",
        "facture_produit__produit__rayon__name",
        "stock_lot__is_divers",
    ).annotate(
        ca_ttc=Coalesce(Sum(F("selling_price") * F("quantity"), output_field=DecimalField()), Decimal("0")),
        cout_achat=Coalesce(Sum(F("cost_price") * F("quantity"), output_field=DecimalField()), Decimal("0")),
        quantite=Coalesce(Sum("quantity"), 0),
    ).order_by("-ca_ttc")[:50]

    top_rows = []
    for row in alloc_prod:
        ca   = float(row["ca_ttc"] or 0)
        cout = float(row["cout_achat"] or 0)
        marge_p = ca - cout
        prod = Produit.objects.filter(pk=row["facture_produit__produit__id"]).values("cip1", "cip2", "cip3").first() or {}
        cip = prod.get("cip1") or prod.get("cip2") or prod.get("cip3") or "—"
        top_rows.append({
            "produit_nom": row["facture_produit__produit__name"] or "—",
            "cip":         cip,
            "rayon":       row["facture_produit__produit__rayon__name"] or "—",
            "is_divers":   row["stock_lot__is_divers"],
            "quantite":    row["quantite"],
            "ca_ttc":      ca,
            "cout_achat":  cout,
            "marge":       marge_p,
            "taux_marge":  (marge_p / ca * 100) if ca else 0,
        })

    # ── Dettes Fournisseurs ──────────────────────────────────────────────────
    from django.db.models import ExpressionWrapper
    comm_sub = CommandeProduit.objects.filter(
        commande__fournisseur=OuterRef("pk"),
        commande__status=Commande.Status.CLOTUREE,
    ).values("commande__fournisseur").annotate(
        t=Sum(F("quantity") * F("price_cost"), output_field=DecimalField())
    ).values("t")
    pay_sub = PaiementFournisseur.objects.filter(
        fournisseur=OuterRef("pk")
    ).values("fournisseur").annotate(
        t=Sum("montant", output_field=DecimalField())
    ).values("t")

    fournisseurs_ann = Fournisseur.objects.filter(is_active=True).annotate(
        total_commande=Coalesce(Subquery(comm_sub[:1]), Value(0, output_field=DecimalField())),
        total_paye=Coalesce(Subquery(pay_sub[:1]), Value(0, output_field=DecimalField())),
        solde_du=ExpressionWrapper(
            F("total_commande") - F("total_paye"), output_field=DecimalField()
        ),
    ).filter(solde_du__gt=0)

    dettes_rows = []
    dettes_total = Decimal("0")
    for f in fournisseurs_ann:
        dettes_total += f.solde_du
        dettes_rows.append({
            "fournisseur":       f.name,
            "total_commande":    float(f.total_commande),
            "total_paye":        float(f.total_paye),
            "solde_du":          float(f.solde_du),
            "echu":              0,
            "a_venir":           float(f.solde_du),
            "prochaine_echeance":"—",
        })

    # ── Créances Clients ─────────────────────────────────────────────────────
    # Filtré sur la période du rapport pour que créances + caisses = CA total mois
    clients_data = Client.objects.filter(
        facture__status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
        facture__date__gte=date_debut,
        facture__date__lt=date_fin,
    ).distinct().annotate(
        nb_factures=Count("facture", filter=Q(
            facture__status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
            facture__date__gte=date_debut,
            facture__date__lt=date_fin,
        )),
    )
    creances_rows = []
    for c in clients_data:
        ca_c = Facture.objects.filter(
            client=c,
            status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
            date__gte=date_debut,
            date__lt=date_fin,
        ).aggregate(t=Coalesce(Sum("total_ttc"), Decimal("0")))["t"]
        paye_c = Caisse.objects.filter(
            facture__client=c,
            facture__date__gte=date_debut,
            facture__date__lt=date_fin,
            statut="completee",
        ).exclude(mode_paiement="en_compte").aggregate(
            t=Coalesce(Sum("montant"), Decimal("0"))
        )["t"]
        reste = ca_c - paye_c
        if reste > Decimal("0.5"):
            creances_rows.append({
                "client_nom":    c.name,
                "client_type":   c.client_type if hasattr(c, "client_type") else "—",
                "nb_factures":   c.nb_factures,
                "ca_total":      float(ca_c),
                "montant_paye":  float(paye_c),
                "reste_a_payer": float(reste),
                "jours_retard":  0,
            })
    creances_rows.sort(key=lambda x: x["reste_a_payer"], reverse=True)

    # ── Stock par rayon ──────────────────────────────────────────────────────
    from django.utils import timezone as tz
    stock_rows = []
    today = tz.now().date()

    # Rayons nommés + rayon None (produits sans rayon)
    rayons_list = list(Rayon.objects.all().order_by("name")) + [None]
    for rayon in rayons_list:
        prods = Produit.objects.filter(rayon=rayon, is_active=True)
        nb_refs      = prods.count()
        if nb_refs == 0:
            continue
        qte_totale   = prods.aggregate(t=Coalesce(Sum("stock"), 0))["t"]
        valeur_pmp   = prods.aggregate(
            t=Coalesce(Sum(F("stock") * F("pmp"), output_field=DecimalField()), Decimal("0"))
        )["t"]
        alertes_rupt = prods.filter(stock__lte=F("stock_minimum")).count()
        alertes_per  = StockLot.objects.filter(
            produit__rayon=rayon,
            quantity_remaining__gt=0,
            date_expiration__lte=today + timedelta(days=30),
            date_expiration__gt=today,
        ).count()
        stock_rows.append({
            "rayon":              rayon.name if rayon else "(Sans rayon)",
            "nb_refs":            nb_refs,
            "qte_totale":         qte_totale,
            "valeur_pmp":         float(valeur_pmp),
            "stock_minimum":      prods.aggregate(t=Coalesce(Sum("stock_minimum"), 0))["t"],
            "alertes_rupture":    alertes_rupt,
            "alertes_peremption": alertes_per,
        })

    # ── Achats fournisseurs ──────────────────────────────────────────────────
    achats_map: dict = {}
    total_precompte_excel = Decimal("0")
    for c in Commande.objects.filter(
        date__gte=date_debut, date__lt=date_fin,
        status=Commande.Status.CLOTUREE,
    ).prefetch_related("produits").select_related("fournisseur"):
        if not c.fournisseur:
            continue
        fid = c.fournisseur.id
        if fid not in achats_map:
            achats_map[fid] = {
                "fournisseur_nom": c.fournisseur.name,
                "nb_commandes": 0, "montant_total": Decimal("0"),
                "nb_avoirs": 0, "montant_avoirs": Decimal("0"),
                "precompte": Decimal("0"),
            }
        achats_map[fid]["montant_total"] += sum(
            cp.quantity * cp.price_cost for cp in c.produits.all()
        )
        achats_map[fid]["nb_commandes"] += 1
        cmd_precompte = c.precompte
        achats_map[fid]["precompte"] += cmd_precompte
        total_precompte_excel += cmd_precompte
    for a in Avoir.objects.filter(
        date__gte=date_debut.date(), date__lt=date_fin.date(), status="VALIDEE"
    ).select_related("fournisseur"):
        if not a.fournisseur:
            continue
        fid = a.fournisseur.id
        if fid not in achats_map:
            achats_map[fid] = {
                "fournisseur_nom": a.fournisseur.name,
                "nb_commandes": 0, "montant_total": Decimal("0"),
                "nb_avoirs": 0, "montant_avoirs": Decimal("0"),
                "precompte": Decimal("0"),
            }
        achats_map[fid]["montant_total"] -= a.total_ht
        achats_map[fid]["montant_avoirs"] += a.total_ht
        achats_map[fid]["nb_avoirs"] += 1
    achats_rows = sorted(achats_map.values(), key=lambda x: x["montant_total"], reverse=True)

    # ── État des caisses ──────────────────────────────────────────────────────
    clotures_rows = []
    for cl in ClotureCaisse.objects.filter(
        date__gte=date_debut, date__lt=date_fin,
    ).select_related("user", "poste_caisse").order_by("date"):
        dt = timezone.localtime(cl.date) if timezone.is_aware(cl.date) else cl.date
        h_ouv = cl.date_debut
        h_fer = cl.date_fin
        if cl.user:
            caissier = cl.user.get_full_name().strip() or cl.user.username
        else:
            caissier = "Inconnu"
        clotures_rows.append({
            "date":             dt.strftime("%d/%m/%Y"),
            "caisse":           (cl.poste_caisse.nom if cl.poste_caisse else "Caisse") +
                                (f" — {caissier}" if caissier else ""),
            "caissier":         caissier,
            "heure_ouverture":  timezone.localtime(h_ouv).strftime("%H:%M") if h_ouv and timezone.is_aware(h_ouv) else (str(h_ouv or "")),
            "heure_fermeture":  timezone.localtime(h_fer).strftime("%H:%M") if h_fer and timezone.is_aware(h_fer) else (str(h_fer or "")),
            "total_ventes":     float(cl.total_ventes or 0),
            "total_entrees":    float(cl.total_entrees or 0),
            "total_sorties":    float(cl.total_sorties or 0),
            "montant_theorique":float(cl.montant_theorique or 0),
            "montant_reel":     float(cl.montant_reel or 0),
            "ecart_caisse":     float(cl.ecart_caisse or 0),
            "observation":      cl.observation or "",
        })

    # ── Dépenses (MouvementCaisse) ───────────────────────────────────────────
    depenses_rows = []
    for m in MouvementCaisse.objects.filter(
        date__gte=date_debut, date__lt=date_fin,
    ).select_related("user", "poste_caisse").order_by("date"):
        dt = timezone.localtime(m.date) if timezone.is_aware(m.date) else m.date
        depenses_rows.append({
            "date":        dt,
            "caisse":      m.poste_caisse.nom if m.poste_caisse else "—",
            "type":        m.type,
            "motif":       m.motif,
            "description": m.description or "",
            "montant":     float(m.montant),
            "user":        m.user.get_full_name() if m.user else "Inconnu",
        })

    return {
        "ca": {
            "ca_ttc":       ca_total,
            "total_remises": total_remises,
            "total_remises_detail": {
                "global":   agg["remises"],
                "fidelite": agg["remises_fid"],
                "lignes":   remises_lignes,
            },
            "nb_ventes":    agg["nb_ventes"],
        },
        "divers_ca": {"ca_ttc": ca_divers},
        "marge": {
            "marge_brute":  marge_totale,
            "marge_pharma": marge_pharma,
            "marge_divers": marge_divers,
            "marge_pct":    float(marge_totale / ca_total * 100) if ca_total else 0,
        },
        "creances": {
            "total":       sum(Decimal(str(r["reste_a_payer"])) for r in creances_rows),
            "nb_factures": len(creances_rows),
        },
        "dettes_total":      dettes_total,
        "daily_rows":        daily_rows,
        "remises_rows":      remises_rows,
        "avoirs_rows":       avoirs_rows,
        "top_rows":          top_rows,
        "dettes_rows":       dettes_rows,
        "creances_rows":     creances_rows,
        "stock_rows":        stock_rows,
        "achats_rows":       achats_rows,
        "clotures_rows":     clotures_rows,
        "depenses_rows":     depenses_rows,
        "ca": {
            "ca_ht": float(agg["ca_total_ht"]),
            "ca_ttc": float(ca_total),
            "tva_collectee": float(ca_total - agg["ca_total_ht"]) if ca_total else 0,
        },
        "achats_fiscaux": {
            "total_achats_ht": int(sum(Decimal(str(r.get("montant_total", 0) or 0)) for r in achats_rows).quantize(Decimal("1"), rounding=ROUND_HALF_UP)),
            "total_precompte": int(total_precompte_excel.quantize(Decimal("1"), rounding=ROUND_HALF_UP)) if total_precompte_excel else 0,
        },
        "accompte_fiscal": _compute_accompte_fiscal_inline(agg["ca_total_ht"], factures_ann),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Point d'entrée public
# ─────────────────────────────────────────────────────────────────────────────

def build_rapport_general_excel(date_debut, date_fin, mois_label: str) -> HttpResponse:
    """
    Construit le rapport Excel général et retourne un HttpResponse prêt à télécharger.
    """
    from api.models.settings import PharmacySettings

    # Infos pharmacie
    try:
        ps = PharmacySettings.objects.get(pk=1)
        pharmacy = {
            "name":    ps.pharmacy_name,
            "address": ps.address,
            "city":    ps.city,
            "phone":   ps.phone,
            "email":   ps.email,
        }
        logo_path = ps.logo.path if ps.logo else None
    except Exception:
        pharmacy = {"name": "PHARMACIE", "address": "", "city": "", "phone": "", "email": ""}
        logo_path = None

    from api.views.rapports.excel_general_extra import (
        collect_extra_data, enrich_synthese,
        sheet_modes_paiement, sheet_retours_annulations,
        sheet_performance_vendeurs, sheet_tresorerie,
        sheet_perimes, sheet_promotions, sheet_clients_pro,
    )

    data  = _collect_data(date_debut, date_fin)
    extra = collect_extra_data(date_debut, date_fin)

    wb = openpyxl.Workbook()
    if wb.active is not None:
        wb.remove(wb.active)

    # ── Feuilles originales (1-10) ──
    _sheet_synthese(wb, data, pharmacy, mois_label, logo_path)
    # Enrichir la Synthèse avec évolution + objectif
    enrich_synthese(wb["Synthèse"], extra)

    _sheet_ca_marges(wb, data["daily_rows"], pharmacy, mois_label, logo_path)
    _sheet_remises_avoirs(wb, data["remises_rows"], data["avoirs_rows"], pharmacy, mois_label, logo_path)
    _sheet_top_produits(wb, data["top_rows"], pharmacy, mois_label, logo_path)
    _sheet_dettes_fournisseurs(wb, data["dettes_rows"], pharmacy, mois_label, logo_path)
    _sheet_creances_clients(wb, data["creances_rows"], pharmacy, mois_label, logo_path)
    _sheet_stock(wb, data["stock_rows"], pharmacy, mois_label, logo_path)
    _sheet_achats(wb, data["achats_rows"], pharmacy, mois_label, logo_path)
    _sheet_etat_caisses(wb, data["clotures_rows"], pharmacy, mois_label, logo_path)
    _sheet_depenses(wb, data["depenses_rows"], pharmacy, mois_label, logo_path)

    # ── Feuille fiscale ──
    _sheet_fiscal(wb, data, pharmacy, mois_label, logo_path)

    # ── Feuille UGs ──
    _sheet_ugs(wb, data, pharmacy, mois_label, logo_path)

    # ── Nouvelles feuilles (11-17) ──
    sheet_modes_paiement(wb, extra, pharmacy, mois_label, logo_path)
    sheet_retours_annulations(wb, extra, pharmacy, mois_label, logo_path)
    sheet_performance_vendeurs(wb, extra, pharmacy, mois_label, logo_path)
    sheet_tresorerie(wb, extra, pharmacy, mois_label, logo_path)
    sheet_perimes(wb, extra, pharmacy, mois_label, logo_path)
    sheet_promotions(wb, extra, pharmacy, mois_label, logo_path)
    sheet_clients_pro(wb, extra, pharmacy, mois_label, logo_path)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Rapport_Pharmacie_{mois_label.replace(' ', '_')}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
