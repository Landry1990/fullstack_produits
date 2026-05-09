"""
Rapports financiers, comptables et analyse de TVA — RapportFinanceMixin.
"""
import csv
import json
from datetime import datetime, time, timedelta
from decimal import Decimal

import openpyxl
from django.db.models import Count, Exists, F, OuterRef, Q, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Font
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response

from api.models import Caisse, Facture, FactureProduit, FactureProduitAllocation, StockLot
from api.views.rapports.base import RapportBaseMixin
from api.views.rapports.pdf_builders import build_rapport_pdf


# ── Helpers privés ────────────────────────────────────────────────────────────

def _write_pharma_header(ws, PharmacySettings, title: str) -> None:
    """Écrit un en-tête pharmacie dans une feuille Excel."""
    from django.utils import timezone as tz
    try:
        pharmacy = PharmacySettings.objects.get(pk=1)
        pharma_name    = pharmacy.pharmacy_name or "ZENITH"
        pharma_address = (
            f"{pharmacy.address} - {pharmacy.city}".strip(" -")
            if (pharmacy.address or pharmacy.city) else ""
        )
        pharma_phone = f"Tél : {pharmacy.phone}" if pharmacy.phone else ""
    except Exception:
        pharma_name, pharma_address, pharma_phone = "ZENITH", "", ""

    now_time = tz.now()
    if tz.is_aware(now_time):
        now_time = tz.localtime(now_time)
    now_str = now_time.strftime("%d/%m/%Y à %H:%M")
    for line in [pharma_name, pharma_address, pharma_phone, f"Édité le : {now_str}", "", title]:
        ws.append([line])
    ws.append([])


def _parse_day_range(request):
    """
    Retourne (date_debut, date_fin_exclusive) depuis les query params
    date_debut / date_fin (format YYYY-MM-DD).
    Lève ValueError si les params sont absents ou mal formés.
    """
    db_s = request.query_params.get('date_debut')
    df_s = request.query_params.get('date_fin')
    if not db_s or not df_s:
        raise ValueError('date_debut et date_fin requis')
    date_debut = timezone.make_aware(
        datetime.combine(datetime.strptime(db_s, '%Y-%m-%d').date(), time.min)
    )
    date_fin_exclusive = timezone.make_aware(
        datetime.combine(
            datetime.strptime(df_s, '%Y-%m-%d').date() + timedelta(days=1),
            time.min,
        )
    )
    return db_s, df_s, date_debut, date_fin_exclusive


def _parse_day_range_inclusive(request):
    """
    Retourne (date_debut, date_fin inclusive à 23:59:59).
    Utilisé par rapport_remises et rapport_detail_marges.
    """
    db_s = request.query_params.get('date_debut')
    df_s = request.query_params.get('date_fin')
    if not db_s or not df_s:
        raise ValueError('date_debut et date_fin requis')
    date_debut = timezone.make_aware(datetime.combine(datetime.strptime(db_s, '%Y-%m-%d'), time.min))
    date_fin   = timezone.make_aware(datetime.combine(datetime.strptime(df_s, '%Y-%m-%d'), time.max))
    return db_s, df_s, date_debut, date_fin


def _group_results(rows: list, group_key: str) -> list:
    """
    Agrège une liste de dicts par la valeur de `group_key`.
    - Les colonnes numériques sont sommées.
    - Les colonnes textuelles (non-numériques) sont ignorées dans les lignes agrégées
      sauf `group_key` lui-même qui devient la clé de groupe.
    - Une colonne '_count' est ajoutée pour indiquer le nombre de lignes fusionnées.
    """
    from collections import defaultdict
    groups: dict = defaultdict(lambda: {'_count': 0})

    for row in rows:
        key_val = row.get(group_key, 'N/A')
        bucket  = groups[key_val]
        bucket['_count'] += 1
        bucket[group_key] = key_val
        for col, val in row.items():
            if col == group_key:
                continue
            try:
                num = float(val)
                bucket[col] = round(float(bucket.get(col, 0)) + num, 2)  # type: ignore[assignment]
            except (TypeError, ValueError):
                # colonne textuelle : on garde la valeur seulement si identique pour tout le groupe
                if col not in bucket:
                    bucket[col] = val
                elif bucket[col] != val:
                    bucket[col] = '—'  # type: ignore  # valeurs hétérogènes dans le groupe

    # Réordonner : group_key en premier, _count en dernier
    result = []
    for bucket in groups.values():
        ordered = {group_key: bucket[group_key]}
        for k, v in bucket.items():
            if k not in (group_key, '_count'):
                ordered[k] = v
        ordered['_count'] = bucket['_count']
        result.append(ordered)
    return result


# ── Mixin ─────────────────────────────────────────────────────────────────────

class RapportFinanceMixin:
    """Rapports financiers, comptables et analyse de TVA."""

    # ── Rapports JSON ─────────────────────────────────────────────────────────

    @action(detail=False, methods=['get'])
    def rapport_mensuel(self, request):
        mois = request.query_params.get('mois')
        if not mois:
            return Response({'detail': 'Mois requis'}, status=400)
        try:
            date_debut = timezone.make_aware(datetime.strptime(f"{mois}-01", '%Y-%m-%d'))
            date_fin   = (date_debut + timedelta(days=32)).replace(day=1)
        except ValueError:
            return Response({'detail': 'Format mois invalide (YYYY-MM)'}, status=400)
        return Response(self._get_rapport_data(date_debut, date_fin, mois))

    @action(detail=False, methods=['get'])
    def rapport_par_dates(self, request):
        """Rapport complet sur une tranche de dates arbitraire."""
        try:
            db_s, df_s, date_debut, date_fin_exclusive = _parse_day_range(request)
        except ValueError as e:
            return Response({'detail': str(e)}, status=400)
        data = self._get_rapport_data(date_debut, date_fin_exclusive, f"{db_s} → {df_s}")
        data['date_debut'] = db_s
        data['date_fin']   = df_s
        return Response(data)

    # ── Rapports PDF ──────────────────────────────────────────────────────────

    @action(detail=False, methods=['get'])
    def rapport_mensuel_pdf(self, request):
        mois = request.query_params.get('mois')
        if not mois:
            return Response({'detail': 'Mois requis'}, status=400)
        date_debut = timezone.make_aware(datetime.strptime(f"{mois}-01", '%Y-%m-%d'))
        date_fin   = (date_debut + timedelta(days=32)).replace(day=1)
        data       = self._get_rapport_data(date_debut, date_fin, mois)
        return build_rapport_pdf(data, f"RAPPORT MENSUEL — {mois}", f"rapport_{mois}.pdf")

    @action(detail=False, methods=['get'])
    def rapport_par_dates_pdf(self, request):
        """PDF du rapport sur une tranche de dates arbitraire."""
        try:
            db_s, df_s, date_debut, date_fin_exclusive = _parse_day_range(request)
        except ValueError as e:
            return Response({'detail': str(e)}, status=400)
        data = self._get_rapport_data(date_debut, date_fin_exclusive, f"{db_s} → {df_s}")
        return build_rapport_pdf(
            data,
            f"RAPPORT D'ACTIVITÉ — {db_s} au {df_s}",
            f"rapport_{db_s}_{df_s}.pdf",
        )

    # ── CA multi-annuel ───────────────────────────────────────────────────────

    @action(detail=False, methods=['get'])
    def rapport_ca_multi_annuel(self, request):
        annees = [
            d.year for d in Facture.objects
            .filter(status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE])
            .exclude(produits__allocations__stock_lot__is_divers=True)
            .dates('date', 'year', order='DESC')
        ]
        if not annees:
            return Response([])

        m_keys  = ['january', 'february', 'march', 'april', 'may', 'june',
                   'july', 'august', 'september', 'october', 'november', 'december']
        results = [{'Mois': k, '_index': i + 1} for i, k in enumerate(m_keys)]
        totaux  = {'Mois': 'total_general', '_index': 13}

        for annee in sorted(annees):
            at_tva = at_exo = Decimal('0.00')
            for m_idx in range(1, 13):
                date_debut = timezone.make_aware(datetime(annee, m_idx, 1))
                date_fin   = (date_debut + timedelta(days=32)).replace(day=1)
                ca_tva = ca_exo = Decimal('0.00')
                for item in self._calculate_ca_par_tva(self._get_factures_periode(date_debut, date_fin)):
                    if item['taux'] > 0:
                        ca_tva += item['ca_ttc']
                    else:
                        ca_exo += item['ca_ttc']
                row = results[m_idx - 1]
                row[f"{annee}_ca_tva"]  = ca_tva
                row[f"{annee}_ca_exo"]  = ca_exo
                row[f"{annee}_total"]   = ca_tva + ca_exo
                at_tva += ca_tva
                at_exo += ca_exo
            totaux[f"{annee}_ca_tva"] = at_tva
            totaux[f"{annee}_ca_exo"] = at_exo
            totaux[f"{annee}_total"]  = at_tva + at_exo

        results.append(totaux)
        results.sort(key=lambda x: x['_index'])
        for r in results:
            del r['_index']
        return Response(results)

    # ── TVA sur vendus ────────────────────────────────────────────────────────

    @action(detail=False, methods=['get'])
    def rapport_tva_vendus(self, request):
        db_str = request.query_params.get('date_debut')
        df_str = request.query_params.get('date_fin')
        if not db_str or not df_str:
            return Response({'error': 'Dates requises'}, status=400)
        try:
            date_debut = timezone.make_aware(datetime.fromisoformat(db_str.replace('Z', '+00:00')))
            date_fin   = timezone.make_aware(datetime.fromisoformat(df_str.replace('Z', '+00:00')))
            if date_fin.hour == 0:
                date_fin += timedelta(days=1)
        except ValueError:
            return Response({'error': 'Format de date invalide'}, status=400)

        lignes = (
            FactureProduit.objects
            .filter(
                facture__date__range=(date_debut, date_fin),
                facture__status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
                tva__gt=0,
            )
            .exclude(facture__produits__allocations__stock_lot__is_divers=True)
            .values('produit__name', 'produit__cip1', 'tva')
            .annotate(
                total_qty=Sum('quantity'),
                total_ttc=Sum(F('quantity') * (F('selling_price') - F('discount'))),
            )
            .order_by('produit__name')
        )
        data = []
        for ligne in lignes:
            tva = ligne['tva'] or 0
            ttc = ligne['total_ttc'] or 0
            mt_tva = (ttc * tva) / (100 + tva) if tva > 0 else 0
            data.append({
                'produit':     ligne['produit__name'],
                'cip':         ligne['produit__cip1'],
                'quantite':    ligne['total_qty'],
                'taux_tva':    f"{float(tva)} %",
                'total_ttc':   round(ttc, 0),
                'montant_tva': round(mt_tva, 0),
            })
        return Response(data)

    # ── Export comptable CSV ──────────────────────────────────────────────────

    @action(detail=False, methods=['get'])
    def export_comptable_csv(self, request):
        db_str = request.query_params.get('date_debut')
        df_str = request.query_params.get('date_fin')
        try:
            date_debut = timezone.make_aware(datetime.fromisoformat(db_str.replace('Z', '+00:00')))
            date_fin   = timezone.make_aware(datetime.fromisoformat(df_str.replace('Z', '+00:00')))
            if date_fin.hour == 0:
                date_fin += timedelta(days=1)
        except (ValueError, AttributeError):
            return Response({'error': 'Date invalide'}, status=status.HTTP_400_BAD_REQUEST)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="export_comptable.csv"'
        response.write('\ufeff'.encode('utf8'))

        writer = csv.writer(response, delimiter=';')
        writer.writerow(['Date', 'Heure', 'Facture #', 'Client', 'Status',
                         'Total HT', 'Total TVA', 'Total TTC', 'Remise',
                         'Mode de Paiement', 'Caissier'])

        factures = (
            Facture.objects
            .filter(date__range=(date_debut, date_fin),
                    status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE])
            .exclude(produits__allocations__stock_lot__is_divers=True)
            .select_related('client', 'created_by')
            .prefetch_related('paiements')
        )
        modes_dict = dict(Caisse.MODES_PAIEMENT)
        for f in factures:
            modes = ", ".join(
                str(modes_dict.get(m, m))
                for m in f.paiements.filter(statut='completee')
                .values_list('mode_paiement', flat=True).distinct()
            )
            writer.writerow([
                f.date.strftime('%d/%m/%Y'),
                f.date.strftime('%H:%M'),
                f.numero_facture or f.id,
                f.client.name if f.client else 'Passage',
                f.get_status_display(),
                str(f.total_ht).replace('.', ','),
                str(f.total_tva).replace('.', ','),
                str(f.total_ttc).replace('.', ','),
                str(f.remise).replace('.', ','),
                modes,
                f.created_by.get_full_name() if f.created_by else 'Système',
            ])
        return response

    # ── Remises par vendeur ───────────────────────────────────────────────────

    @action(detail=False, methods=['get'])
    def rapport_remises(self, request):
        try:
            _, _, date_debut, date_fin = _parse_day_range_inclusive(request)
        except ValueError as e:
            return Response({"error": str(e)}, status=400)

        factures = (
            Facture.objects
            .filter(status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
                    date__gte=date_debut, date__lte=date_fin)
            .exclude(produits__allocations__stock_lot__is_divers=True)
            .select_related('validated_by')
        )
        stats = (
            factures
            .values('validated_by__id', 'validated_by__username',
                    'validated_by__first_name', 'validated_by__last_name')
            .annotate(
                remise_globale  = Coalesce(Sum('remise'),          Decimal('0.00')),
                remise_fidelite = Coalesce(Sum('montant_fidelite'), Decimal('0.00')),
                ca_ttc          = Coalesce(Sum('total_ttc'),        Decimal('0.00')),
                nb_factures     = Count('id'),
            )
            .order_by('-remise_globale')
        )
        line_remises = {
            s['facture__validated_by__id']: s['remise_lignes']
            for s in (
                FactureProduit.objects
                .filter(facture__in=factures)
                .values('facture__validated_by__id')
                .annotate(remise_lignes=Coalesce(Sum(F('discount') * F('quantity')), Decimal('0.00')))
            )
        }
        res = []
        for s in stats:
            uid = s['validated_by__id']
            rl  = line_remises.get(uid, Decimal('0.00'))
            total = s['remise_globale'] + s['remise_fidelite'] + rl
            res.append({
                'user_id':          uid,
                'username':         s['validated_by__username'],
                'full_name':        f"{s['validated_by__first_name'] or ''} {s['validated_by__last_name'] or ''}".strip(),
                'nb_factures':      s['nb_factures'],
                'ca_ttc':           s['ca_ttc'],
                'remise_globale':   s['remise_globale'],
                'remise_lignes':    rl,
                'remise_fidelite':  s['remise_fidelite'],
                'total_remise':     total,
                'ratio_remise_pct': float(total / s['ca_ttc'] * 100) if s['ca_ttc'] > 0 else 0,
            })
        return Response(res)

    @action(detail=False, methods=['get'])
    def rapport_remises_details(self, request):
        """Liste détaillée des factures avec remises sur la période."""
        try:
            _, _, date_debut, date_fin = _parse_day_range_inclusive(request)
        except ValueError as e:
            return Response({"error": str(e)}, status=400)

        factures = (
            Facture.objects
            .filter(status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
                    date__gte=date_debut, date__lte=date_fin)
            .exclude(produits__allocations__stock_lot__is_divers=True)
            .select_related('client', 'validated_by')
            .prefetch_related('produits')
        )
        results = []
        for f in factures:
            remise_globale  = float(f.remise or 0)
            remise_fidelite = float(f.montant_fidelite or 0)
            remise_lignes   = sum(float(l.discount or 0) * l.quantity for l in f.produits.all())
            total_remise    = remise_globale + remise_fidelite + remise_lignes
            if total_remise <= 0:
                continue
            results.append({
                'numero_facture':  f.numero_facture or f'#{f.id}',
                'date':            f.date.strftime('%d/%m/%Y %H:%M'),
                'client':          f.client.name if f.client else 'Passage',
                'total_ttc':       float(f.total_ttc or 0),
                'remise_globale':  remise_globale,
                'remise_lignes':   round(remise_lignes, 2),
                'remise_fidelite': remise_fidelite,
                'total_remise':    round(total_remise, 2),
                'ratio_remise_pct': (
                    round(total_remise / float(f.total_ttc) * 100, 2)
                    if f.total_ttc and f.total_ttc > 0 else 0
                ),
                'vendeur': (
                    f.validated_by.get_full_name() or f.validated_by.username
                    if f.validated_by else 'N/A'
                ),
            })
        results.sort(key=lambda x: -x['total_remise'])
        return Response(results)

    @action(detail=False, methods=['get'])
    def rapport_remises_details_excel(self, request):
        from api.models import PharmacySettings
        data = self.rapport_remises_details(request).data or []
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Détails Remises"
        _write_pharma_header(ws, PharmacySettings, "Détail des Remises par Facture")
        ws.append(["Facture", "Date", "Client", "Total TTC", "Remise Globale",
                   "Remise Lignes", "Remise Fidélité", "Total Remise", "% / CA", "Vendeur"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        for item in data:
            ws.append([
                item['numero_facture'], item['date'], item['client'],
                item['total_ttc'], item['remise_globale'], item['remise_lignes'],
                item['remise_fidelite'], item['total_remise'],
                f"{item['ratio_remise_pct']:.2f}%", item['vendeur'],
            ])
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Details_Remises.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def rapport_remises_excel(self, request):
        from api.models import PharmacySettings
        data = self.rapport_remises(request).data or []
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Remises"
        _write_pharma_header(ws, PharmacySettings, "Rapport des Remises par Utilisateur")
        ws.append(["Utilisateur", "Nb Factures", "CA TTC", "Remise Globale",
                   "Remise Lignes", "Remise Fidélité", "Total Remise", "% / CA"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        for item in data:
            ws.append([
                item['full_name'], item['nb_factures'], item['ca_ttc'],
                item['remise_globale'], item['remise_lignes'], item['remise_fidelite'],
                item['total_remise'], f"{item['ratio_remise_pct']:.2f}%",
            ])
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Remises.xlsx"'
        wb.save(response)
        return response

    # ── Marges détaillées ─────────────────────────────────────────────────────

    @action(detail=False, methods=['get'])
    def rapport_detail_marges(self, request):
        """
        Rapport détaillé des marges par produit et par lot.
        Utilise le coût exact du lot ou le PMP si non alloué.
        ?grouper_par=produit → agrège par produit (vue pertes/gains)
        """
        try:
            _, _, date_debut, date_fin = _parse_day_range_inclusive(request)
        except ValueError as e:
            return Response({"error": str(e)}, status=400)

        factures = Facture.objects.filter(
            date__range=(date_debut, date_fin),
            status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
        ).exclude(produits__allocations__stock_lot__is_divers=True)
        results = []

        # 1. Lignes allouées (lot connu)
        allocations = (
            FactureProduitAllocation.objects
            .filter(facture_produit__facture__in=factures)
            .select_related(
                'facture_produit', 'facture_produit__facture',
                'facture_produit__produit', 'stock_lot',
            )
            .order_by('-facture_produit__facture__date')
        )
        for alloc in allocations:
            item     = alloc.facture_produit
            f        = item.facture
            p        = item.produit
            qty      = float(alloc.quantity)
            price    = float(item.selling_price - item.discount)
            cost     = float(alloc.cost_price)
            mt_vente = round(price * qty, 2)
            mt_achat = round(cost * qty, 2)
            marge    = round(mt_vente - mt_achat, 2)
            results.append({
                'date':           f.date.strftime('%d/%m/%Y'),
                'facture':        f.numero_facture or f'#{f.id}',
                'produit':        p.name,
                'lot':            alloc.stock_lot.lot if alloc.stock_lot else 'N/A',
                'quantite':       qty,
                'prix_vente_net': round(price, 2),
                'cout_achat':     round(cost, 2),
                'mt_vente':       mt_vente,
                'mt_achat':       mt_achat,
                'marge':          marge,
                'taux_marge':     round(marge / mt_vente * 100, 1) if mt_vente > 0 else 0,
            })

        # 2. Lignes non allouées (fallback dernier lot reçu)
        unallocated = (
            FactureProduit.objects
            .filter(facture__in=factures)
            .annotate(has_alloc=Exists(
                FactureProduitAllocation.objects.filter(facture_produit=OuterRef('pk'))
            ))
            .filter(has_alloc=False)
            .select_related('facture', 'produit')
        )
        # Pré-charger le dernier lot reçu par produit pour éviter N+1
        produit_ids = list(unallocated.values_list('produit_id', flat=True).distinct())
        last_lot_cost: dict[int, float] = {}
        for lot in (
            StockLot.objects
            .filter(produit_id__in=produit_ids)
            .order_by('produit_id', '-date_reception')
            .distinct('produit_id')
            .values('produit_id', 'price_cost')
        ):
            last_lot_cost[lot['produit_id']] = float(lot['price_cost'])

        for item in unallocated:
            f        = item.facture
            p        = item.produit
            qty      = float(item.quantity)
            price    = float(item.selling_price - item.discount)
            cost     = last_lot_cost.get(p.id, float(p.pmp or 0))
            mt_vente = round(price * qty, 2)
            mt_achat = round(cost * qty, 2)
            marge    = round(mt_vente - mt_achat, 2)
            results.append({
                'date':           f.date.strftime('%d/%m/%Y'),
                'facture':        f.numero_facture or f'#{f.id}',
                'produit':        p.name,
                'lot':            'SANS LOT',
                'quantite':       qty,
                'prix_vente_net': round(price, 2),
                'cout_achat':     round(cost, 2),
                'mt_vente':       mt_vente,
                'mt_achat':       mt_achat,
                'marge':          marge,
                'taux_marge':     round(marge / mt_vente * 100, 1) if mt_vente > 0 else 0,
            })
        # --- Filtre marge avant pagination ---
        filtre_marge = request.query_params.get('filtre_marge', '')
        if filtre_marge == 'negative':
            results = [r for r in results if r['taux_marge'] < 0]
        elif filtre_marge == 'low':
            results = [r for r in results if r['taux_marge'] < 25]

        # --- Regroupement par produit si demandé ---
        if request.query_params.get('grouper_par') == 'produit':
            grouped_map: dict[str, dict] = {}
            for r in results:
                key = r['produit']
                if key not in grouped_map:
                    grouped_map[key] = {'nb_ventes': 0, 'quantite': 0.0, 'mt_vente': 0.0, 'mt_achat': 0.0}
                grouped_map[key]['nb_ventes'] += 1
                grouped_map[key]['quantite']  += float(r['quantite'])
                grouped_map[key]['mt_vente']  += float(r['mt_vente'])
                grouped_map[key]['mt_achat']  += float(r['mt_achat'])

            aggregated = []
            for produit_name, g in grouped_map.items():
                mt_v  = round(g['mt_vente'], 2)
                mt_a  = round(g['mt_achat'], 2)
                marge = round(mt_v - mt_a, 2)
                aggregated.append({
                    'produit':    produit_name,
                    'nb_ventes':  g['nb_ventes'],
                    'quantite':   round(g['quantite'], 2),
                    'mt_vente':   mt_v,
                    'mt_achat':   mt_a,
                    'marge':      marge,
                    'taux_marge': round(marge / mt_v * 100, 1) if mt_v > 0 else 0,
                    'statut':     'PERTE' if marge < 0 else ('FAIBLE' if mt_v > 0 and marge / mt_v * 100 < 25 else 'OK'),
                })
            aggregated.sort(key=lambda x: float(x['marge']))
            page = self.paginator.paginate_queryset(aggregated, request)
            if page is not None:
                return self.paginator.get_paginated_response(page)
            return Response(aggregated)

        page = self.paginator.paginate_queryset(results, request)
        if page is not None:
            return self.paginator.get_paginated_response(page)
        return Response(results)

    # ── Rapport dynamique ─────────────────────────────────────────────────────

    # Mapping champ logique → champ ORM par source
    _DYNAMIC_FIELD_MAP = {
        'quantite':      {'ventes': 'quantity',       'achats': 'quantity',    'stock': 'quantity_remaining', 'produits': 'stock'},
        'total_ht':      {'stock':  'price_cost'},
        'prix_vente':    {'ventes': 'selling_price',  'produits': 'selling_price'},
        'cout_achat':    {'ventes': 'selling_price',  'achats': 'price_cost',  'stock': 'price_cost',         'produits': 'pmp'},
        'tva':           {'ventes': 'produit__tva',   'achats': 'produit__tva','stock': 'produit__tva',       'produits': 'tva'},
        'stock_minimum': {'produits': 'stock_minimum'},
        'cip':           {'ventes': 'produit__cip1',  'achats': 'produit__cip1','stock': 'produit__cip1',     'produits': 'cip1'},
    }

    _DYNAMIC_LABEL_MAP = {
        'pourcentage_marge': ['Marge (%)'],
        'tva':               ['TVA (%)'],
        'quantite':          ['Quantité'],
        'total_ht':          ['Total HT'],
        'prix_vente':        ['Prix Vente'],
        'cout_achat':        ['Coût Achat', 'P.U Achat'],
    }

    @action(detail=False, methods=['get'])
    def rapport_dynamique(self, request):
        try:
            return self._rapport_dynamique(request)
        except Exception as e:
            return Response({"error": f"Erreur interne: {e}"}, status=500)

    def _rapport_dynamique(self, request):
        source         = request.query_params.get('source', 'ventes')
        db_s           = request.query_params.get('date_debut')
        df_s           = request.query_params.get('date_fin')
        vendeur_id     = request.query_params.get('vendeur_id')
        client_id      = request.query_params.get('client_id')
        fournisseur_id = request.query_params.get('fournisseur_id')
        famille_id     = request.query_params.get('famille_id')
        requested_fields = request.query_params.get('fields', '').split(',')
        logic          = request.query_params.get('logic', 'AND').upper()
        sort_by        = request.query_params.get('sort_by')          # label de colonne, ex: 'Total HT'
        sort_order     = request.query_params.get('sort_order', 'desc').lower()  # 'asc' | 'desc'
        group_by       = request.query_params.get('group_by')         # label de colonne, ex: 'Produit'

        try:
            conditions = json.loads(request.query_params.get('conditions', '[]'))
        except (ValueError, TypeError):
            conditions = []

        if source != 'produits' and (not db_s or not df_s):
            return Response({"error": "Dates requises"}, status=400)

        date_debut = date_fin = None
        if db_s and df_s:
            try:
                date_debut = timezone.make_aware(datetime.combine(datetime.strptime(db_s, '%Y-%m-%d'), time.min))
                date_fin   = timezone.make_aware(datetime.combine(datetime.strptime(df_s, '%Y-%m-%d'), time(23, 59, 59)))
            except ValueError:
                return Response({"error": "Format de date invalide (YYYY-MM-DD)"}, status=400)

        def _apply_conditions(qs, source_type):
            master_q = Q()
            for cond in conditions:
                field = cond.get('field')
                op    = cond.get('operator')
                val   = cond.get('value')
                if not field or not op:
                    continue
                db_field = self._DYNAMIC_FIELD_MAP.get(field, {}).get(source_type)
                if not db_field:
                    continue
                if op == 'isnull':
                    q_obj = Q(**{f"{db_field}__isnull": True})
                elif op == 'notnull':
                    q_obj = Q(**{f"{db_field}__isnull": False})
                else:
                    if val is None or val == '':
                        continue
                    try:
                        if field == 'cip':
                            q_obj = Q(**{db_field: val})
                        else:
                            n = float(val)
                            ops_map = {'gte': '__gte', 'lte': '__lte', 'gt': '__gt', 'lt': '__lt', 'eq': ''}
                            suffix = ops_map.get(op)
                            if suffix is None:
                                continue
                            q_obj = Q(**{f"{db_field}{suffix}": n})
                    except (ValueError, TypeError):
                        continue
                master_q = master_q | q_obj if logic == 'OR' else master_q & q_obj
            return qs.filter(master_q)

        def _check_row(row_data):
            if not conditions:
                return True
            checks = []
            for cond in conditions:
                field = cond.get('field')
                op    = cond.get('operator')
                val   = cond.get('value')
                if not field or not op:
                    continue
                labels = self._DYNAMIC_LABEL_MAP.get(field)
                if not labels:
                    continue
                row_val = next((v for k, v in row_data.items() if k in labels), None)
                if op == 'isnull':
                    checks.append(row_val is None)
                    continue
                if op == 'notnull':
                    checks.append(row_val is not None)
                    continue
                if row_val is None or val is None or val == '':
                    continue
                try:
                    n_row, n_cond = float(row_val), float(val)
                    res_map = {'gte': n_row >= n_cond, 'lte': n_row <= n_cond,
                               'gt':  n_row >  n_cond, 'lt':  n_row <  n_cond,
                               'eq':  n_row == n_cond}
                    if op in res_map:
                        checks.append(res_map[op])
                except (ValueError, TypeError):
                    continue
            if not checks:
                return True
            return any(checks) if logic == 'OR' else all(checks)

        results = []

        if source == 'ventes':
            from api.models.billing import FactureProduit as FP, FactureProduitAllocation as FPA
            filters = {
                'facture__date__range': (date_debut, date_fin),
                'facture__status__in':  ['VAL', 'PAY'],
            }
            if vendeur_id: filters['facture__created_by_id'] = vendeur_id
            if client_id:  filters['facture__client_id']     = client_id
            if famille_id: filters['produit__famille_risque_id'] = famille_id
            qs = _apply_conditions(FP.objects.filter(**filters), 'ventes')
            for item in qs.select_related(
                'facture', 'facture__client', 'facture__created_by',
                'produit', 'produit__famille_risque', 'produit__rayon', 'produit__forme',
            ):
                p = item.produit
                f = item.facture
                prix_net      = float((item.selling_price or 0) - (item.discount or 0))
                qty           = float(item.quantity or 0)
                total_ht_ligne = round(prix_net * qty, 2)
                row = {}
                if 'date'    in requested_fields: row['Date']    = f.date.strftime('%d/%m/%Y') if f.date else 'N/A'
                if 'facture' in requested_fields: row['Facture'] = f.numero_facture
                if 'client'  in requested_fields: row['Client']  = f.client.name if f.client else 'Passage'
                if 'vendeur' in requested_fields: row['Vendeur'] = f.created_by.username if f.created_by else 'N/A'
                if 'produit' in requested_fields: row['Produit'] = p.name if p else item.produit_nom
                if 'famille' in requested_fields: row['Famille'] = p.famille_risque.nom if p and p.famille_risque else 'N/A'
                if 'quantite'   in requested_fields: row['Quantité']  = int(qty)
                if 'prix_vente' in requested_fields: row['Prix Vente'] = round(prix_net, 2)
                if 'total_ht'   in requested_fields: row['Total HT']   = total_ht_ligne
                if 'tva'        in requested_fields: row['TVA (%)']    = float(p.tva or 0) if p else 0
                if 'rayon'      in requested_fields: row['Rayon']      = p.rayon.name if p and p.rayon else 'N/A'
                if 'forme'      in requested_fields: row['Forme']      = p.forme.name if p and p.forme else 'N/A'
                if 'cip'        in requested_fields: row['Code CIP']   = p.cip1 if p else 'N/A'
                if any(x in requested_fields for x in ['cout_achat', 'marge', 'pourcentage_marge']):
                    allocs     = FPA.objects.filter(facture_produit=item)
                    total_cost = sum(float(a.quantity or 0) * float(a.cost_price or 0) for a in allocs)
                    if not allocs.exists() and p:
                        total_cost = float(p.pmp or 0) * qty
                    if 'cout_achat'        in requested_fields: row['Coût Achat']  = round(total_cost / qty, 2) if qty > 0 else 0
                    if 'marge'             in requested_fields: row['Marge Brute'] = round(total_ht_ligne - total_cost, 2)
                    if 'pourcentage_marge' in requested_fields:
                        row['Marge (%)'] = round((total_ht_ligne - total_cost) / total_ht_ligne * 100, 2) if total_ht_ligne > 0 else 0
                if _check_row(row):
                    results.append(row)

        elif source == 'achats':
            from api.models.orders import CommandeProduit as CP
            filters = {'commande__date_cloture__range': (date_debut, date_fin), 'commande__status': 'CLOT'}
            if fournisseur_id: filters['commande__fournisseur_id'] = fournisseur_id
            if famille_id:     filters['produit__famille_risque_id'] = famille_id
            qs = _apply_conditions(CP.objects.filter(**filters), 'achats')
            for cp in qs.select_related(
                'commande', 'commande__fournisseur',
                'produit', 'produit__rayon', 'produit__forme', 'produit__famille_risque',
            ):
                p   = cp.produit
                cmd = cp.commande
                qty = float(cp.quantity or 0)
                pu  = float(cp.price_cost or 0)
                row = {}
                if 'date'        in requested_fields: row['Réception']  = cmd.date_cloture.strftime('%d/%m/%Y') if cmd.date_cloture else 'N/A'
                if 'fournisseur' in requested_fields: row['Fournisseur'] = cmd.fournisseur.name if cmd.fournisseur else (cmd.fournisseur_nom or 'N/A')
                if 'produit'     in requested_fields: row['Produit']     = p.name if p else cp.produit_nom
                if 'famille'     in requested_fields: row['Famille']     = p.famille_risque.nom if p and p.famille_risque else 'N/A'
                if 'quantite'    in requested_fields: row['Quantité']    = int(qty)
                if 'cout_achat'  in requested_fields: row['P.U Achat']   = round(pu, 2)
                if 'total_ht'    in requested_fields: row['Total HT']    = round(pu * qty, 2)
                if 'tva'         in requested_fields: row['TVA (%)']     = float(p.tva or 0) if p else 0
                if 'cip'         in requested_fields: row['Code CIP']    = p.cip1 if p else 'N/A'
                if 'rayon'       in requested_fields: row['Rayon']       = p.rayon.name if p and p.rayon else 'N/A'
                if 'forme'       in requested_fields: row['Forme']       = p.forme.name if p and p.forme else 'N/A'
                if _check_row(row):
                    results.append(row)

        elif source == 'stock':
            from api.models.stock import StockLot
            filters = {'date_reception__range': (date_debut, date_fin)}
            if fournisseur_id: filters['fournisseur_id'] = fournisseur_id
            if famille_id:     filters['produit__famille_risque_id'] = famille_id
            qs = _apply_conditions(StockLot.objects.filter(**filters), 'stock')
            for lot in qs.select_related(
                'produit', 'produit__rayon', 'produit__forme',
                'produit__famille_risque', 'fournisseur',
            ):
                p = lot.produit
                row = {}
                if 'date'        in requested_fields: row['Réception']  = lot.date_reception.strftime('%d/%m/%Y') if lot.date_reception else 'N/A'
                if 'produit'     in requested_fields: row['Produit']     = p.name if p else lot.produit_nom
                if 'fournisseur' in requested_fields: row['Fournisseur'] = lot.fournisseur.name if lot.fournisseur else 'N/A'
                if 'famille'     in requested_fields: row['Famille']     = p.famille_risque.nom if p and p.famille_risque else 'N/A'
                if 'lot'         in requested_fields: row['Lot']         = lot.lot
                if 'quantite'    in requested_fields: row['Quantité']    = lot.quantity_remaining
                if 'cout_achat'  in requested_fields: row['P.U Achat']   = round(float(lot.price_cost or 0), 2)
                if 'total_ht'    in requested_fields: row['Total HT']    = round(float((lot.price_cost or 0) * (lot.quantity_remaining or 0)), 2)
                if 'tva'         in requested_fields: row['TVA (%)']     = float(p.tva or 0) if p else 0
                if 'rayon'       in requested_fields: row['Rayon']       = p.rayon.name if p and p.rayon else 'N/A'
                if 'forme'       in requested_fields: row['Forme']       = p.forme.name if p and p.forme else 'N/A'
                if 'cip'         in requested_fields: row['Code CIP']    = p.cip1 if p else 'N/A'
                if _check_row(row):
                    results.append(row)

        elif source == 'produits':
            from api.models.products import Produit
            filters = {}
            if famille_id: filters['famille_risque_id'] = famille_id
            qs = _apply_conditions(Produit.objects.filter(**filters), 'produits')
            for p in qs.select_related('rayon', 'forme', 'fournisseur', 'famille_risque'):
                row = {}
                if 'produit'         in requested_fields: row['Produit']     = p.name
                if 'famille'         in requested_fields: row['Famille']     = p.famille_risque.nom if p.famille_risque else 'N/A'
                if 'fournisseur'     in requested_fields: row['Fournisseur'] = p.fournisseur.name if p.fournisseur else 'N/A'
                if 'rayon'           in requested_fields: row['Rayon']       = p.rayon.name if p.rayon else 'N/A'
                if 'forme'           in requested_fields: row['Forme']       = p.forme.name if p.forme else 'N/A'
                if 'prix_vente'      in requested_fields: row['Prix Vente']  = round(float(p.selling_price or 0), 2)
                if 'cout_achat'      in requested_fields: row['Coût Achat']  = round(float(p.pmp or 0), 2)
                if 'quantite'        in requested_fields: row['Quantité']    = p.stock
                if 'total_ht'        in requested_fields: row['Total HT']    = round(float((p.stock or 0) * (p.pmp or 0)), 2)
                if 'tva'             in requested_fields: row['TVA (%)']     = float(p.tva or 0)
                if 'cip'             in requested_fields: row['Code CIP']    = p.cip1 or 'N/A'
                if 'stock_minimum'   in requested_fields: row['Stock Min']   = p.stock_minimum
                if 'pourcentage_marge' in requested_fields: row['Marge (%)'] = round(float(p.pourcentage_marge or 0), 2)
                if _check_row(row):
                    results.append(row)

        else:
            return Response({"error": f"Source inconnue: {source}"}, status=400)

        # ── Grouper par ───────────────────────────────────────────────────────
        if group_by and results:
            results = _group_results(results, group_by)

        # ── Trier par ─────────────────────────────────────────────────────────
        if results:
            reverse = (sort_order != 'asc')
            if sort_by and sort_by in results[0]:
                # Tri sur la colonne demandée (numérique ou texte)
                def _sort_key(row):
                    val = row.get(sort_by)
                    try:
                        return (0, float(val))
                    except (TypeError, ValueError):
                        return (1, str(val or ''))
                results.sort(key=_sort_key, reverse=reverse)
            else:
                # Tri par défaut : première colonne, alphabétique
                try:
                    first_key = list(results[0].keys())[0]
                    results.sort(key=lambda x: str(x.get(first_key, '')), reverse=reverse)
                except (IndexError, KeyError):
                    pass

        return Response(results)

    @action(detail=False, methods=['get'])
    def export_sage_i7(self, request):
        """
        Export comptable au format Sage Sari i7 (Sage 100).
        Génère un fichier CSV (délimiteur ;) avec Journal, Date, CompteG, CompteT, Libellé, Débit, Crédit.
        Comptes par défaut (OHADA/Cameroun) :
        - Ventes : 701100
        - TVA Collectée : 443100
        - Clients : 411100
        - Caisse : 571100
        - Banque : 521100
        """
        db_str = request.query_params.get('date_debut')
        df_str = request.query_params.get('date_fin')
        try:
            # On attend des dates au format YYYY-MM-DD
            date_debut = timezone.make_aware(datetime.combine(datetime.strptime(db_str, '%Y-%m-%d').date(), time.min))
            date_fin   = timezone.make_aware(datetime.combine(datetime.strptime(df_str, '%Y-%m-%d').date(), time.max))
        except (ValueError, AttributeError, TypeError):
            return Response({'error': 'Dates invalides (format requis: YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="export_sage_i7.csv"'
        
        # Encodage UTF-8 avec BOM pour une ouverture directe dans Excel sans soucis d'accents
        response.write('\ufeff'.encode('utf8'))
        writer = csv.writer(response, delimiter=';')
        
        # 1. Écritures de Ventes (Journal VT)
        factures = Facture.objects.filter(
            date__range=(date_debut, date_fin),
            status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE]
        ).exclude(produits__allocations__stock_lot__is_divers=True).select_related('client').order_by('date')

        for f in factures:
            date_sage = f.date.strftime('%d%m%y')
            num_fact = f.numero_facture or f"F{f.id}"
            libelle = f"Facture {num_fact}"
            if f.client:
                libelle += f" - {f.client.name}"
            
            # Débit Client (411100) - Total TTC
            writer.writerow(['VT', date_sage, '411100', f.client.id if f.client else '', libelle, str(f.total_ttc).replace('.', ','), '0'])
            
            # Crédit Ventes (701100) - Total HT
            writer.writerow(['VT', date_sage, '701100', '', libelle, '0', str(f.total_ht).replace('.', ',')])
            
            # Crédit TVA (443100) si applicable
            if f.total_tva > 0:
                writer.writerow(['VT', date_sage, '443100', '', libelle, '0', str(f.total_tva).replace('.', ',')])

        # 2. Écritures de Règlements (Journal CA / BQ)
        paiements = Caisse.objects.filter(
            date_paiement__range=(date_debut, date_fin),
            statut='completee'
        ).select_related('facture', 'facture__client').order_by('date_paiement')

        for p in paiements:
            date_sage = p.date_paiement.strftime('%d%m%y')
            # Mapping journal et compte de trésorerie
            is_cash = p.mode_paiement == 'especes'
            journal = 'CA' if is_cash else 'BQ'
            compte_t = '571100' if is_cash else '521100'
            
            num_fact = p.facture.numero_facture or p.facture.id
            libelle = f"Regl {p.get_mode_paiement_display()} Fact {num_fact}"
            
            # Débit Trésorerie (Caisse ou Banque)
            writer.writerow([journal, date_sage, compte_t, '', libelle, str(p.montant).replace('.', ','), '0'])
            
            # Crédit Client (411100)
            writer.writerow([journal, date_sage, '411100', p.facture.client.id if p.facture.client else '', libelle, '0', str(p.montant).replace('.', ',')])

        return response
