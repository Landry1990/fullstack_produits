from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Sum, F, DecimalField, Value
from django.db.models.functions import TruncDate, Coalesce
from django.utils import timezone
from datetime import datetime, timedelta, time
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm, mm
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
from api.models import Produit, CommandeProduit, Facture, FactureProduit, MouvementStock, StockLot

class RapportInventoryMixin:
    """
    Rapports liés à la gestion des stocks et de l'inventaire.
    """
    
    @action(detail=False, methods=['get'])
    def valeur_stock_journalier(self, request):
        date_debut_str = request.query_params.get('date_debut')
        date_fin_str = request.query_params.get('date_fin')
        if not date_debut_str or not date_fin_str:
            return Response({'error': 'Dates requises'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            date_debut = datetime.fromisoformat(date_debut_str.replace('Z', '+00:00')).date()
            date_fin = datetime.fromisoformat(date_fin_str.replace('Z', '+00:00')).date()
        except ValueError:
            return Response({'error': 'Format de date invalide'}, status=status.HTTP_400_BAD_REQUEST)

        stock_totals = Produit.objects.filter(stock__gt=0).aggregate(
            total_cost=Coalesce(Sum(F('stock') * F('pmp'), output_field=DecimalField()), Value(0, output_field=DecimalField())),
            total_ttc=Coalesce(Sum(F('stock') * F('selling_price'), output_field=DecimalField()), Value(0, output_field=DecimalField())),
        )
        current_stock_cost = stock_totals['total_cost']
        current_stock_ttc = stock_totals['total_ttc']
        today = timezone.now().date()
        
        ventes_ca = Facture.objects.filter(date__date__gte=date_debut, status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE]).annotate(jour=TruncDate('date')).values('jour').annotate(ca_net=Sum('total_ttc')).order_by('-jour')
        ventes_details = FactureProduit.objects.filter(facture__date__date__gte=date_debut, facture__status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE]).annotate(jour=TruncDate('facture__date')).values('jour').annotate(ventes_ttc_brut=Sum(F('quantity') * F('selling_price'), output_field=DecimalField()), cout_ventes=Sum(F('quantity') * F('produit__pmp'), output_field=DecimalField())).order_by('-jour')
        achats = CommandeProduit.objects.filter(commande__date_cloture__date__gte=date_debut, commande__status='CLOT').annotate(jour=TruncDate('commande__date_cloture')).values('jour').annotate(achats_cout=Sum((F('quantity') + F('unites_gratuites')) * F('price_cost'), output_field=DecimalField()), achats_ttc_virtuel=Sum((F('quantity') + F('unites_gratuites')) * F('produit__selling_price'), output_field=DecimalField())).order_by('-jour')
        
        mouvements_map = {}
        for v in ventes_ca:
            d = v['jour']
            if d not in mouvements_map: mouvements_map[d] = {'ventes_ttc_net': 0, 'ventes_ttc_brut': 0, 'cout_ventes': 0, 'achats_cout': 0, 'achats_ttc': 0}
            mouvements_map[d]['ventes_ttc_net'] = v['ca_net'] or 0
        for v in ventes_details:
            d = v['jour']
            if d not in mouvements_map: mouvements_map[d] = {'ventes_ttc_net': 0, 'ventes_ttc_brut': 0, 'cout_ventes': 0, 'achats_cout': 0, 'achats_ttc': 0}
            mouvements_map[d]['ventes_ttc_brut'] = v['ventes_ttc_brut'] or 0
            mouvements_map[d]['cout_ventes'] = v['cout_ventes'] or 0
        for a in achats:
            d = a['jour']
            if d not in mouvements_map: mouvements_map[d] = {'ventes_ttc_net': 0, 'ventes_ttc_brut': 0, 'cout_ventes': 0, 'achats_cout': 0, 'achats_ttc': 0}
            mouvements_map[d]['achats_cout'] = a['achats_cout'] or 0
            mouvements_map[d]['achats_ttc'] = a['achats_ttc_virtuel'] or 0
            
        resultats = []
        running_cost, running_ttc = float(current_stock_cost), float(current_stock_ttc)
        delta = (today - date_debut).days
        for i in range(delta + 1):
            current_day = today - timedelta(days=i)
            mops = mouvements_map.get(current_day, {'ventes_ttc_net': 0, 'ventes_ttc_brut': 0, 'cout_ventes': 0, 'achats_cout': 0, 'achats_ttc': 0})
            st_cost, st_ttc = running_cost - float(mops['achats_cout']) + float(mops['cout_ventes']), running_ttc - float(mops['achats_ttc']) + float(mops['ventes_ttc_brut'])
            if date_debut <= current_day <= date_fin:
                marge = float(mops['ventes_ttc_net']) - float(mops['cout_ventes'])
                resultats.append({'date': current_day.strftime('%Y-%m-%d'), 'valeur_stock_cout': round(running_cost, 0), 'valeur_stock_ttc': round(running_ttc, 0), 'achats_jour': round(float(mops['achats_cout']), 0), 'ventes_jour': round(float(mops['ventes_ttc_net']), 0), 'cout_ventes': round(float(mops['cout_ventes']), 0), 'marge': round(marge, 0), 'marge_pourcent': round((marge / float(mops['ventes_ttc_net']) * 100), 1) if mops['ventes_ttc_net'] > 0 else 0})
            running_cost, running_ttc = st_cost, st_ttc
        return Response(resultats)

    @action(detail=False, methods=['get'])
    def stocks_morts(self, request):
        try:
            min_value = Decimal(request.query_params.get('min_value', 100000))
            months = int(request.query_params.get('months', 6))
            export_format = request.query_params.get('format')
        except (ValueError, TypeError): return Response({'error': 'Paramètres invalides'}, status=status.HTTP_400_BAD_REQUEST)

        limit_date = (timezone.now() - timedelta(days=months*30)).date()
        produits = Produit.objects.filter(stock__gt=0).select_related('rayon', 'fournisseur')
        results = []
        for p in produits:
            valeur = (p.pmp or Decimal(0)) * p.stock
            if valeur >= min_value and (not p.dernier_vente or p.dernier_vente < limit_date):
                results.append({'id': p.id, 'name': p.name, 'cip': p.cip1, 'stock': p.stock, 'valeur': valeur, 'pmp': p.pmp, 'dernier_vente': p.dernier_vente, 'rayon': p.rayon.name if p.rayon else '', 'fournisseur': p.fournisseur.name if p.fournisseur else ''})
        
        results.sort(key=lambda x: x['valeur'], reverse=True)
        if export_format == 'csv':
            import csv
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="stocks_morts_{timezone.now().date()}.csv"'
            response.write(u'\ufeff'.encode('utf8'))
            writer = csv.writer(response, delimiter=';')
            writer.writerow(['Produit', 'CIP', 'Rayon', 'Fournisseur', 'Stock', 'PMP', 'Valeur Stock', 'Dernière Vente'])
            for r in results: writer.writerow([r['name'], r['cip'], r['rayon'], r['fournisseur'], str(r['stock']).replace('.', ','), str(r['pmp']).replace('.', ','), str(r['valeur']).replace('.', ','), r['dernier_vente'].strftime('%d/%m/%Y') if r['dernier_vente'] else 'Jamais'])
            return response
        return Response(results)

    @action(detail=False, methods=['get'])
    def balance_stock_excel(self, request):
        db_param, df_param = request.query_params.get('date_debut'), request.query_params.get('date_fin')
        lang, exclude_zero = request.query_params.get('lang', 'fr'), request.query_params.get('exclude_zero') == 'true'
        if not db_param or not df_param: return Response({'error': 'Dates requises'}, status=400)
        try:
            from django.utils.dateparse import parse_date
            date_debut, date_fin = datetime.combine(parse_date(db_param), time.min), datetime.combine(parse_date(df_param), time.max)
        except: return Response({'error': 'Date invalide'}, status=400)

        produits = Produit.objects.filter(is_active=True).only('id', 'name', 'cip1', 'stock', 'stock_reserve')
        stock_initial_dict = {item['produit_id']: item['total'] or 0 for item in MouvementStock.objects.filter(date__lt=date_debut).values('produit_id').annotate(total=Sum('quantite'))}
        mouvements_dict = {}
        for item in MouvementStock.objects.filter(date__range=(date_debut, date_fin)).values('produit_id', 'type_mouvement').annotate(total=Sum('quantite')):
            pid, tm, val = item['produit_id'], item['type_mouvement'], item['total'] or 0
            if pid not in mouvements_dict: mouvements_dict[pid] = {}
            mouvements_dict[pid][tm] = val

        wb = openpyxl.Workbook()
        ws = wb.active
        h = {'fr': {'title': "Balance des Stocks", 'cip': "Code CIP", 'designation': "Désignation", 'stock_initial': "Stock Initial", 'achats': "Achats", 'ventes': "Ventes", 'ajustements': "Ajustements", 'stock_final': "Stock Final"}, 'en': {'title': "Stock Balance", 'cip': "CIP Code", 'designation': "Designation", 'stock_initial': "Initial Stock", 'achats': "Purchases", 'ventes': "Sales", 'ajustements': "Adjustments", 'stock_final': "Final Stock"}}.get(lang, 'fr')
        
        ws.merge_cells('A1:G1'); ws['A1'] = f"{h['title']} - {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        ws['A1'].font = Font(bold=True, size=14); ws.append([]); ws.append([h['cip'], h['designation'], h['stock_initial'], h['achats'], h['ventes'], h['ajustements'], h['stock_final']])
        
        for p in produits:
            si, pm = stock_initial_dict.get(p.id, 0), mouvements_dict.get(p.id, {})
            ac, vt = pm.get(MouvementStock.TypeMouvement.ENTREE, 0), pm.get(MouvementStock.TypeMouvement.SORTIE, 0)
            aj = sum(v for t, v in pm.items() if t not in [MouvementStock.TypeMouvement.ENTREE, MouvementStock.TypeMouvement.SORTIE])
            sf = si + sum(pm.values())
            if not(exclude_zero and si==0 and ac==0 and vt==0 and aj==0 and sf==0):
                ws.append([p.cip1 or "", p.name, si, ac, -vt, aj, sf])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Balance_Stocks.xlsx"'; wb.save(response)
        return response

    def _get_valeur_stock_summary_data(self, valorisation, group_by=None):
        """Méthode interne pour calculer les agrégats de valeur de stock avec option de groupement."""
        is_pmp = valorisation == 'ACHAT'
        produits = Produit.objects.filter(stock__gt=0, is_active=True)
        
        # Optimisation si groupement demandé
        if group_by in ['rayon', 'forme', 'groupe']:
            produits = produits.select_related(group_by)
            
        tva_map = {}
        group_map = {}
        total_ttc_global = Decimal('0')
        total_ht_global = Decimal('0')
        total_tva_global = Decimal('0')
        
        for p in produits:
            qty = Decimal(str(p.stock))
            price_ttc = (p.pmp if is_pmp else p.selling_price) or Decimal('0')
            tva_rate = p.tva or Decimal('0')
            
            ttc_line = qty * price_ttc
            if tva_rate > 0:
                ht_line = (ttc_line / (1 + tva_rate / Decimal('100'))).quantize(Decimal('0.01'))
                tva_line = ttc_line - ht_line
            else:
                ht_line = ttc_line
                tva_line = Decimal('0')
                
            total_ttc_global += ttc_line
            total_ht_global += ht_line
            total_tva_global += tva_line
            
            # 1. Groupement par TVA
            rate_key = str(float(tva_rate))
            if rate_key not in tva_map:
                tva_map[rate_key] = {'rate': float(tva_rate), 'ht': Decimal('0'), 'tva': Decimal('0'), 'ttc': Decimal('0')}
            tva_map[rate_key]['ht'] += ht_line
            tva_map[rate_key]['tva'] += tva_line
            tva_map[rate_key]['ttc'] += ttc_line
            
            # 2. Groupement par catégorie (Rayon, Forme, Groupe)
            if group_by in ['rayon', 'forme', 'groupe']:
                group_obj = getattr(p, group_by)
                if group_by == 'rayon':
                    group_name = group_obj.name if group_obj else "Non classé"
                else:
                    group_name = group_obj.nom if group_obj else "Non classé"
                    
                if group_name not in group_map:
                    group_map[group_name] = {'name': group_name, 'ht': Decimal('0'), 'tva': Decimal('0'), 'ttc': Decimal('0')}
                
                group_map[group_name]['ht'] += ht_line
                group_map[group_name]['tva'] += tva_line
                group_map[group_name]['ttc'] += ttc_line
            
        res = {
            'is_pmp': is_pmp,
            'type_valorisation': 'PMP' if is_pmp else 'VENTE',
            'total_ht': total_ht_global,
            'total_tva': total_tva_global,
            'total_ttc': total_ttc_global,
            'tva_breakdown': sorted(tva_map.values(), key=lambda x: x['rate']),
            'date': timezone.now()
        }
        
        if group_by in ['rayon', 'forme', 'groupe']:
            res['group_by'] = group_by
            # Tri par valeur décroissante pour mettre en avant les catégories pesant le plus lourd
            res['group_breakdown'] = sorted(group_map.values(), key=lambda x: x['ttc'], reverse=True)
            
        return res

    def _get_valeur_stock_pharmacie_data(self, valorisation='ACHAT', group_by=None):
        """
        Méthode interne pour calculer les agrégats de valeur de stock PHARMACIE uniquement
        (exclut les lots divers - cadeaux, voyages, etc.)
        """
        is_pmp = valorisation == 'ACHAT'
        
        # Récupérer tous les lots NON divers avec stock > 0
        lots_pharmacie = StockLot.objects.filter(
            is_divers=False, 
            quantity_remaining__gt=0
        ).select_related('produit', 'produit__rayon')
        
        tva_map = {}
        group_map = {}
        total_ttc_global = Decimal('0')
        total_ht_global = Decimal('0')
        total_tva_global = Decimal('0')
        
        for lot in lots_pharmacie:
            qty = Decimal(str(lot.quantity_remaining))
            # Utiliser price_cost (PMP du lot) ou selling_price selon valorisation
            if is_pmp:
                price_ttc = lot.price_cost or Decimal('0')
            else:
                price_ttc = lot.produit.selling_price or Decimal('0')
            
            tva_rate = lot.produit.tva or Decimal('0')
            
            ttc_line = qty * price_ttc
            if tva_rate > 0:
                ht_line = (ttc_line / (1 + tva_rate / Decimal('100'))).quantize(Decimal('0.01'))
                tva_line = ttc_line - ht_line
            else:
                ht_line = ttc_line
                tva_line = Decimal('0')
                
            total_ttc_global += ttc_line
            total_ht_global += ht_line
            total_tva_global += tva_line
            
            # 1. Groupement par TVA
            rate_key = str(float(tva_rate))
            if rate_key not in tva_map:
                tva_map[rate_key] = {'rate': float(tva_rate), 'ht': Decimal('0'), 'tva': Decimal('0'), 'ttc': Decimal('0')}
            tva_map[rate_key]['ht'] += ht_line
            tva_map[rate_key]['tva'] += tva_line
            tva_map[rate_key]['ttc'] += ttc_line
            
            # 2. Groupement par catégorie (Rayon, Forme, Groupe)
            if group_by in ['rayon', 'forme', 'groupe']:
                group_obj = getattr(lot.produit, group_by, None)
                if group_by == 'rayon':
                    group_name = group_obj.name if group_obj else "Non classé"
                else:
                    group_name = group_obj.nom if group_obj else "Non classé"
                    
                if group_name not in group_map:
                    group_map[group_name] = {'name': group_name, 'ht': Decimal('0'), 'tva': Decimal('0'), 'ttc': Decimal('0')}
                
                group_map[group_name]['ht'] += ht_line
                group_map[group_name]['tva'] += tva_line
                group_map[group_name]['ttc'] += ttc_line
        
        res = {
            'is_pmp': is_pmp,
            'type_valorisation': 'PMP' if is_pmp else 'VENTE',
            'total_ht': total_ht_global,
            'total_tva': total_tva_global,
            'total_ttc': total_ttc_global,
            'tva_breakdown': sorted(tva_map.values(), key=lambda x: x['rate']),
            'date': timezone.now()
        }
        
        if group_by in ['rayon', 'forme', 'groupe']:
            res['group_by'] = group_by
            res['group_breakdown'] = sorted(group_map.values(), key=lambda x: x['ttc'], reverse=True)
            
        return res

    @action(detail=False, methods=['get'])
    def valeur_stock_json(self, request):
        """
        Retourne les données de valorisation au format JSON pour l'impression frontend.
        
        Paramètres:
        - valorisation: 'ACHAT' (PMP) ou 'VENTE' (prix de vente)
        - group_by: 'rayon', 'forme', 'groupe' (optionnel)
        - type: 'tous' (défaut), 'pharmacie' (sans divers), 'divers' (lots divers uniquement)
        """
        valorisation = request.query_params.get('valorisation', 'ACHAT')
        group_by = request.query_params.get('group_by')
        stock_type = request.query_params.get('type', 'tous')
        
        if stock_type == 'pharmacie':
            data = self._get_valeur_stock_pharmacie_data(valorisation, group_by=group_by)
        elif stock_type == 'divers':
            data = self._get_valeur_stock_divers_data(valorisation)
        else:
            # Type 'tous' - comportement historique (tous les produits)
            data = self._get_valeur_stock_summary_data(valorisation, group_by=group_by)
        
        return Response(data)

    @action(detail=False, methods=['get'])
    def valeur_stock_pdf(self, request):
        """
        Génère un récapitulatif PDF (Legacy/Direct) de la valeur du stock.
        
        Paramètres:
        - valorisation: 'ACHAT' (PMP) ou 'VENTE' (prix de vente)
        - type: 'tous' (défaut), 'pharmacie' (sans divers), 'divers' (lots divers uniquement)
        """
        from api.pdf_utils import (
            get_pharma_styles, draw_pharma_header, draw_pharma_footer, 
            format_currency, PharmaColors, get_pharma_table_style, 
            get_pharma_summary_table_style
        )
        
        valorisation = request.query_params.get('valorisation', 'ACHAT')
        stock_type = request.query_params.get('type', 'tous')
        
        # Sélectionner la méthode de calcul selon le type
        if stock_type == 'pharmacie':
            data = self._get_valeur_stock_pharmacie_data(valorisation)
            type_suffix = " - PHARMACIE (sans divers)"
        elif stock_type == 'divers':
            data = self._get_valeur_stock_divers_data(valorisation)
            type_suffix = " - DIVERS"
        else:
            data = self._get_valeur_stock_summary_data(valorisation)
            type_suffix = ""
        
        is_pmp = data['is_pmp']

        # 2. Construction du PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4, rightMargin=15*mm, leftMargin=15*mm, topMargin=45*mm, bottomMargin=20*mm
        )
        story = []
        styles = get_pharma_styles()
        
        story.append(Spacer(1, 5*mm))
        
        # --- Section 1: Récapitulatif Global ---
        type_label = "COÛT D'ACHAT (PMP)" if is_pmp else "PRIX DE VENTE (TTC)"
        doc_title = f"VALEUR STOCK{type_suffix} ({'ACHAT PMP' if is_pmp else 'VENTE'})" if is_pmp else f"VALEUR STOCK{type_suffix} ({'VENTE'})" if stock_type != 'divers' else f"VALEUR STOCK DIVERS ({'VENTE'})"
        doc_title = f"VALEUR STOCK{type_suffix} ({'PMP' if is_pmp else 'VENTE'})" if stock_type != 'divers' else f"VALEUR STOCK DIVERS ({'PMP' if is_pmp else 'VENTE'})"
        
        story.append(Paragraph(f"RÉCAPITULATIF GÉNÉRAL — {type_label}", styles['PharmaSubtitle']))
        story.append(Spacer(1, 3*mm))
        
        summary_data = [
            ["MÉTHODE DE VALORISATION", "MONTANT RECONSTITUÉ"],
            ["Valeur Totale HT", format_currency(data['total_ht'])],
            ["Montant Total TVA", format_currency(data['total_tva'])],
            [f"VALEUR TOTALE {('PMP' if is_pmp else 'TTC')}", format_currency(data['total_ttc'])]
        ]
        
        t_summary = Table(summary_data, colWidths=[9*cm, 7*cm])
        t_summary.setStyle(get_pharma_summary_table_style())
        story.append(t_summary)
        
        story.append(Spacer(1, 12*mm))
        
        # --- Section 2: Répartition par TVA ---
        story.append(Paragraph("RÉPARTITION DÉTAILLÉE PAR TAUX DE TVA", styles['PharmaSubtitle']))
        story.append(Spacer(1, 3*mm))
        
        tva_header = ["Taux TVA", "Base HT", "Montant TVA", "Total Reconstitué"]
        tva_data = [tva_header]
        
        for item in data['tva_breakdown']:
            tva_data.append([
                f"{item['rate']}%",
                format_currency(item['ht']),
                format_currency(item['tva']),
                format_currency(item['ttc'])
            ])
            
        t_tva = Table(tva_data, colWidths=[3.5*cm, 4*cm, 4*cm, 4.5*cm])
        t_tva.setStyle(get_pharma_table_style())
        story.append(t_tva)
        
        # --- Section 3: Notes ---
        story.append(Spacer(1, 20*mm))
        methode_desc = "fondée sur le PMP stocké en base." if is_pmp else "fondée sur les prix de vente publics actuels."
        story.append(Paragraph(f"<b>Note comptable :</b> Cette valorisation est {methode_desc}", styles['PharmaSmall']))
        
        doc.build(
            story, 
            onFirstPage=lambda c, d: (draw_pharma_header(c, d, title=doc_title), draw_pharma_footer(c, d)),
            onLaterPages=lambda c, d: (draw_pharma_header(c, d, title=doc_title), draw_pharma_footer(c, d))
        )
        
        suffix = "pmp" if is_pmp else "vente"
        filename = f"recap_valeur_{suffix}_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(buffer.getvalue())
        return response

    def _get_valeur_stock_divers_data(self, valorisation='ACHAT'):
        """
        Méthode interne pour calculer les agrégats de valeur de stock pour les lots divers uniquement.
        """
        is_pmp = valorisation == 'ACHAT'
        lots_divers = StockLot.objects.filter(is_divers=True, quantity_remaining__gt=0).select_related('produit', 'produit__rayon')
        
        tva_map = {}
        rayon_map = {}
        total_ttc_global = Decimal('0')
        total_ht_global = Decimal('0')
        total_tva_global = Decimal('0')
        
        for lot in lots_divers:
            qty = Decimal(str(lot.quantity_remaining))
            price_ttc = (lot.price_cost if is_pmp else lot.produit.selling_price) or Decimal('0')
            tva_rate = lot.produit.tva or Decimal('0')
            
            ttc_line = qty * price_ttc
            if tva_rate > 0:
                ht_line = (ttc_line / (1 + tva_rate / Decimal('100'))).quantize(Decimal('0.01'))
                tva_line = ttc_line - ht_line
            else:
                ht_line = ttc_line
                tva_line = Decimal('0')
                
            total_ttc_global += ttc_line
            total_ht_global += ht_line
            total_tva_global += tva_line
            
            # Groupement par TVA
            rate_key = str(float(tva_rate))
            if rate_key not in tva_map:
                tva_map[rate_key] = {'rate': float(tva_rate), 'ht': Decimal('0'), 'tva': Decimal('0'), 'ttc': Decimal('0')}
            tva_map[rate_key]['ht'] += ht_line
            tva_map[rate_key]['tva'] += tva_line
            tva_map[rate_key]['ttc'] += ttc_line
            
            # Groupement par rayon
            rayon_name = lot.produit.rayon.name if lot.produit.rayon else "Non classé"
            if rayon_name not in rayon_map:
                rayon_map[rayon_name] = {'name': rayon_name, 'ht': Decimal('0'), 'tva': Decimal('0'), 'ttc': Decimal('0')}
            rayon_map[rayon_name]['ht'] += ht_line
            rayon_map[rayon_name]['tva'] += tva_line
            rayon_map[rayon_name]['ttc'] += ttc_line
        
        return {
            'is_pmp': is_pmp,
            'type_valorisation': 'PMP' if is_pmp else 'VENTE',
            'total_ht': total_ht_global,
            'total_tva': total_tva_global,
            'total_ttc': total_ttc_global,
            'tva_breakdown': sorted(tva_map.values(), key=lambda x: x['rate']),
            'rayon_breakdown': sorted(rayon_map.values(), key=lambda x: x['ttc'], reverse=True),
            'date': timezone.now()
        }

    @action(detail=False, methods=['get'])
    def valeur_stock_divers_json(self, request):
        """Retourne les données de valorisation des lots divers au format JSON."""
        valorisation = request.query_params.get('valorisation', 'ACHAT')
        data = self._get_valeur_stock_divers_data(valorisation)
        return Response(data)
