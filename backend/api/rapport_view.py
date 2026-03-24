from django.utils.formats import date_format
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Sum, F, DecimalField, Q, Count, Value
from django.db.models.functions import TruncDate, Coalesce
from django.utils import timezone
from datetime import datetime, timedelta, time
from decimal import Decimal
from api.pagination import StandardResultsSetPagination
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from api.models import Facture, FactureProduitAllocation, Caisse, FactureProduit, CommandeProduit, Produit, MouvementStock


class RapportViewSet(viewsets.ViewSet):
    """
    API endpoint pour les rapports mensuels.
    """
    permission_classes = [IsAuthenticated]

    # ============== HELPER METHODS FOR RAPPORT ==============
    
    def _get_factures_periode(self, date_debut, date_fin):
        """Récupère les factures validées/payées de la période."""
        return Facture.objects.filter(
            status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
            date__gte=date_debut,
            date__lt=date_fin
        ).prefetch_related('produits', 'produits__produit', 'paiements')

    def _calculate_ca_stats(self, factures):
        """Calcule CA TTC, CA HT, nombre de ventes et remises."""
        ca_stats = factures.aggregate(
            ca_ttc=Coalesce(Sum('total_ttc'), Decimal('0.00')),
            ca_ht=Coalesce(Sum('total_ht'), Decimal('0.00')),
            total_remises_global=Coalesce(Sum('remise'), Decimal('0.00')),
            total_remises_fidelite=Coalesce(Sum('montant_fidelite'), Decimal('0.00')),
            part_client=Coalesce(Sum('part_client'), Decimal('0.00'))
        )
        
        # Remises sur lignes et Valeur des Unités Gratuites (UG)
        from api.models import FactureProduit
        prod_stats = FactureProduit.objects.filter(
            facture__in=factures
        ).aggregate(
            total_remises_lignes=Coalesce(Sum(F('discount') * F('quantity'), output_field=DecimalField()), Decimal('0.00')),
            total_valeur_ug=Coalesce(Sum(F('free_quantity') * F('selling_price'), output_field=DecimalField()), Decimal('0.00'))
        )

        part_assurance = ca_stats['ca_ttc'] - ca_stats['part_client']
        
        return {
            'ca_ttc': ca_stats['ca_ttc'],
            'ca_ht': ca_stats['ca_ht'],
            'total_remises': (
                ca_stats['total_remises_global'] + 
                ca_stats['total_remises_fidelite'] + 
                prod_stats['total_remises_lignes'] + 
                prod_stats['total_valeur_ug']
            ),
            'total_remises_detail': {
                'global': ca_stats['total_remises_global'],
                'fidelite': ca_stats['total_remises_fidelite'],
                'lignes': prod_stats['total_remises_lignes'],
                'unites_gratuites': prod_stats['total_valeur_ug'],
            },
            'part_assurance': part_assurance,
            'part_client': ca_stats['part_client'],
            'nb_ventes': factures.count()
        }

    def _calculate_margin(self, factures):
        """Calcule la marge brute via les allocations FIFO."""
        allocations = FactureProduitAllocation.objects.filter(
            facture_produit__facture__in=factures
        )

        cout_achat_total = allocations.aggregate(
            total=Coalesce(Sum(F('cost_price') * F('quantity'), output_field=DecimalField()), Decimal('0.00'))
        )['total']
        
        ca_ht_total = factures.aggregate(
            total=Coalesce(Sum('total_ht'), Decimal('0.00'))
        )['total']
        
        marge_brute = ca_ht_total - cout_achat_total
        marge_pct = (marge_brute / ca_ht_total * 100) if ca_ht_total > 0 else Decimal('0.00')
        
        return {
            'cout_achat': cout_achat_total,
            'marge_brute': marge_brute,
            'marge_pct': round(marge_pct, 2)
        }

    def _calculate_encaissements(self, date_debut, date_fin, factures):
        """Calcule les encaissements par mode, ventes à crédit et coupons."""
        # Détection des recouvrements : mode spécifique ou mention [RECOUV]
        recouvrement_q = Q(mode_paiement='recouvrement') | Q(reference__icontains='[RECOUV]')

        # Encaissements réels (Cash flow) - Exclure recouvrement, en_compte et depot
        encaissements_qs = Caisse.objects.filter(
            date_paiement__gte=date_debut,
            date_paiement__lt=date_fin,
            statut='completee'
        ).exclude(
            recouvrement_q | Q(mode_paiement__in=['en_compte', 'coupon', 'depot'])
        ).values('mode_paiement').annotate(
            total=Sum('montant')
        )
        
        # Mapper les résultats pour garantir la présence de tous les modes essentiels
        results_map = {enc['mode_paiement']: enc['total'] for enc in encaissements_qs}
        
        # Liste des modes à afficher systématiquement
        modes_a_afficher = ['especes', 'momo', 'om', 'carte', 'virement', 'cheque']
        
        encaissements_data = []
        caisse_modes_dict = dict(Caisse.MODES_PAIEMENT)
        
        for m in modes_a_afficher:
            encaissements_data.append({
                'mode': m,
                'mode_label': caisse_modes_dict.get(m, m.replace('_', ' ').title()),
                'montant': results_map.get(m, Decimal('0.00'))
            })
            
        # Ajouter d'autres modes qui auraient pu être utilisés mais ne sont pas dans la liste fixe
        for m_code, m_total in results_map.items():
            if m_code not in modes_a_afficher:
                encaissements_data.append({
                    'mode': m_code,
                    'mode_label': caisse_modes_dict.get(m_code, m_code),
                    'montant': m_total
                })

        # Recouvrements (Encaissements de créances anciennes - Hybride)
        recouvrements_total = Caisse.objects.filter(
            Q(date_paiement__gte=date_debut),
            Q(date_paiement__lt=date_fin),
            Q(statut='completee'),
            recouvrement_q
        ).aggregate(total=Sum('montant'))['total'] or Decimal('0.00')

        # Ventes à crédit
        ventes_credit = Caisse.objects.filter(
            date_paiement__gte=date_debut,
            date_paiement__lt=date_fin,
            statut='completee',
            mode_paiement='en_compte'
        ).aggregate(total=Sum('montant'))['total'] or Decimal('0.00')

        # Dépôts utilisés (Non-physique)
        depots_total = Caisse.objects.filter(
            date_paiement__gte=date_debut,
            date_paiement__lt=date_fin,
            statut='completee',
            mode_paiement='depot'
        ).aggregate(total=Sum('montant'))['total'] or Decimal('0.00')

        # Coupons utilisés
        from api.models import CouponMonnaie
        coupons_total = CouponMonnaie.objects.filter(
            facture_utilisation__in=factures,
            status='UTILISE'
        ).aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
        
        return {
            'encaissements': encaissements_data,
            'recouvrements_total': recouvrements_total,
            'ventes_credit': ventes_credit,
            'depots_total': depots_total,
            'coupons_total': coupons_total
        }

    def _calculate_creances(self):
        """Calcule les créances globales à percevoir."""
        
        # Optimize to a single query instead of N+1
        stats = Facture.objects.filter(
            status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE]
        ).annotate(
            paid_amount=Coalesce(
                Sum('paiements__montant', filter=Q(paiements__statut='completee') & ~Q(paiements__mode_paiement='en_compte')),
                Value(0, output_field=DecimalField())
            ),
            reste=F('total_ttc') - F('paid_amount')
        ).filter(
            reste__gt=0.5
        ).aggregate(
            total=Coalesce(Sum('reste'), Decimal('0.00')),
            nb_factures=Count('id')
        )
        
        return {
            'total': stats['total'],
            'nb_factures': stats['nb_factures'],
            'creances_a_percevoir': stats['total']  # Backward compatibility
        }

    def _calculate_ca_par_tva(self, factures):
        """Calcule la répartition du CA par taux de TVA."""
        # This is strictly accurate if remises are apportioned proportionally 
        # to the contribution of each line's gross total per TVA rate.
        
        ca_par_tva_stats = {}

        # 1. Total brut TTC par facture (avant remise)
        # Needed for proportional discount apportionment
        # Or faster: we fetch all FactureProduit with their parent facture remise & total_brut
        from api.models import FactureProduit
        
        lignes = FactureProduit.objects.filter(
            facture__in=factures
        ).values(
            'tva',
            'facture__id',
            'facture__remise'
        ).annotate(
            ligne_brut_ttc=Sum(
                F('quantity') * (F('selling_price') - Coalesce(F('discount'), Value(0, output_field=DecimalField()))), 
                output_field=DecimalField()
            )
        )
        
        # We need the facture total brut to calculate ratio
        from collections import defaultdict
        
        facture_totals = defaultdict(Decimal)
        for ligne in lignes:
            facture_totals[ligne['facture__id']] += ligne['ligne_brut_ttc']
            
        for ligne in lignes:
            taux = ligne['tva']
            fid = ligne['facture__id']
            ligne_brut = ligne['ligne_brut_ttc']
            remise_facture = ligne['facture__remise']
            
            total_brut_facture = facture_totals[fid]
            
            # Ratio
            ratio = ligne_brut / total_brut_facture if total_brut_facture > 0 else Decimal('0.00')
            part_remise = remise_facture * ratio
            
            ttc_net = ligne_brut - part_remise
            
            if taux > 0:
                ht_net = (ttc_net / (1 + taux / Decimal('100.00'))).quantize(Decimal('0.01'))
                tva_montant = ttc_net - ht_net
            else:
                ht_net = ttc_net
                tva_montant = Decimal('0.00')
                
            if taux not in ca_par_tva_stats:
                ca_par_tva_stats[taux] = {'ca_ht': Decimal('0.00'), 'montant_tva': Decimal('0.00'), 'ca_ttc': Decimal('0.00')}
                
            ca_par_tva_stats[taux]['ca_ht'] += ht_net
            ca_par_tva_stats[taux]['montant_tva'] += tva_montant
            ca_par_tva_stats[taux]['ca_ttc'] += ttc_net

        return [
            {
                'taux': float(taux),
                'ca_ht': stats['ca_ht'],
                'montant_tva': stats['montant_tva'],
                'ca_ttc': stats['ca_ttc']
            }
            for taux, stats in sorted(ca_par_tva_stats.items(), key=lambda x: x[0], reverse=True)
        ]

    def _calculate_achats_fournisseurs(self, date_debut, date_fin):
        """Calcule les achats nets par fournisseur (Commandes - Avoirs)."""
        from api.models import Commande, Avoir
        
        # Commandes clôturées
        commandes_mois = Commande.objects.filter(
            date__gte=date_debut,
            date__lt=date_fin,
            status='CLOT'
        ).prefetch_related('produits')
        
        achats_stats = {}
        for commande in commandes_mois:
            if not commande.fournisseur:
                continue
            fid = commande.fournisseur.id
            if fid not in achats_stats:
                achats_stats[fid] = {
                    'fournisseur_id': fid,
                    'fournisseur_nom': commande.fournisseur.name,
                    'montant_total': Decimal('0.00'),
                    'nb_commandes': 0,
                    'nb_avoirs': 0,
                    'montant_avoirs': Decimal('0.00')
                }
            cout_cmd = sum(cp.quantity * cp.price for cp in commande.produits.all())
            achats_stats[fid]['montant_total'] += cout_cmd
            achats_stats[fid]['nb_commandes'] += 1
            
        # Avoirs validés (retours)
        avoirs_mois = Avoir.objects.filter(
            date__gte=date_debut.date(),
            date__lt=date_fin.date(),
            status='VALIDEE'
        )
        
        for avoir in avoirs_mois:
            if not avoir.fournisseur:
                continue
            fid = avoir.fournisseur.id
            if fid not in achats_stats:
                achats_stats[fid] = {
                    'fournisseur_id': fid,
                    'fournisseur_nom': avoir.fournisseur.name,
                    'montant_total': Decimal('0.00'),
                    'nb_commandes': 0,
                    'nb_avoirs': 0,
                    'montant_avoirs': Decimal('0.00')
                }
            montant_avoir = avoir.total_ht
            achats_stats[fid]['montant_total'] -= montant_avoir
            achats_stats[fid]['montant_avoirs'] += montant_avoir
            achats_stats[fid]['nb_avoirs'] += 1
            
        return sorted(achats_stats.values(), key=lambda x: x['montant_total'], reverse=True)

    def _calculate_clients_pro(self, factures):
        """Calcule les statistiques des clients professionnels sans inflation d'agrégation."""
        pro_factures = factures.filter(client__client_type='PROFESSIONNEL')
        
        # 1. Billing stats (No joins here, so total_ttc is safe)
        billing_stats = pro_factures.values(
            'client__id', 
            'client__name'
        ).annotate(
            total_billed=Sum('total_ttc', output_field=DecimalField()),
            nb_factures=Count('id')
        )

        # 2. Payment stats (Separate query to avoid multiplying total_ttc by pay count)
        payment_stats_query = Caisse.objects.filter(
            facture__in=pro_factures,
            statut='completee'
        ).exclude(
            mode_paiement='en_compte'
        ).values('facture__client__id').annotate(
            total_paid=Sum('montant')
        )
        
        payments_map = {p['facture__client__id']: p['total_paid'] for p in payment_stats_query}

        # 3. Merge and summarize
        ca_pro_total = Decimal('0.00')
        montant_paye_pro = Decimal('0.00')
        nb_factures_total = 0
        top_clients_pro = []

        for b_stat in billing_stats:
            cid = b_stat['client__id']
            cnom = b_stat['client__name']
            
            total_billed = b_stat['total_billed'] or Decimal('0.00')
            paid_amount = payments_map.get(cid, Decimal('0.00'))
            nb_factures = b_stat['nb_factures']
            
            ca_pro_total += total_billed
            montant_paye_pro += paid_amount
            nb_factures_total += nb_factures
            
            top_clients_pro.append({
                'client_id': cid,
                'client_nom': cnom,
                'ca_total': total_billed,
                'montant_paye': paid_amount,
                'reste_a_payer': total_billed - paid_amount
            })

        reste_a_payer_pro = ca_pro_total - montant_paye_pro
        taux_recouvrement_pro = (montant_paye_pro / ca_pro_total * 100) if ca_pro_total > 0 else Decimal('0.00')
        
        top_clients_pro.sort(key=lambda x: x['reste_a_payer'], reverse=True)
        
        return {
            'ca_total': ca_pro_total,
            'montant_paye': montant_paye_pro,
            'reste_a_payer': reste_a_payer_pro,
            'taux_recouvrement_pct': round(taux_recouvrement_pro, 2),
            'nb_factures': nb_factures_total,
            'top_clients': top_clients_pro[:10]
        }

    def _calculate_unites_gratuites(self, date_debut, date_fin):
        """Calcule les statistiques des unités gratuites reçues."""
        commandes_produits_ug = CommandeProduit.objects.filter(
            commande__date__gte=date_debut,
            commande__date__lt=date_fin,
            commande__status='CLOT',
            unites_gratuites__gt=0
        ).select_related('produit')
        
        valeur_ug_total = Decimal('0.00')
        qty_ug_total = 0
        ug_par_produit = {}
        
        for cp in commandes_produits_ug:
            qty_gratuite = cp.unites_gratuites
            valeur = qty_gratuite * cp.produit.selling_price
            valeur_ug_total += valeur
            qty_ug_total += qty_gratuite
            
            pid = cp.produit.id
            if pid not in ug_par_produit:
                ug_par_produit[pid] = {
                    'produit_id': pid,
                    'produit_nom': cp.produit.name,
                    'quantite_gratuite': 0,
                    'valeur_totale': Decimal('0.00')
                }
            ug_par_produit[pid]['quantite_gratuite'] += qty_gratuite
            ug_par_produit[pid]['valeur_totale'] += valeur
        
        top_ug = sorted(ug_par_produit.values(), key=lambda x: x['valeur_totale'], reverse=True)[:10]
        
        return {
            'valeur_totale': valeur_ug_total,
            'quantite_totale': qty_ug_total,
            'pct_du_ca': Decimal('0.00'),  # Will be calculated in main method
            'nb_produits_distincts': len(ug_par_produit),
            'top_produits': top_ug
        }

    def _calculate_mouvements_caisse(self, date_debut, date_fin):
        """Calcule les entrées/sorties de caisse hors ventes."""
        from api.models import MouvementCaisse
        
        mouvements = MouvementCaisse.objects.filter(
            date__gte=date_debut,
            date__lt=date_fin
        ).select_related('user')
        
        total_entrees = Decimal('0.00')
        total_sorties = Decimal('0.00')
        mouvements_data = []
        
        for mvt in mouvements:
            if mvt.type == 'ENTREE':
                total_entrees += mvt.montant
            else:
                total_sorties += mvt.montant
            mouvements_data.append({
                'id': mvt.id,
                'date': mvt.date,
                'type': mvt.type,
                'montant': mvt.montant,
                'motif': mvt.motif,
                'user': mvt.user.get_full_name() if mvt.user else 'Inconnu'
            })
        
        return {
            'total_entrees': total_entrees,
            'total_sorties': total_sorties,
            'solde': total_entrees - total_sorties,
            'liste': mouvements_data
        }

    # ============== MAIN RAPPORT METHOD ==============

    def _get_rapport_data(self, date_debut, date_fin, mois_str):
        """
        Méthode principale pour calculer toutes les données du rapport.
        Orchestrate les appels aux méthodes spécialisées.
        """
        # 1. Récupérer les factures une seule fois
        factures = self._get_factures_periode(date_debut, date_fin)
        
        # 2. Calculer chaque section
        ca_stats = self._calculate_ca_stats(factures)
        marge = self._calculate_margin(factures)
        encaissements = self._calculate_encaissements(date_debut, date_fin, factures)
        creances = self._calculate_creances()
        ca_par_tva = self._calculate_ca_par_tva(factures)
        achats = self._calculate_achats_fournisseurs(date_debut, date_fin)
        clients_pro = self._calculate_clients_pro(factures)
        ug = self._calculate_unites_gratuites(date_debut, date_fin)
        mouvements = self._calculate_mouvements_caisse(date_debut, date_fin)
        
        # 3. Calculer le % UG par rapport au CA
        if ca_stats['ca_ttc'] > 0:
            ug['pct_du_ca'] = round((ug['valeur_totale'] / ca_stats['ca_ttc'] * 100), 2)
        
        # 4. Assembler le résultat final
        return {
            'mois': mois_str,
            'periode': {
                'debut': date_debut.isoformat(),
                'fin': date_fin.isoformat()
            },
            'ca': ca_stats,
            'marge': marge,
            'encaissements': encaissements['encaissements'],
            'recouvrements_total': encaissements['recouvrements_total'],
            'ventes_credit': encaissements['ventes_credit'],
            'coupons_total': encaissements['coupons_total'],
            'creances_a_percevoir': creances['total'],
            'creances': {
                'total': creances['total'],
                'nb_factures': creances['nb_factures']
            },
            'depots_total': encaissements['depots_total'],
            'ca_par_tva': ca_par_tva,
            'achats_par_fournisseur': achats,
            'clients_professionnels': clients_pro,
            'unites_gratuites': ug,
            'mouvements_caisse': mouvements
        }


    @action(detail=False, methods=['get'])
    def valeur_stock_journalier(self, request):
        """
        Reconstruit la valeur du stock jour par jour (Back-casting).
        Retourne : Date, Stock (Coût), Stock (TTC), Achats (Coût), Ventes (TTC), Coût Ventes.
        """
        date_debut_str = request.query_params.get('date_debut')
        date_fin_str = request.query_params.get('date_fin')

        if not date_debut_str or not date_fin_str:
            return Response(
                {'error': 'Les paramètres date_debut et date_fin sont requis.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            date_debut = datetime.fromisoformat(date_debut_str.replace('Z', '+00:00')).date()
            date_fin = datetime.fromisoformat(date_fin_str.replace('Z', '+00:00')).date()
        except ValueError:
            return Response({'error': 'Format de date invalide (ISO attendu).'}, status=status.HTTP_400_BAD_REQUEST)

        # 1. État Initial (Aujourd'hui / Maintenant)
        # Optimisation : Utilisation d'agrégation SQL au lieu d'une boucle Python
        stock_totals = Produit.objects.filter(stock__gt=0).aggregate(
            total_cost=Coalesce(Sum(F('stock') * F('pmp'), output_field=DecimalField()), Value(0, output_field=DecimalField())),
            total_ttc=Coalesce(Sum(F('stock') * F('selling_price'), output_field=DecimalField()), Value(0, output_field=DecimalField())),
        )
        current_stock_cost = stock_totals['total_cost']
        current_stock_ttc = stock_totals['total_ttc']

        today = timezone.now().date()
        
        # Mouvements : Ventes (FactureProduit pour le coût, Facture pour le CA Net)
        # On utilise Facture pour le CA TTC Net afin d'inclure remises et remises lignes correctement
        ventes_ca = Facture.objects.filter(
            date__date__gte=date_debut,
            status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE]
        ).annotate(
            jour=TruncDate('date')
        ).values('jour').annotate(
            ca_net=Sum('total_ttc')
        ).order_by('-jour')

        # Pour le back-casting du stock TTC virtuel et le coût des ventes
        ventes_details = FactureProduit.objects.filter(
            facture__date__date__gte=date_debut,
            facture__status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE]
        ).annotate(
            jour=TruncDate('facture__date')
        ).values('jour').annotate(
            ventes_ttc_brut=Sum(F('quantity') * F('selling_price'), output_field=DecimalField()),
            cout_ventes=Sum(F('quantity') * F('produit__pmp'), output_field=DecimalField())
        ).order_by('-jour')
        
        # Mouvements : Achats (CommandeProduit Cloturée)
        achats = CommandeProduit.objects.filter(
            commande__date_cloture__date__gte=date_debut,
            commande__status='CLOT'
        ).annotate(
            jour=TruncDate('commande__date_cloture')
        ).values('jour').annotate(
            achats_cout=Sum((F('quantity') + F('unites_gratuites')) * F('price_cost'), output_field=DecimalField()),
            achats_ttc_virtuel=Sum((F('quantity') + F('unites_gratuites')) * F('produit__selling_price'), output_field=DecimalField())
        ).order_by('-jour')
        
        # Indexer par date
        mouvements_map = {}
        
        for v in ventes_ca:
            d = v['jour']
            if d not in mouvements_map: mouvements_map[d] = {'ventes_ttc_net': 0, 'ventes_ttc_brut': 0, 'cout_ventes': 0, 'achats_cout': 0, 'achats_ttc': 0}
            mouvements_map[d]['ventes_ttc_net'] = v['ca_net'] or 0
            
        for v in ventes_details:
            d = v['jour']
            if d not in mouvements_map: mouvements_map[d] = {'ventes_ttc_net': 0, 'ventes_ttc_brut': 0, 'cout_ventes': 0, 'achats_cout': 0, 'achats_ttc': 0}
            mouvements_map[d]['ventes_ttc_brut'] = v['ventes_ttc_brut'] or 0
            mouvements_map[d]['cout_ventes'] = v['cout_ventes'] or 0
            
        for a in achats:
            d = a['jour']
            if d not in mouvements_map: mouvements_map[d] = {'ventes_ttc_net': 0, 'ventes_ttc_brut': 0, 'cout_ventes': 0, 'achats_cout': 0, 'achats_ttc': 0}
            mouvements_map[d]['achats_cout'] = a['achats_cout'] or 0
            mouvements_map[d]['achats_ttc'] = a['achats_ttc_virtuel'] or 0
            
        # Back-casting
        resultats = []
        running_cost = float(current_stock_cost)
        running_ttc = float(current_stock_ttc)
        delta = (today - date_debut).days
        
        for i in range(delta + 1):
            current_day = today - timedelta(days=i)
            
            mops = mouvements_map.get(current_day, {'ventes_ttc_net': 0, 'ventes_ttc_brut': 0, 'cout_ventes': 0, 'achats_cout': 0, 'achats_ttc': 0})
            
            ventes_ttc_net = float(mops['ventes_ttc_net'] or 0)
            ventes_ttc_brut = float(mops['ventes_ttc_brut'] or 0)
            cout_ventes = float(mops['cout_ventes'] or 0)
            achats_cout = float(mops['achats_cout'] or 0)
            achats_ttc = float(mops['achats_ttc'] or 0)
            
            end_day_cost = running_cost
            end_day_ttc = running_ttc
            
            # Back-casting : Stock Début = Stock Fin - Achats + Sorties(Ventes)
            # IMPORTANT : Pour le stock TTC virtuel, on utilise le prix de vente BRUT (valeur théorique du stock)
            start_day_cost = end_day_cost - achats_cout + cout_ventes
            start_day_ttc = end_day_ttc - achats_ttc + ventes_ttc_brut
            
            if date_debut <= current_day <= date_fin:
                marge = ventes_ttc_net - cout_ventes
                marge_pourcent = 0
                if ventes_ttc_net > 0:
                    marge_pourcent = (marge / ventes_ttc_net) * 100
                
                resultats.append({
                    'date': current_day.strftime('%Y-%m-%d'),
                    'valeur_stock_cout': round(end_day_cost, 0),
                    'valeur_stock_ttc': round(end_day_ttc, 0),
                    'achats_jour': round(achats_cout, 0),
                    'ventes_jour': round(ventes_ttc_net, 0), # CA Net pour comparaison avec rapports CA
                    'cout_ventes': round(cout_ventes, 0),
                    'marge': round(marge, 0),
                    'marge_pourcent': round(marge_pourcent, 1)
                })
            
            running_cost = start_day_cost
            running_ttc = start_day_ttc


        return Response(resultats)

    @action(detail=False, methods=['get'])
    def rapport_ca_multi_annuel(self, request):
        """
        Calcule le CA par mois et par année (TVA vs Exonéré).
        Retourne un tableau pivoté exploitable directement par le frontend avec i18n.
        """
        annees_dispo = Facture.objects.filter(
            status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE]
        ).dates('date', 'year', order='DESC')
        
        annees = [d.year for d in annees_dispo]
        if not annees:
            return Response([])

        month_keys = [
            'january', 'february', 'march', 'april', 'may', 'june',
            'july', 'august', 'september', 'october', 'november', 'december'
        ]
        
        # Structure de retour
        resultats = []
        for i in range(12):
            resultats.append({
                'Mois': month_keys[i],
                '_index': i + 1
            })

        # Ligne de totaux
        totaux_generaux = {
            'Mois': 'total_general',
            '_index': 13
        }

        # 2. Calculer pour chaque année
        for annee in sorted(annees):
            annee_total_tva = Decimal('0.00')
            annee_total_exo = Decimal('0.00')
            
            for m_idx in range(1, 13):
                date_debut = timezone.make_aware(datetime(annee, m_idx, 1))
                if m_idx == 12:
                    date_fin = timezone.make_aware(datetime(annee + 1, 1, 1))
                else:
                    date_fin = timezone.make_aware(datetime(annee, m_idx + 1, 1))
                
                factures = self._get_factures_periode(date_debut, date_fin)
                ca_par_tva = self._calculate_ca_par_tva(factures)
                
                ca_tva = Decimal('0.00')
                ca_exonerer = Decimal('0.00')
                
                for item in ca_par_tva:
                    if item['taux'] > 0:
                        ca_tva += item['ca_ttc']
                    else:
                        ca_exonerer += item['ca_ttc']
                
                # Ajouter les colonnes dynamiques
                row = resultats[m_idx - 1]
                row[f"{annee}_ca_tva"] = ca_tva
                row[f"{annee}_ca_exo"] = ca_exonerer
                row[f"{annee}_total"] = ca_tva + ca_exonerer
                
                # Accumuler pour le total annuel de la ligne finale
                annee_total_tva += ca_tva
                annee_total_exo += ca_exonerer

            # Ajouter les totaux pour l'année dans la ligne finale
            totaux_generaux[f"{annee}_ca_tva"] = annee_total_tva
            totaux_generaux[f"{annee}_ca_exo"] = annee_total_exo
            totaux_generaux[f"{annee}_total"] = annee_total_tva + annee_total_exo

        resultats.append(totaux_generaux)

        # Trier les mois
        resultats.sort(key=lambda x: x['_index'])
        
        # Nettoyer
        for r in resultats:
            del r['_index']

        return Response(resultats)

    @action(detail=False, methods=['get'])
    def rapport_mensuel(self, request):
        mois = request.query_params.get('mois')
        if not mois: return Response({'detail': 'Mois requis'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            date_debut = datetime.strptime(f"{mois}-01", '%Y-%m-%d')
            if date_debut.month == 12: date_fin = date_debut.replace(year=date_debut.year + 1, month=1, day=1)
            else: date_fin = date_debut.replace(month=date_debut.month + 1, day=1)
            date_debut = timezone.make_aware(date_debut)
            date_fin = timezone.make_aware(date_fin)
        except ValueError: return Response({'detail': 'Format mois invalide'}, status=status.HTTP_400_BAD_REQUEST)
        
        data = self._get_rapport_data(date_debut, date_fin, mois)
        return Response(data)

    @action(detail=False, methods=['get'])
    def produits_annules(self, request):
        """
        Liste les produits issus de factures annulées.
        """
        date_debut = request.query_params.get('date_debut')
        date_fin = request.query_params.get('date_fin')

        queryset = FactureProduit.objects.filter(
            facture__status=Facture.Status.ANNULEE
        ).select_related('facture', 'produit', 'facture__cancelled_by').order_by('-facture__date_annulation')

        if date_debut:
            queryset = queryset.filter(facture__date_annulation__gte=date_debut)
        if date_fin:
            # Ensure date_fin includes the whole day (23:59:59)
            if len(date_fin) == 10: # YYYY-MM-DD
                from django.utils import timezone
                from datetime import datetime, time
                end_dt = datetime.combine(datetime.strptime(date_fin, '%Y-%m-%d').date(), time.max)
                end_dt = timezone.make_aware(end_dt)
                queryset = queryset.filter(facture__date_annulation__lte=end_dt)
            else:
                queryset = queryset.filter(facture__date_annulation__lte=date_fin)

        # Pagination handles large results
        paginator = self.paginator
        page = paginator.paginate_queryset(queryset, self.request, view=self)
        if page is not None:
            data = []
            for fp in page:
                notes = fp.facture.notes or ""
                motif = notes.split('Motif: ')[-1] if 'Motif: ' in notes else ""
                
                # Determine source from motif
                source = "VENTES"
                if "Annulation depuis Caisse Centrale" in notes:
                    source = "CAISSE_CENTRALE"
                elif "Modification (Reload)" in notes:
                    source = "VENTES_MODIF"

                data.append({
                    'date_annulation': fp.facture.date_annulation.strftime('%d/%m/%Y %H:%M') if fp.facture.date_annulation else "",
                    'numero_facture': fp.facture.numero_facture or f"#{fp.facture.id}",
                    'nom_produit': fp.produit.name if fp.produit else fp.produit_nom,
                    'quantite_annulee': fp.quantity,
                    'lot': fp.lot,
                    'stock_actuel': fp.produit.stock if fp.produit else 0,
                    'annule_par': fp.facture.cancelled_by.username if fp.facture.cancelled_by else "Système",
                    'motif': motif,
                    'source': source
                })
            return paginator.get_paginated_response(data)

        data = []
        for fp in queryset:
            notes = fp.facture.notes or ""
            motif = notes.split('Motif: ')[-1] if 'Motif: ' in notes else ""
            
            # Determine source from motif
            source = "VENTES"
            if "Annulation depuis Caisse Centrale" in notes:
                source = "CAISSE_CENTRALE"
            elif "Modification (Reload)" in notes:
                source = "VENTES_MODIF"

            data.append({
                'date_annulation': fp.facture.date_annulation.strftime('%d/%m/%Y %H:%M') if fp.facture.date_annulation else "",
                'numero_facture': fp.facture.numero_facture or f"#{fp.facture.id}",
                'nom_produit': fp.produit.name if fp.produit else fp.produit_nom,
                'quantite_annulee': fp.quantity,
                'lot': fp.lot,
                'stock_actuel': fp.produit.stock if fp.produit else 0,
                'annule_par': fp.facture.cancelled_by.username if fp.facture.cancelled_by else "Système",
                'motif': motif,
                'source': source
            })
        return Response(data)

    @property
    def paginator(self):
        """
        Instancie le paginateur si nécessaire et le stocke.
        """
        if not hasattr(self, '_paginator'):
            self._paginator = StandardResultsSetPagination()
        return self._paginator

    @action(detail=False, methods=['get'])
    def rapport_mensuel_pdf(self, request):
        """
        Génère le PDF du rapport mensuel avec support multi-langue.
        """
        mois = request.query_params.get('mois')
        lang = request.query_params.get('lang', 'fr')
        if not mois: return Response({'detail': 'Mois requis'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            date_debut = datetime.strptime(f"{mois}-01", '%Y-%m-%d')
            if date_debut.month == 12: date_fin = date_debut.replace(year=date_debut.year + 1, month=1, day=1)
            else: date_fin = date_debut.replace(month=date_debut.month + 1, day=1)
            date_debut = timezone.make_aware(date_debut)
            date_fin = timezone.make_aware(date_fin)
        except ValueError: return Response({'detail': 'Format mois invalide'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 1. Calculer les données
        data = self._get_rapport_data(date_debut, date_fin, mois)
        
        # 2. Traductions
        trans = {
            'fr': {
                'title': "RAPPORT MENSUEL",
                'encaissements': "Encaissements",
                'mode': "Mode",
                'montant': "Montant",
                'total': "TOTAL",
                'total_expl': "TOTAL EXPLIQUÉ",
                'credit': "Ventes à Crédit",
                'coupons': "Coupons",
                'recouvrements': "Recouvrements",
                'depots_util': "Utilisation Dépôts",
                'mouvements': "Mouvements Caisse",
                'entrees': "Entrées",
                'sorties': "Sorties",
                'solde': "Solde",
                'tva_anal': "Analyse TVA",
                'taux': "Taux",
                'ht': "HT",
                'tva': "TVA",
                'ttc': "TTC",
                'top_fourn': "Top 3 Fournisseurs",
                'fourn': "Fournisseur",
                'cmd': "Cmd",
                'clients_pro': "Clients Professionnels",
                'ca_pro': "CA Pro",
                'paye': "Payé",
                'reste': "Reste",
                'taux_rec': "Taux",
                'insurance': "Assurances / Tiers-Payant",
                'grand_total': "TOTAL GÉNÉRAL",
                'subtotal_enc': "Sous-total Encaissements",
                'months': ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            },
            'en': {
                'title': "MONTHLY REPORT",
                'encaissements': "Cash In",
                'mode': "Method",
                'montant': "Amount",
                'total': "TOTAL",
                'total_expl': "TOTAL EXPLAINED",
                'credit': "Credit Sales",
                'coupons': "Coupons",
                'recouvrements': "Recoveries",
                'insurance': "Insurance / Third-party",
                'grand_total': "GRAND TOTAL",
                'subtotal_enc': "Total Encashments",
                'depots_util': "Deposit Usage",
                'mouvements': "Cash Movements",
                'entrees': "Cash In",
                'sorties': "Cash Out",
                'solde': "Balance",
                'tva_anal': "VAT Analysis",
                'taux': "Rate",
                'ht': "EX-VAT",
                'tva': "VAT",
                'ttc': "INC-VAT",
                'top_fourn': "Top 3 Suppliers",
                'fourn': "Supplier",
                'cmd': "Ord",
                'clients_pro': "Professional Clients",
                'ca_pro': "Pro Revenue",
                'paye': "Paid",
                'reste': "Balance",
                'taux_rec': "Rate",
                'months': ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
            }
        }
        t = trans.get(lang, trans['fr'])
        
        # Nom du mois localisé
        mois_dt = datetime.strptime(mois, '%Y-%m')
        nom_mois = t['months'][mois_dt.month - 1]
        titre_periode = f"{nom_mois} {mois_dt.year}".upper()

        # 3. Générer le PDF
        from django.http import HttpResponse
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm, mm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from io import BytesIO
        from api.pdf_utils import (
            get_pharma_styles, draw_pharma_header, draw_pharma_footer,
            format_currency, PharmaColors
        )
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=15, leftMargin=15,
            topMargin=80, bottomMargin=40
        )
        
        story = []
        styles = get_pharma_styles()
        compact_title = ParagraphStyle(
            'CompactTitle', fontSize=9, fontName='Helvetica-Bold',
            textColor=PharmaColors.GREEN, spaceAfter=3, spaceBefore=6,
        )
        
        def compact_table_style():
            return [
                ('BACKGROUND', (0, 0), (-1, 0), PharmaColors.GREEN),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, PharmaColors.GRAY_LIGHT),
            ]
        
        # Titre principal
        story.append(Paragraph(
            f"<b>{t['title']} - {titre_periode}</b>",
            ParagraphStyle('MainTitle', fontSize=12, fontName='Helvetica-Bold', 
                          textColor=PharmaColors.GREEN, alignment=TA_CENTER, spaceAfter=8)
        ))
        
        # === SECTION 1: KPIs PRINCIPAUX ===
        ca_ttc_label = "CA TTC" if lang == 'fr' else "TOTAL REV"
        ca_ht_label = "CA HT" if lang == 'fr' else "NET REV"
        marge_label = "Marge" if lang == 'fr' else "Margin"
        creances_label = "Créances" if lang == 'fr' else "Receivables"
        remises_label = "Remises" if lang == 'fr' else "Discounts"
        
        kpi_data = [[
            f"{ca_ttc_label}\n{format_currency(data['ca']['ca_ttc'])}",
            f"{ca_ht_label}\n{format_currency(data['ca']['ca_ht'])}",
            f"{marge_label} ({data['marge']['marge_pct']}%)\n{format_currency(data['marge']['marge_brute'])}",
            f"{remises_label}\n{format_currency(data['ca']['total_remises'])}",
            f"{creances_label}\n{format_currency(data['creances']['total'])}"
        ]]
        t_kpi = Table(kpi_data, colWidths=[3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm])
        t_kpi.setStyle([
            ('BACKGROUND', (0, 0), (-1, -1), PharmaColors.GREEN_LIGHT),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 1, PharmaColors.GREEN),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, PharmaColors.GREEN),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ])
        story.append(t_kpi)
        story.append(Spacer(1, 6))
        
        # Encaissements
        story.append(Paragraph(t['encaissements'], compact_title))
        enc_rows = [[t['mode'], t['montant']]]
        for enc in data['encaissements']:
            enc_rows.append([enc['mode_label'], format_currency(enc['montant'])])
        
        # Ajouter les dépôts avant le total
        if data.get('depots_total', 0) > 0:
            enc_rows.append([t['depots_util'], format_currency(data['depots_total'])])

        # Sous-total des encaissements (Cash + Dépôts)
        total_enc = sum((Decimal(str(e['montant'])) for e in data['encaissements']), Decimal('0.00')) + Decimal(str(data.get('depots_total', 0)))
        enc_rows.append([t['subtotal_enc'], format_currency(total_enc)])
        
        t_enc = Table(enc_rows, colWidths=[5*cm, 3*cm])
        t_enc.setStyle(compact_table_style())
        story.append(t_enc)
        
        if data.get('ventes_credit', 0) > 0 or data.get('coupons_total', 0) > 0 or data.get('recouvrements_total', 0) > 0:
            extra_rows = []
            
            # Titre pour les éléments hors-caisse
            titre_hors_caisse = "HORS ENCAISSEMENT" if lang == 'fr' else "NON-CASH ITEMS"
            extra_rows.append([titre_hors_caisse, ""])

            if data.get('ventes_credit', 0) > 0:
                extra_rows.append([t['credit'], format_currency(data['ventes_credit'])])
            if data.get('coupons_total', 0) > 0:
                extra_rows.append([t['coupons'], format_currency(data['coupons_total'])])
            if data['ca'].get('part_assurance', 0) > 0:
                extra_rows.append([t['insurance'], format_currency(data['ca']['part_assurance'])])
            if data.get('recouvrements_total', 0) > 0:
                extra_rows.append([t['recouvrements'], format_currency(data['recouvrements_total'])])
            
            # TOTAL GENERAL (CA)
            extra_rows.append([t['grand_total'], format_currency(data['ca']['ca_ttc'])])
            
            t_extra = Table(extra_rows, colWidths=[5*cm, 3*cm])
            style = compact_table_style()
            # Mettre en gras la dernière ligne (Total Général)
            style.append(('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'))
            style.append(('FONTSIZE', (0, -1), (-1, -1), 9))
            style.append(('BACKGROUND', (0, -1), (-1, -1), PharmaColors.GRAY_LIGHT))
            t_extra.setStyle(style)
            story.append(Spacer(1, 2))
            story.append(t_extra)

        story.append(Spacer(1, 4))
        
        # Mouvements Caisse
        if data['mouvements_caisse']['total_entrees'] > 0 or data['mouvements_caisse']['total_sorties'] > 0:
            story.append(Paragraph(t['mouvements'], compact_title))
            mvt_data = [
                [t['entrees'], format_currency(data['mouvements_caisse']['total_entrees'])],
                [t['sorties'], format_currency(data['mouvements_caisse']['total_sorties'])],
                [t['solde'], format_currency(data['mouvements_caisse']['solde'])]
            ]
            t_mvt = Table(mvt_data, colWidths=[5*cm, 3*cm])
            t_mvt.setStyle(compact_table_style())
            story.append(t_mvt)
            story.append(Spacer(1, 4))
        
        # TVA
        story.append(Paragraph(t['tva_anal'], compact_title))
        tva_rows = [[t['taux'], t['ht'], t['tva'], t['ttc']]]
        for tt in data['ca_par_tva']:
            tva_rows.append([
                f"{tt['taux']}%",
                format_currency(tt['ca_ht']),
                format_currency(tt['montant_tva']),
                format_currency(tt['ca_ttc'])
            ])
        t_tva = Table(tva_rows, colWidths=[2*cm, 4*cm, 3*cm, 4*cm])
        t_tva.setStyle(compact_table_style())
        story.append(t_tva)
        story.append(Spacer(1, 4))
        
        # Fournisseurs
        if data['achats_par_fournisseur']:
            story.append(Paragraph(t['top_fourn'], compact_title))
            achats_rows = [[t['fourn'], t['cmd'], t['montant']]]
            for a in data['achats_par_fournisseur'][:3]:
                achats_rows.append([
                    a['fournisseur_nom'][:20],
                    str(a['nb_commandes']),
                    format_currency(a['montant_total'])
                ])
            t_achats = Table(achats_rows, colWidths=[6*cm, 1.5*cm, 3.5*cm])
            t_achats.setStyle(compact_table_style())
            story.append(t_achats)
            story.append(Spacer(1, 4))
        
        # Clients Pro
        pro_data = data['clients_professionnels']
        if pro_data['ca_total'] > 0 or data.get('recouvrements_total', 0) > 0:
            story.append(Paragraph(t['clients_pro'], compact_title))
            recov_label = "Recouvrements" if lang == 'fr' else "Recoveries"
            p_rows = [
                [t['ca_pro'], format_currency(pro_data['ca_total']), t['paye'], format_currency(pro_data['montant_paye'])],
                [t['reste'], format_currency(pro_data['reste_a_payer']), recov_label, format_currency(data.get('recouvrements_total', 0))]
            ]
            t_pro = Table(p_rows, colWidths=[2.5*cm, 4*cm, 2.5*cm, 4*cm])
            t_pro.setStyle(compact_table_style())
            story.append(t_pro)
        
        # Build PDF
        doc.build(
            story,
            onFirstPage=lambda c, d: [draw_pharma_header(c, d, title="RAPPORT", lang=lang), draw_pharma_footer(c, d, lang=lang)],
            onLaterPages=lambda c, d: [draw_pharma_header(c, d, title="RAPPORT", lang=lang), draw_pharma_footer(c, d, lang=lang)]
        )
        
        pdf = buffer.getvalue()
        buffer.close()
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="rapport_mensuel_{mois}.pdf"'
        response.write(pdf)
        return response

    @action(detail=False, methods=['get'])
    def stocks_morts(self, request):
        """
        Rapport des stocks dormants (Dead Stock).
        Produits avec forte valeur en stock mais sans vente depuis X mois.
        """
        try:
            min_value = Decimal(request.query_params.get('min_value', 100000))
            months = int(request.query_params.get('months', 6))
            export_format = request.query_params.get('format')
        except (ValueError, TypeError):
             return Response({'error': 'Paramètres invalides'}, status=status.HTTP_400_BAD_REQUEST)

        # Date limite : aujourd'hui - X mois
        limit_date = (timezone.now() - timedelta(days=months*30)).date()
        
        # 1. Filtre de base : Stock positif
        produits = Produit.objects.filter(stock__gt=0).select_related('rayon', 'fournisseur')
        
        results = []
        for p in produits:
            # Calcul valeur stock
            valeur = (p.pmp or Decimal(0)) * p.stock
            
            # Filtre valeur min
            if valeur < min_value:
                continue
                
            # Filtre inactivité
            # Si jamais vendu (None) ou vendu avant la date limite
            dernier_vente = p.dernier_vente
            
            is_dead = False
            if not dernier_vente:
                is_dead = True
            else:
                if dernier_vente < limit_date:
                    is_dead = True
                    
            if is_dead:
                results.append({
                    'id': p.id,
                    'name': p.name,
                    'cip': p.cip1,
                    'stock': p.stock,
                    'valeur': valeur,
                    'pmp': p.pmp,
                    'dernier_vente': dernier_vente,
                    'rayon': p.rayon.name if p.rayon else '',
                    'fournisseur': p.fournisseur.name if p.fournisseur else ''
                })
        
        # Tri par valeur décroissante
        results.sort(key=lambda x: x['valeur'], reverse=True)
        
        if export_format == 'csv':
            import csv
            from django.http import HttpResponse
            
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="stocks_morts_{timezone.now().date()}.csv"'
            
            # En-tête UTF-8 BOM pour Excel
            response.write(u'\ufeff'.encode('utf8'))
            
            writer = csv.writer(response, delimiter=';')
            writer.writerow([
                'Produit', 'CIP', 'Rayon', 'Fournisseur', 
                'Stock', 'PMP', 'Valeur Stock', 'Dernière Vente'
            ])

            for r in results:
                writer.writerow([
                    r['name'],
                    r['cip'],
                    r['rayon'],
                    r['fournisseur'],
                    str(r['stock']).replace('.', ','),
                    str(r['pmp']).replace('.', ','),
                    str(r['valeur']).replace('.', ','),
                    r['dernier_vente'].strftime('%d/%m/%Y') if r['dernier_vente'] else 'Jamais'
                ])

            return response

        return Response(results)

    @action(detail=False, methods=['get'])
    def stats_vendeurs(self, request):
        """
        Statistiques de ventes par vendeur (hors caissiers).
        Retourne : Nom vendeur, Nombre de ventes, Chiffre d'affaires.
        """
        date_debut_str = request.query_params.get('date_debut')
        date_fin_str = request.query_params.get('date_fin')

        if not date_debut_str or not date_fin_str:
            return Response(
                {'error': 'Les paramètres date_debut et date_fin sont requis.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Gérer le format datetime complet (ISO 8601)
            # Utiliser timezone.make_aware pour être cohérent avec le stockage DB
            d_debut = datetime.fromisoformat(date_debut_str.replace('Z', '+00:00'))
            d_fin = datetime.fromisoformat(date_fin_str.replace('Z', '+00:00'))
            
            if timezone.is_naive(d_debut):
                date_debut = timezone.make_aware(d_debut)
            else:
                date_debut = d_debut
                
            if timezone.is_naive(d_fin):
                date_fin = timezone.make_aware(d_fin)
            else:
                date_fin = d_fin
                
            # Si c'est une date "pure" (H=0, M=0), on s'assure d'inclure toute la journée de fin
            if date_fin.hour == 0 and date_fin.minute == 0 and date_fin.second == 0:
                date_fin = date_fin + timedelta(days=1)
                
        except ValueError:
            return Response({'error': 'Format de date invalide (ISO attendu).'}, status=status.HTTP_400_BAD_REQUEST)

        # Filtrer TOUTES les factures valides sur la période (sans exclusion pour avoir le bon total)
        factures = Facture.objects.filter(
            status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
            date__gte=date_debut,
            date__lt=date_fin
        ).select_related('created_by', 'created_by__profile')

        # Agrégation par vendeur
        stats = {}
        autres_stats = {
            'vendeur': 'Ventes Non Attribuées',
            'nbre_ventes': 0,
            'chiffre_affaires': Decimal('0.00')
        }
        
        for f in factures:
            # Identifier si C'est un vendeur valide ou "Autre"
            is_vendeur = False
            vendeur_nom = "Inconnu"
            vendeur_id = None

            if f.created_by:
                # Vérifier le rôle
                role = 'INCONNU'
                if hasattr(f.created_by, 'profile') and f.created_by.profile:
                    role = f.created_by.profile.role
                
                # Nous incluons TOUS les utilisateurs identifiés comme vendeurs, quel que soit leur rôle
                # Car un caissier peut aussi faire de la vente (e.g. Laure)
                is_vendeur = True
                vendeur_id = f.created_by.id
                vendeur_nom = f.created_by.get_full_name() or f.created_by.username
            
            # Agrégation
            if is_vendeur and vendeur_id:
                if vendeur_id not in stats:
                    stats[vendeur_id] = {
                        'vendeur': vendeur_nom,
                        'nbre_ventes': 0,
                        'chiffre_affaires': Decimal('0.00')
                    }
                stats[vendeur_id]['nbre_ventes'] += 1
                stats[vendeur_id]['chiffre_affaires'] += f.total_ttc
            else:
                # Caissiers ou factures sans créateur (Système/Import)
                autres_stats['nbre_ventes'] += 1
                autres_stats['chiffre_affaires'] += f.total_ttc

        # Conversion en liste et tri par CA décroissant pour les vendeurs
        results = list(stats.values())
        results.sort(key=lambda x: x['chiffre_affaires'], reverse=True)
        
        # Ajouter la ligne "Autres" si non vide
        if autres_stats['chiffre_affaires'] > 0 or autres_stats['nbre_ventes'] > 0:
             results.append(autres_stats)

        # Ajouter une ligne de TOTAL global
        if results:
            total_ventes = sum(r['nbre_ventes'] for r in results)
            total_ca = sum(r['chiffre_affaires'] for r in results)
            
            results.append({
                'vendeur': 'TOTAL',
                'nbre_ventes': total_ventes,
                'chiffre_affaires': total_ca
            })

        return Response(results)


    @action(detail=False, methods=['get'])
    def rapport_tva_vendus(self, request):
        """
        Rapport des produits vendus soumis à la TVA (> 0) sur une période.
        """
        date_debut_str = request.query_params.get('date_debut')
        date_fin_str = request.query_params.get('date_fin')

        if not date_debut_str or not date_fin_str:
            return Response(
                {'error': 'Les paramètres date_debut et date_fin sont requis.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Handle both date only (YYYY-MM-DD) and datetime (ISO) inputs gracefully
            d_debut = datetime.fromisoformat(date_debut_str.replace('Z', '+00:00'))
            d_fin = datetime.fromisoformat(date_fin_str.replace('Z', '+00:00'))
            
            if timezone.is_naive(d_debut):
                date_debut = timezone.make_aware(d_debut)
            else:
                date_debut = d_debut
                
            if timezone.is_naive(d_fin):
                date_fin = timezone.make_aware(d_fin)
            else:
                date_fin = d_fin
                
            # Inclusion totale du dernier jour si l'heure n'est pas précisée
            if date_fin.hour == 0 and date_fin.minute == 0 and date_fin.second == 0:
                date_fin = date_fin + timedelta(days=1)
                
        except ValueError:
            return Response({'error': 'Format de date invalide.'}, status=status.HTTP_400_BAD_REQUEST)

        # On filtre les FactureProduit
        # 1. Facture Validee ou Payee
        # 2. Date dans la plage
        # 3. Produit avec TVA > 0
        
        lignes = FactureProduit.objects.filter(
            facture__date__range=(date_debut, date_fin),
            facture__status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
            tva__gt=0 # Utiliser le taux sauvegardé sur la ligne
        ).values(
            'produit__name', 'produit__cip1', 'tva'
        ).annotate(
            total_qty=Sum('quantity'),
            total_ttc=Sum(F('quantity') * (F('selling_price') - F('discount')), output_field=DecimalField()),
        ).order_by('produit__name')
        
        data = []
        for l in lignes:
            tva_rate = l['tva'] or Decimal(0)
            ttc = l['total_ttc'] or Decimal(0)
            qty = l['total_qty'] or 0
            
            # Calcul montant TVA: TTC - HT = TTC - (TTC / (1 + rate/100))
            # = TTC * (1 - 1/(1 + rate/100))
            # = TTC * (rate/100) / (1 + rate/100)
            # = TTC * rate / (100 + rate)
            
            if tva_rate > 0:
                montant_tva = (ttc * tva_rate) / (100 + tva_rate)
            else:
                montant_tva = Decimal(0)
                
            data.append({
                'produit': l['produit__name'],
                'cip': l['produit__cip1'],
                'quantite': qty,
                'taux_tva': f"{float(tva_rate)} %",
                'total_ttc': round(ttc, 0),
                'montant_tva': round(montant_tva, 0)
            })
            
        return Response(data)
    @action(detail=False, methods=['get'])
    def export_comptable_csv(self, request):
        """
        Export détaillé des ventes pour la comptabilité (CSV).
        Inclut: Date, Facture, Client, HT, TVA, TTC, Remise, Mode Paiement.
        """
        import csv
        from django.http import HttpResponse
        
        date_debut_str = request.query_params.get('date_debut')
        date_fin_str = request.query_params.get('date_fin')
        
        if not date_debut_str or not date_fin_str:
             return Response({'error': 'Les paramètres date_debut et date_fin sont requis.'}, status=status.HTTP_400_BAD_REQUEST)
             
        try:
            d_debut = datetime.fromisoformat(date_debut_str.replace('Z', '+00:00'))
            d_fin = datetime.fromisoformat(date_fin_str.replace('Z', '+00:00'))
            
            if timezone.is_naive(d_debut):
                date_debut = timezone.make_aware(d_debut)
            else:
                date_debut = d_debut
                
            if timezone.is_naive(d_fin):
                date_fin = timezone.make_aware(d_fin)
            else:
                date_fin = d_fin
                
            if date_fin.hour == 0 and date_fin.minute == 0 and date_fin.second == 0:
                date_fin = date_fin + timedelta(days=1)
        except ValueError:
            return Response({'error': 'Format de date invalide.'}, status=status.HTTP_400_BAD_REQUEST)

        # 1. Récupérer les factures validées/payées
        factures = Facture.objects.filter(
            date__range=(date_debut, date_fin),
            status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE]
        ).select_related('client', 'created_by').prefetch_related('paiements')

        # 2. Préparer la réponse CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="export_comptable_{date_debut.date()}.csv"'
        
        # En-tête UTF-8 BOM pour Excel
        response.write(u'\ufeff'.encode('utf8'))
        
        writer = csv.writer(response, delimiter=';')
        writer.writerow([
            'Date', 'Heure', 'Facture #', 'Client', 'Status', 
            'Total HT', 'Total TVA', 'Total TTC', 'Remise', 
            'Mode de Paiement', 'Caissier'
        ])

        for f in factures:
            # Modes de paiement consolidés
            modes = list(f.paiements.filter(statut='completee').values_list('mode_paiement', flat=True).distinct())
            modes_labels = [dict(Caisse.MODES_PAIEMENT).get(m, m) for m in modes]
            modes_str = ", ".join(modes_labels)

            writer.writerow([
                f.date.strftime('%d/%m/%Y'),
                f.date.strftime('%H:%M'),
                f.numero_facture or f.id,
                f.client.name if f.client else (f.client_name_override or 'Client de passage'),
                f.get_status_display(),
                str(f.total_ht).replace('.', ','),
                str(f.total_tva).replace('.', ','),
                str(f.total_ttc).replace('.', ','),
                str(f.remise).replace('.', ','),
                modes_str,
                f.created_by.get_full_name() if f.created_by else 'Système'
            ])

        return response

    @action(detail=False, methods=['get'])
    def meilleurs_clients(self, request):
        """
        Classement des meilleurs clients par CA et nombre de ventes.
        Paramètres: date_debut, date_fin, format (csv optionnel)
        """
        import csv
        from django.http import HttpResponse
        
        date_debut_str = request.query_params.get('date_debut')
        date_fin_str = request.query_params.get('date_fin')
        export_format = request.query_params.get('format')
        
        if not date_debut_str or not date_fin_str:
            return Response(
                {'error': 'Les paramètres date_debut et date_fin sont requis.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            d_debut = datetime.fromisoformat(date_debut_str.replace('Z', '+00:00'))
            d_fin = datetime.fromisoformat(date_fin_str.replace('Z', '+00:00'))
            
            if timezone.is_naive(d_debut):
                date_debut = timezone.make_aware(d_debut)
            else:
                date_debut = d_debut
                
            if timezone.is_naive(d_fin):
                date_fin = timezone.make_aware(d_fin)
            else:
                date_fin = d_fin
                
            if date_fin.hour == 0 and date_fin.minute == 0 and date_fin.second == 0:
                date_fin = date_fin + timedelta(days=1)
        except ValueError:
            return Response({'error': 'Format de date invalide.'}, status=status.HTTP_400_BAD_REQUEST)

        # Récupérer toutes les factures validées/payées avec client
        factures = Facture.objects.filter(
            date__range=(date_debut, date_fin),
            status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
            client__isnull=False
        ).select_related('client')

        # Agrégation par client
        stats = {}
        for f in factures:
            cid = f.client.id
            if cid not in stats:
                stats[cid] = {
                    'client_id': cid,
                    'client_name': f.client.name,
                    'client_type': f.client.client_type,
                    'nb_ventes': 0,
                    'chiffre_affaires': Decimal('0.00')
                }
            stats[cid]['nb_ventes'] += 1
            stats[cid]['chiffre_affaires'] += f.total_ttc

        # Conversion en liste et tri par CA décroissant
        results = list(stats.values())
        results.sort(key=lambda x: x['chiffre_affaires'], reverse=True)
        
        # Ajouter rang et panier moyen
        for i, r in enumerate(results, 1):
            r['rang'] = i
            r['panier_moyen'] = round(r['chiffre_affaires'] / r['nb_ventes'], 0) if r['nb_ventes'] > 0 else Decimal('0')
            r['chiffre_affaires'] = float(r['chiffre_affaires'])
            r['panier_moyen'] = float(r['panier_moyen'])

        # Export CSV
        if export_format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="meilleurs_clients_{date_debut.date()}.csv"'
            
            # En-tête UTF-8 BOM pour Excel
            response.write(u'\ufeff'.encode('utf8'))
            
            writer = csv.writer(response, delimiter=';')
            writer.writerow([
                'Rang', 'Client', 'Type', 'Nb Ventes', 'Chiffre Affaires', 'Panier Moyen'
            ])

            for r in results:
                writer.writerow([
                    r['rang'],
                    r['client_name'],
                    r['client_type'],
                    r['nb_ventes'],
                    str(r['chiffre_affaires']).replace('.', ','),
                    str(r['panier_moyen']).replace('.', ',')
                ])

            return response

        return Response(results)

    @action(detail=False, methods=['get'])
    def classement_vendeurs_mensuel(self, request):
        """
        Classement des vendeurs par mois avec rang, CA, nb ventes, panier moyen.
        Params: mois (YYYY-MM), periode (mois|trimestre|annee)
        """
        from django.db.models.functions import TruncMonth
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        mois_str = request.query_params.get('mois')  # Format: YYYY-MM
        periode = request.query_params.get('periode', 'mois')  # mois, trimestre, annee
        
        now = timezone.now()
        
        # Déterminer la période
        if mois_str:
            try:
                year, month = map(int, mois_str.split('-'))
                # Utilisation de timezone.make_aware pour éviter les décalages UTC/Local
                date_debut = timezone.make_aware(datetime(year, month, 1))
                if month == 12:
                    date_fin = timezone.make_aware(datetime(year + 1, 1, 1))
                else:
                    date_fin = timezone.make_aware(datetime(year, month + 1, 1))
            except:
                return Response({'error': 'Format mois invalide (YYYY-MM attendu)'}, status=400)
        else:
            # Par défaut: mois en cours
            date_debut = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            if now.month == 12:
                date_fin = now.replace(year=now.year + 1, month=1, day=1)
            else:
                date_fin = now.replace(month=now.month + 1, day=1)
        
        # Ajustement selon la période
        if periode == 'trimestre':
            quarter = (date_debut.month - 1) // 3
            date_debut = date_debut.replace(month=quarter * 3 + 1, day=1)
            end_month = (quarter + 1) * 3 + 1
            if end_month > 12:
                date_fin = date_debut.replace(year=date_debut.year + 1, month=1, day=1)
            else:
                date_fin = date_debut.replace(month=end_month, day=1)
        elif periode == 'annee':
            date_debut = date_debut.replace(month=1, day=1)
            date_fin = date_debut.replace(year=date_debut.year + 1, month=1, day=1)
        
        # Récupérer les factures (Inclure tout, même sans créateur pour le total pharmacy)
        factures = Facture.objects.filter(
            status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
            date__gte=date_debut,
            date__lt=date_fin
        ).select_related('created_by', 'created_by__profile')
        
        # Agrégation par vendeur
        stats = {}
        autres_stats = {
            'vendeur_id': 0,
            'vendeur': 'Ventes Non Attribuées',
            'nbre_ventes': 0,
            'chiffre_affaires': Decimal('0.00')
        }
        
        for f in factures:
            if f.created_by:
                vendeur_id = f.created_by.id
                vendeur_nom = f.created_by.get_full_name() or f.created_by.username
                
                if vendeur_id not in stats:
                    stats[vendeur_id] = {
                        'vendeur_id': vendeur_id,
                        'vendeur': vendeur_nom,
                        'nbre_ventes': 0,
                        'chiffre_affaires': Decimal('0.00')
                    }
                stats[vendeur_id]['nbre_ventes'] += 1
                stats[vendeur_id]['chiffre_affaires'] += f.total_ttc
            else:
                autres_stats['nbre_ventes'] += 1
                autres_stats['chiffre_affaires'] += f.total_ttc
        
        # Conversion en liste et tri
        results = list(stats.values())
        results.sort(key=lambda x: x['chiffre_affaires'], reverse=True)
        
        # Ajouter les "Autres" si significatif
        if autres_stats['chiffre_affaires'] > 0:
            results.append(autres_stats)
        
        # Ajouter rang et panier moyen
        for i, r in enumerate(results, 1):
            r['rang'] = i
            r['panier_moyen'] = round(float(r['chiffre_affaires']) / r['nbre_ventes'], 2) if r['nbre_ventes'] > 0 else 0
            r['chiffre_affaires'] = float(r['chiffre_affaires'])
        
        # Période M-1 pour évolution
        if periode == 'mois':
            if date_debut.month == 1:
                prev_debut = date_debut.replace(year=date_debut.year - 1, month=12)
            else:
                prev_debut = date_debut.replace(month=date_debut.month - 1)
            if prev_debut.month == 12:
                prev_fin = prev_debut.replace(year=prev_debut.year + 1, month=1, day=1)
            else:
                prev_fin = prev_debut.replace(month=prev_debut.month + 1, day=1)
            
            prev_factures = Facture.objects.filter(
                status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
                date__gte=prev_debut,
                date__lt=prev_fin,
                created_by__isnull=False
            ).values('created_by').annotate(
                ca=Sum('total_ttc')
            )
            prev_ca = {p['created_by']: float(p['ca']) for p in prev_factures}
            
            for r in results:
                prev = prev_ca.get(r['vendeur_id'], 0)
                if prev > 0:
                    r['evolution'] = round(((r['chiffre_affaires'] - prev) / prev) * 100, 1)
                else:
                    r['evolution'] = None
        
        return Response({
            'periode': {
                'debut': date_debut.strftime('%Y-%m-%d'),
                'fin': date_fin.strftime('%Y-%m-%d'),
                'type': periode
            },
            'data': results
        })

    @action(detail=False, methods=['get'])
    def evolution_vendeur(self, request):
        """
        Historique mensuel d'un ou plusieurs vendeurs sur 12 mois.
        Params: vendeur_id (ID ou 'all')
        """
        vendeur_id_param = request.query_params.get('vendeur_id')
        
        if not vendeur_id_param:
            return Response({'error': 'Le paramètre vendeur_id est requis.'}, status=400)
        
        from django.contrib.auth import get_user_model
        User = get_user_model()
        now = timezone.now()
        
        # Liste des vendeurs à traiter
        vendeurs_a_traiter = []
        
        if vendeur_id_param == 'all':
             # Récupérer tous les vendeurs ayant fait au moins une vente dans les 12 derniers mois
             # On remonte 12 mois en arrière
            date_debut_global = now - timedelta(days=365)
            active_sellers_ids = Facture.objects.filter(
                status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
                date__gte=date_debut_global,
                created_by__isnull=False
            ).values_list('created_by', flat=True).distinct()
            
            vendeurs_a_traiter = User.objects.filter(id__in=active_sellers_ids)
        else:
            try:
                vid = int(vendeur_id_param)
                vendeurs_a_traiter = [User.objects.get(id=vid)]
            except:
                 return Response({'error': 'vendeur_id invalide'}, status=400)

        # Préparer les étiquettes de mois (communs à tous)
        months_labels = []
        for i in range(11, -1, -1):
            month_date = now - timedelta(days=i * 30)
            months_labels.append({
                'month_key': month_date.strftime('%Y-%m'),
                'label': date_format(month_date, "M Y"), # Ex: Fev 2026
                'date_obj': month_date
            })

        response_data = []

        for vendeur in vendeurs_a_traiter:
            historique = []
            for m in months_labels:
                year = m['date_obj'].year
                month = m['date_obj'].month
                
                date_debut = datetime(year, month, 1)
                if month == 12:
                    date_fin = datetime(year + 1, 1, 1)
                else:
                    date_fin = datetime(year, month + 1, 1)
                
                # Optimisation possible: faire une seule requête aggrégée par mois pour tous les vendeurs d'un coup
                # Mais boucle simple acceptable pour < 20 vendeurs
                agg = Facture.objects.filter(
                    status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
                    date__gte=date_debut,
                    date__lt=date_fin,
                    created_by=vendeur
                ).aggregate(
                    ca=Sum('total_ttc')
                )
                
                historique.append({
                    'mois': m['month_key'],
                    'label': m['label'],
                    'chiffre_affaires': float(agg['ca'] or 0)
                })
            
            response_data.append({
                'vendeur': vendeur.get_full_name() or vendeur.username,
                'vendeur_id': vendeur.id,
                'data': historique
            })
            
        return Response(response_data)

    @action(detail=False, methods=['get'])
    def balance_stock_excel(self, request):
        """
        Génère un état de balance des stocks (Stock Initial, Achats, Ventes, Stock Final)
        pour une période donnée au format Excel.
        Params: date_debut, date_fin, lang, exclude_zero
        """
        date_debut_param = request.query_params.get('date_debut')
        date_fin_param = request.query_params.get('date_fin')
        lang = request.query_params.get('lang', 'fr')
        exclude_zero = request.query_params.get('exclude_zero') == 'true'

        if not date_debut_param or not date_fin_param:
            return Response({'error': 'Date debut and date fin are required.'}, status=400)

        try:
            from django.utils.dateparse import parse_date
            date_debut = datetime.combine(parse_date(date_debut_param), time.min)
            date_fin = datetime.combine(parse_date(date_fin_param), time.max)
        except (ValueError, TypeError):
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

        # 1. Récupérer tous les produits actifs
        produits = Produit.objects.filter(is_active=True).only('id', 'name', 'cip1', 'stock', 'stock_reserve')
        
        # 2. OPTIMISATION : Calculer tous les stocks initiaux en une seule requête (Bulk)
        # Somme de TOUS les mouvements avant la date de début pour chaque produit
        stock_initial_bulk = MouvementStock.objects.filter(
            date__lt=date_debut
        ).values('produit_id').annotate(total=Sum('quantite'))
        stock_initial_dict = {item['produit_id']: item['total'] or 0 for item in stock_initial_bulk}

        # 3. OPTIMISATION : Calculer tous les mouvements de la période en une seule requête (Bulk)
        mouvements_periode_bulk = MouvementStock.objects.filter(
            date__range=(date_debut, date_fin)
        ).values('produit_id', 'type_mouvement').annotate(total=Sum('quantite'))
        
        # Organiser en dictionnaire {produit_id: {type: total}}
        mouvements_dict = {}
        for item in mouvements_periode_bulk:
            pid = item['produit_id']
            tm = item['type_mouvement']
            val = item['total'] or 0
            if pid not in mouvements_dict:
                mouvements_dict[pid] = {}
            mouvements_dict[pid][tm] = val

        # 4. Préparer le classeur Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Traductions des en-têtes
        headers_lang = {
            'fr': {
                'title': f"Balance des Stocks - {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}",
                'cip': "Code CIP",
                'designation': "Désignation",
                'stock_initial': "Stock Initial",
                'achats': "Achats",
                'ventes': "Ventes",
                'ajustements': "Ajustements",
                'stock_final': "Stock Final"
            },
            'en': {
                'title': f"Stock Balance - {date_debut.strftime('%d/%m/%Y')} to {date_fin.strftime('%d/%m/%Y')}",
                'cip': "CIP Code",
                'designation': "Designation",
                'stock_initial': "Initial Stock",
                'achats': "Purchases",
                'ventes': "Sales",
                'ajustements': "Adjustments",
                'stock_final': "Final Stock"
            }
        }
        
        h = headers_lang.get(lang, headers_lang['fr'])
        
        # Style des titres
        title_font = Font(name='Arial', size=14, bold=True, color="1B4F72")
        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells('A1:G1')
        ws['A1'] = h['title']
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.append([]) # Ligne vide
        
        headers = [h['cip'], h['designation'], h['stock_initial'], h['achats'], h['ventes'], h['ajustements'], h['stock_final']]
        ws.append(headers)
        
        for cell in ws[3]:
            cell.font = header_font
            cell.fill = header_fill
            
        # 5. Remplir les données
        row_idx = 4
        for produit in produits:
            # Récupérer les données pré-calculées
            stock_initial = stock_initial_dict.get(produit.id, 0)
            
            prod_mouv = mouvements_dict.get(produit.id, {})
            achats = prod_mouv.get(MouvementStock.TypeMouvement.ENTREE, 0)
            ventes = prod_mouv.get(MouvementStock.TypeMouvement.SORTIE, 0)
            
            # Ajustements (Total mouvements - Entrées - Sorties)
            ajustements = sum(val for tm, val in prod_mouv.items() if tm not in [MouvementStock.TypeMouvement.ENTREE, MouvementStock.TypeMouvement.SORTIE])
            
            # Calcul du stock final
            mouvements_totaux_periode = sum(prod_mouv.values())
            stock_final = stock_initial + mouvements_totaux_periode

            # Filtre exclusion si tout est à zéro
            if exclude_zero and stock_initial == 0 and achats == 0 and ventes == 0 and ajustements == 0 and stock_final == 0:
                continue

            ws.append([
                produit.cip1 or "",
                produit.name,
                stock_initial,
                achats,
                -ventes, # Afficher en positif pour la lecture
                ajustements,
                stock_final
            ])
            
            # Appliquer des bordures
            for cell in ws[row_idx]:
                cell.border = border
                
            row_idx += 1

        # Ajuster la largeur des colonnes
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value + 2

        # 6. Envoyer la réponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Balance_Stocks_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def rapport_remises(self, request):
        date_debut_str = request.query_params.get('date_debut')
        date_fin_str = request.query_params.get('date_fin')
        
        try:
            date_debut = timezone.make_aware(datetime.combine(datetime.strptime(date_debut_str, '%Y-%m-%d'), time.min))
            date_fin = timezone.make_aware(datetime.combine(datetime.strptime(date_fin_str, '%Y-%m-%d'), time.max))
        except (ValueError, TypeError):
            return Response({"error": "Dates invalides. Format attendu: YYYY-MM-DD"}, status=status.HTTP_400_BAD_REQUEST)

        factures = Facture.objects.filter(
            status__in=[Facture.Status.VALIDEE, Facture.Status.PAYEE],
            date__gte=date_debut,
            date__lte=date_fin
        ).select_related('validated_by')

        # 1. Stats par utilisateur (Remises Globales et Fidélité)
        stats_users = factures.values(
            'validated_by__id', 
            'validated_by__username', 
            'validated_by__first_name', 
            'validated_by__last_name'
        ).annotate(
            remise_globale=Coalesce(Sum('remise'), Decimal('0.00')),
            remise_fidelite=Coalesce(Sum('montant_fidelite'), Decimal('0.00')),
            ca_ttc=Coalesce(Sum('total_ttc'), Decimal('0.00')),
            nb_factures=Count('id')
        ).order_by('-remise_globale')

        # 2. Remises par ligne (Discount et Unités Gratuites)
        line_stats = FactureProduit.objects.filter(
            facture__in=factures
        ).values('facture__validated_by__id').annotate(
            remise_lignes=Coalesce(Sum(F('discount') * F('quantity'), output_field=DecimalField()), Decimal('0.00')),
            valeur_ug=Coalesce(Sum(F('free_quantity') * F('selling_price'), output_field=DecimalField()), Decimal('0.00'))
        )

        # 3. Fusionner les stats
        line_stats_dict = {s['facture__validated_by__id']: s for s in line_stats}
        
        results = []
        for s in stats_users:
            user_id = s['validated_by__id']
            ls = line_stats_dict.get(user_id, {})
            
            remise_lignes = ls.get('remise_lignes', Decimal('0.00'))
            valeur_ug = ls.get('valeur_ug', Decimal('0.00'))
            
            total_remise = s['remise_globale'] + s['remise_fidelite'] + remise_lignes + valeur_ug
            
            results.append({
                'user_id': user_id,
                'username': s['validated_by__username'],
                'full_name': f"{s['validated_by__first_name'] or ''} {s['validated_by__last_name'] or ''}".strip() or s['validated_by__username'],
                'nb_factures': s['nb_factures'],
                'ca_ttc': s['ca_ttc'],
                'remise_globale': s['remise_globale'],
                'remise_lignes': remise_lignes,
                'remise_fidelite': s['remise_fidelite'],
                'valeur_ug': valeur_ug,
                'total_remise': total_remise,
                'ratio_remise_pct': float(total_remise / s['ca_ttc'] * 100) if s['ca_ttc'] > 0 else 0
            })

        return Response(results)

    @action(detail=False, methods=['get'])
    def rapport_remises_excel(self, request):
        # Utilise la même logique que rapport_remises
        data = self.rapport_remises(request).data
        if isinstance(data, dict) and "error" in data:
            return Response(data, status=status.HTTP_400_BAD_REQUEST)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport des Remises"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal="center")
        right_align = Alignment(horizontal="right")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # En-têtes
        headers = [
            "Utilisateur", "Nb Factures", "CA TTC", "Remise Globale", 
            "Remise Lignes", "Remise Fidélité", "Valeur UG", "Total Remise", "% / CA"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # Données
        for item in data:
            row = [
                item['full_name'],
                item['nb_factures'],
                item['ca_ttc'],
                item['remise_globale'],
                item['remise_lignes'],
                item['remise_fidelite'],
                item['valeur_ug'],
                item['total_remise'],
                f"{item['ratio_remise_pct']:.2f}%"
            ]
            ws.append(row)
            # Appliquer les formats et bordures
            curr_row = ws.max_row
            for i, val in enumerate(row):
                cell = ws.cell(row=curr_row, column=i+1)
                cell.border = border
                if i >= 2 and i <= 7: # Colonnes monétaires
                    cell.number_format = '#,##0.00'
                    cell.alignment = right_align

        # Ajuster les colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Rapport_Remises_{request.query_params.get("date_debut")}.xlsx"'
        wb.save(response)
        return response
